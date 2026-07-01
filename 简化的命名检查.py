#!/usr/bin/env python3
"""
简化的物料命名规范检查
使用已有的数据进行检查
"""

import os
import re
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 规范规则
VAGUE_NAMES = {
    '断路器', '互感器', '开关', '灯具', '线缆', '继电器', '接触器',
    '控制器', '仪表', '电缆', '电线', '按钮',
}

PRECISE_EXAMPLES = {
    '塑壳断路器', '微型断路器', '小型断路器', '真空断路器', '万能式框架断路器',
    '框架断路器', '漏电断路器', '电动机断路器',
    '电流互感器', '电压互感器', '开口式电流互感器',
    '交流接触器', '直流接触器',
    '热继电器', '中间继电器', '时间继电器',
    '指示灯', '转换开关', '隔离开关', '负荷开关', '刀开关',
    '浪涌保护器', '电涌保护器',
    '无功补偿控制器', '滤波控制器',
}

NAME_FORBIDDEN_PATTERNS = [
    (r'\d+\.?\d*\s*(?:A|V|kW|W|kV|mm|cm|m\b|kg|吨|匹|HP|Hz)', '名称含规格参数'),
    (r'[\d]+\s*[×xX*]\s*[\d]', '名称含尺寸'),
    (r'[红黄绿蓝黑白灰橙紫]', '名称含颜色'),
    (r'[（(]\s*[春夏秋冬南北东西上下左右内外]\s*[)）]', '名称含方向/位置'),
]

DESC_FORBIDDEN_CHARS = re.compile(r'[★☆●○◆◇■□▲△▼▽※§※✓✔★☆♪♫♬@#$%^&_`~]')

def check_naming_issues(materials):
    """检查命名规范问题"""
    all_issues = []
    
    for mat in materials:
        issues = []
        code = mat.get('code', '')
        name = mat.get('name', '')
        desc = mat.get('desc', '')
        
        # 检查名称中是否包含模糊词汇
        for vague_name in VAGUE_NAMES:
            if vague_name in name and vague_name not in PRECISE_EXAMPLES:
                issues.append(f"名称含模糊词: {vague_name}")
        
        # 检查名称中是否包含禁止模式
        for pattern, desc in NAME_FORBIDDEN_PATTERNS:
            if re.search(pattern, name):
                issues.append(desc)
        
        # 检查描述中是否有非标符号
        if DESC_FORBIDDEN_CHARS.search(desc):
            issues.append("描述含非标符号")
        
        # 检查乘号规范
        if re.search(r'\d\s*[×xX]\s*\d', name) and '×' in name:
            issues.append("名称应使用 x 而非 ×")
        
        if issues:
            all_issues.append({
                'code': code,
                'name': name,
                'desc': desc,
                'issues': issues
            })
    
    return all_issues

def export_excel(issues, output_file, date_str):
    """导出Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"命名规范检查-{date_str}"
    
    # 表头
    headers = ['物料编号', '物料名称', '物料描述', '问题类型']
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
    
    # 数据
    for idx, mat in enumerate(issues, 2):
        ws.cell(row=idx, column=1, value=mat['code'])
        ws.cell(row=idx, column=2, value=mat['name'])
        ws.cell(row=idx, column=3, value=mat['desc'])
        ws.cell(row=idx, column=4, value='; '.join(mat['issues']))
    
    wb.save(output_file)
    return output_file

def load_materials_from_latest():
    """从最新的物料数据文件加载"""
    data_dir = "IN3数据"
    
    # 找到最新的物料主数据文件
    files = []
    for f in os.listdir(data_dir):
        if f.startswith('物料主数据导出结果') and f.endswith('.xlsx'):
            files.append(f)
    
    if not files:
        print("未找到物料数据文件")
        return []
    
    # 按修改时间排序
    files.sort(key=lambda x: os.path.getmtime(os.path.join(data_dir, x)), reverse=True)
    latest_file = os.path.join(data_dir, files[0])
    print(f"使用数据文件: {latest_file}")
    
    # 加载Excel
    wb = openpyxl.load_workbook(latest_file, data_only=True)
    ws = wb['物料主数据']
    
    materials = []
    for r in range(2, ws.max_row + 1):
        mat = {
            'code': str(ws.cell(row=r, column=2).value or '').strip(),
            'name': str(ws.cell(row=r, column=6).value or '').strip(),
            'desc': str(ws.cell(row=r, column=7).value or '').strip(),
        }
        materials.append(mat)
    
    wb.close()
    return materials

def main():
    today = date.today().strftime('%Y-%m-%d')
    print(f"检查物料命名规范 - {today}")
    
    # 加载物料数据
    materials = load_materials_from_latest()
    print(f"加载了 {len(materials)} 条物料")
    
    # 检查命名问题
    issues = check_naming_issues(materials)
    print(f"发现 {len(issues)} 个物料有命名问题")
    
    if issues:
        # 导出结果
        output_file = f"IN3数据/命名规范检查-{today.replace('-','')}.xlsx"
        export_excel(issues, output_file, today.replace('-',''))
        print(f"结果已保存到: {output_file}")
    else:
        print("✅ 无命名问题")

if __name__ == '__main__':
    main()