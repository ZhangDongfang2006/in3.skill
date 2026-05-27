#!/usr/bin/env python3
"""采购订单全量数据分析 - 生成多Sheet Excel报告"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

INPUT = 'IN3数据/采购订单明细-全量-20260527.xlsx'
OUTPUT = 'IN3数据/采购订单分析报告-全量-20260527.xlsx'

print("读取数据...")
df = pd.read_excel(INPUT, header=0)
print(f"原始行数: {len(df)}, 列数: {len(df.columns)}")

# 清理列名前后空格
df.columns = df.columns.str.strip()

# 关键列重命名方便使用
col_map = {
    '采购订单号': 'order_no', '行号': 'line_no', '供应商代码': 'supplier_code',
    '供应商名称': 'supplier_name', '物料编号': 'mat_code', '物料名称': 'mat_name',
    '物料描述': 'mat_desc', '物料类别': 'mat_category', '物料子类别': 'mat_subcat',
    '采购单位': 'unit', '制造商': 'manufacturer', '采购数量': 'qty',
    '未税单价': 'price_ex_tax', '税率（%）': 'tax_rate', '含税单价': 'price_inc_tax',
    '未税总额': 'total_ex_tax', '价税合计': 'total_inc_tax', '状态': 'status',
    '创建时间': 'create_time', '下单时间': 'order_time', '采购员': 'buyer',
    '订单类型': 'order_type', '采购类型': 'purchase_type',
    '订单含税总金额': 'order_total_tax', '要求交期': 'req_delivery',
}
df.rename(columns=col_map, inplace=True)

# 数值列转numeric
for c in ['price_inc_tax', 'price_ex_tax', 'total_inc_tax', 'total_ex_tax', 'qty', 'tax_rate']:
    df[c] = pd.to_numeric(df[c], errors='coerce')

# 时间列
df['create_dt'] = pd.to_datetime(df['create_time'], errors='coerce')
df['order_dt'] = pd.to_datetime(df['order_time'], errors='coerce')
df['create_month'] = df['create_dt'].dt.to_period('M').astype(str)

# 过滤掉金额为0或空的行（非实质采购）
df_valid = df[df['total_inc_tax'].notna() & (df['total_inc_tax'] > 0)].copy()
print(f"有效行数（金额>0）: {len(df_valid)}")

total_amount = df_valid['total_inc_tax'].sum()
print(f"含税总金额: {total_amount:,.2f}")

# ========== 分析1: 供应商价格对比 ==========
print("\n分析1: 同一物料不同供应商价格对比...")
# 按物料+供应商聚合
mat_sup = df_valid.groupby(['mat_code', 'mat_name', 'supplier_name']).agg(
    avg_price=('price_inc_tax', 'mean'),
    min_price=('price_inc_tax', 'min'),
    max_price=('price_inc_tax', 'max'),
    total_amount=('total_inc_tax', 'sum'),
    total_qty=('qty', 'sum'),
    count=('total_inc_tax', 'count')
).reset_index()

# 找出有多个供应商的物料
sup_count = mat_sup.groupby('mat_code')['supplier_name'].nunique().reset_index()
sup_count.columns = ['mat_code', 'supplier_cnt']
multi_sup = sup_count[sup_count['supplier_cnt'] >= 2]['mat_code']
mat_multi = mat_sup[mat_sup['mat_code'].isin(multi_sup)].copy()

# 计算每个物料的价格差异
price_diff = mat_multi.groupby('mat_code').agg(
    price_min=('avg_price', 'min'),
    price_max=('avg_price', 'max'),
    suppliers=('supplier_name', lambda x: ' | '.join(x)),
    supplier_cnt=('supplier_name', 'nunique'),
    mat_name=('mat_name', 'first'),
).reset_index()
price_diff['price_range'] = price_diff['price_max'] - price_diff['price_min']
price_diff['diff_pct'] = np.where(
    price_diff['price_min'] > 0,
    (price_diff['price_max'] - price_diff['price_min']) / price_diff['price_min'] * 100,
    0
)
price_diff = price_diff.sort_values('diff_pct', ascending=False)

# 展开明细：每个物料的各供应商价格
detail_rows = []
for _, row in price_diff.head(500).iterrows():  # Top 500
    sub = mat_multi[mat_multi['mat_code'] == row['mat_code']].sort_values('avg_price')
    for _, s in sub.iterrows():
        detail_rows.append({
            '物料编号': row['mat_code'],
            '物料名称': row['mat_name'],
            '供应商数量': row['supplier_cnt'],
            '最低价': row['price_min'],
            '最高价': row['price_max'],
            '价差%': round(row['diff_pct'], 1),
            '供应商': s['supplier_name'],
            '供应商均价': round(s['avg_price'], 2),
            '采购次数': s['count'],
            '采购总量': s['total_qty'],
            '采购总额': round(s['total_amount'], 2),
        })

df_price_compare = pd.DataFrame(detail_rows)
print(f"  多供应商物料: {len(price_diff)}, 价差>30%: {(price_diff['diff_pct']>30).sum()}")

# ========== 分析2: 重复采购统计 ==========
print("分析2: 重复采购统计...")
repeat = df_valid.groupby(['mat_code', 'mat_name', 'mat_category']).agg(
    purchase_count=('total_inc_tax', 'count'),
    total_amount=('total_inc_tax', 'sum'),
    total_qty=('qty', 'sum'),
    avg_price=('price_inc_tax', 'mean'),
    supplier_cnt=('supplier_name', 'nunique'),
    first_date=('create_dt', 'min'),
    last_date=('create_dt', 'max'),
).reset_index()
repeat = repeat.sort_values('total_amount', ascending=False)
repeat['avg_price'] = repeat['avg_price'].round(2)
repeat['total_amount'] = repeat['total_amount'].round(2)
repeat.columns = ['物料编号', '物料名称', '物料类别', '采购次数', '采购总额', '采购总量', '平均含税单价', '供应商数', '首次采购', '最近采购']
print(f"  物料总数: {len(repeat)}, 采购≥3次: {(repeat['采购次数']>=3).sum()}")

# ========== 分析3: 电器元件分析 ==========
print("分析3: 电器元件分析...")
df_elec = df_valid[df_valid['mat_category'].str.contains('电器元件', na=False)].copy()
print(f"  电器元件行数: {len(df_elec)}, 总额: {df_elec['total_inc_tax'].sum():,.2f}")

elec_repeat = df_elec.groupby(['mat_code', 'mat_name']).agg(
    purchase_count=('total_inc_tax', 'count'),
    total_amount=('total_inc_tax', 'sum'),
    total_qty=('qty', 'sum'),
    avg_price=('price_inc_tax', 'mean'),
    supplier_cnt=('supplier_name', 'nunique'),
    first_date=('create_dt', 'min'),
    last_date=('create_dt', 'max'),
).reset_index()
elec_repeat = elec_repeat.sort_values('total_amount', ascending=False)
elec_repeat['avg_price'] = elec_repeat['avg_price'].round(2)
elec_repeat['total_amount'] = elec_repeat['total_amount'].round(2)
elec_repeat.columns = ['物料编号', '物料名称', '采购次数', '采购总额', '采购总量', '平均含税单价', '供应商数', '首次采购', '最近采购']

# 电器元件供应商价格对比
elec_sup = df_elec.groupby(['mat_code', 'mat_name', 'supplier_name']).agg(
    avg_price=('price_inc_tax', 'mean'),
    total_amount=('total_inc_tax', 'sum'),
    count=('total_inc_tax', 'count'),
).reset_index()
elec_sup_cnt = elec_sup.groupby('mat_code')['supplier_name'].nunique()
elec_multi = elec_sup_cnt[elec_sup_cnt >= 2].index
elec_multi_detail = elec_sup[elec_sup['mat_code'].isin(elec_multi)].copy()

elec_price_diff = elec_multi_detail.groupby('mat_code').agg(
    price_min=('avg_price', 'min'),
    price_max=('avg_price', 'max'),
    mat_name=('mat_name', 'first'),
).reset_index()
elec_price_diff['diff_pct'] = np.where(
    elec_price_diff['price_min'] > 0,
    (elec_price_diff['price_max'] - elec_price_diff['price_min']) / elec_price_diff['price_min'] * 100,
    0
)
elec_price_diff = elec_price_diff.sort_values('diff_pct', ascending=False)

elec_detail_rows = []
for _, row in elec_price_diff.head(200).iterrows():
    sub = elec_multi_detail[elec_multi_detail['mat_code'] == row['mat_code']].sort_values('avg_price')
    for _, s in sub.iterrows():
        elec_detail_rows.append({
            '物料编号': row['mat_code'],
            '物料名称': row['mat_name'],
            '价差%': round(row['diff_pct'], 1),
            '供应商': s['supplier_name'],
            '均价': round(s['avg_price'], 2),
            '采购次数': s['count'],
            '总额': round(s['total_amount'], 2),
        })
df_elec_compare = pd.DataFrame(elec_detail_rows)
print(f"  电器元件多供应商物料: {len(elec_price_diff)}")

# ========== 分析4: 价格趋势 ==========
print("分析4: 价格趋势...")
freq_mats = repeat[repeat['采购次数'] >= 3]['物料编号'].tolist()
df_trend = df_valid[df_valid['mat_code'].isin(freq_mats) & df_valid['create_dt'].notna()].copy()
df_trend = df_trend.sort_values(['mat_code', 'create_dt'])

# 采样：取采购总额top 100的物料做趋势
top100_mats = repeat[repeat['物料编号'].isin(freq_mats)].head(100)['物料编号'].tolist()
df_trend_sample = df_trend[df_trend['mat_code'].isin(top100_mats)][
    ['mat_code', 'mat_name', 'create_dt', 'create_month', 'price_inc_tax', 'qty', 'total_inc_tax', 'supplier_name']
].copy()
df_trend_sample.columns = ['物料编号', '物料名称', '创建时间', '月份', '含税单价', '数量', '价税合计', '供应商']
print(f"  趋势数据行数: {len(df_trend_sample)}")

# 按物料+月汇总趋势
trend_monthly = df_trend_sample.groupby(['物料编号', '物料名称', '月份']).agg(
    avg_price=('含税单价', 'mean'),
    total_qty=('数量', 'sum'),
    total_amount=('价税合计', 'sum'),
).reset_index()
trend_monthly['avg_price'] = trend_monthly['avg_price'].round(2)

# 计算每个物料的首末价格变化
trend_change = trend_monthly.groupby(['物料编号', '物料名称']).agg(
    first_price=('avg_price', 'first'),
    last_price=('avg_price', 'last'),
    months=('月份', 'count'),
).reset_index()
trend_change['change_pct'] = np.where(
    trend_change['first_price'] > 0,
    (trend_change['last_price'] - trend_change['first_price']) / trend_change['first_price'] * 100,
    0
)
trend_change = trend_change.sort_values('change_pct', ascending=False)
trend_change.columns = ['物料编号', '物料名称', '首次均价', '最近均价', '月份点数', '价格变化%']
print(f"  有趋势的物料: {len(trend_change)}")

# ========== 分析5: 供应商排名 ==========
print("分析5: 供应商排名...")
supplier_rank = df_valid.groupby(['supplier_code', 'supplier_name']).agg(
    total_amount=('total_inc_tax', 'sum'),
    line_count=('total_inc_tax', 'count'),
    mat_count=('mat_code', 'nunique'),
    avg_unit_price=('price_inc_tax', 'mean'),
).reset_index()
supplier_rank = supplier_rank.sort_values('total_amount', ascending=False)
supplier_rank['total_amount'] = supplier_rank['total_amount'].round(2)
supplier_rank['avg_unit_price'] = supplier_rank['avg_unit_price'].round(2)
supplier_rank['amount_pct'] = (supplier_rank['total_amount'] / total_amount * 100).round(2)
supplier_rank['cum_pct'] = supplier_rank['amount_pct'].cumsum().round(2)
supplier_rank.columns = ['供应商代码', '供应商名称', '采购总额', '明细条数', '物料种数', '平均单价', '金额占比%', '累计占比%']
print(f"  供应商总数: {len(supplier_rank)}")

# ========== 生成Excel ==========
print("\n生成Excel报告...")

wb = Workbook()

# 样式
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
red_fill = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')
orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
light_blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def write_sheet(ws, data, title=None):
    """写入DataFrame到worksheet"""
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(data.columns), 1))
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color='2F5496')
        start = 3
    else:
        start = 1
    
    # 表头
    for j, col in enumerate(data.columns, 1):
        cell = ws.cell(row=start, column=j, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # 数据
    for i, (_, row) in enumerate(data.iterrows(), start + 1):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j)
            if isinstance(val, (pd.Timestamp, datetime)):
                cell.value = val
                cell.number_format = 'YYYY-MM-DD'
            elif isinstance(val, float):
                cell.value = round(val, 2)
                cell.number_format = '#,##0.00'
            else:
                cell.value = val
            cell.border = thin_border
    
    # 自动列宽
    for j, col in enumerate(data.columns, 1):
        max_len = max(len(str(col)), data.iloc[:, j-1].astype(str).str.len().max() if len(data) > 0 else 0)
        ws.column_dimensions[chr(64 + j) if j <= 26 else 'A' + chr(64 + j - 26)].width = min(max_len + 4, 35)
    
    return start

def apply_conditional(ws, df, start, price_diff_col=None, multi_sup_col=None):
    """条件格式"""
    for i in range(start + 1, start + 1 + len(df)):
        row_idx = i - start - 1
        # 高价差红色
        if price_diff_col is not None and price_diff_col < len(df.columns):
            val = df.iloc[row_idx, price_diff_col]
            try:
                if float(val) > 30:
                    for j in range(1, len(df.columns) + 1):
                        ws.cell(row=i, column=j).fill = red_fill
            except:
                pass

# ========== Sheet 1: 概览 ==========
ws1 = wb.active
ws1.title = '概览'

overview_data = pd.DataFrame({
    '指标': [
        '数据总行数', '有效行数（金额>0）', '含税总金额（元）',
        '物料种类数', '供应商数量', '采购订单数',
        '电器元件行数', '电器元件金额（元）', '电器元件金额占比',
        '多供应商物料数', '价差>30%物料数',
        '最早订单时间', '最近订单时间',
    ],
    '值': [
        f'{len(df):,}',
        f'{len(df_valid):,}',
        f'{total_amount:,.2f}',
        f'{df_valid["mat_code"].nunique():,}',
        f'{df_valid["supplier_name"].nunique():,}',
        f'{df_valid["order_no"].nunique():,}',
        f'{len(df_elec):,}',
        f'{df_elec["total_inc_tax"].sum():,.2f}',
        f'{df_elec["total_inc_tax"].sum()/total_amount*100:.1f}%',
        f'{len(price_diff):,}',
        f'{(price_diff["diff_pct"]>30).sum():,}',
        str(df_valid['create_dt'].min())[:10],
        str(df_valid['create_dt'].max())[:10],
    ]
})

write_sheet(ws1, overview_data, '采购订单全量分析概览')

# 关键发现
findings_row = len(overview_data) + 6
findings = [
    '关键发现：',
    f'1. 共有 {len(price_diff)} 种物料由多个供应商供货，其中 {(price_diff["diff_pct"]>30).sum()} 种价差超过30%',
    f'2. 最大价差物料: {price_diff.iloc[0]["mat_name"] if len(price_diff)>0 else "N/A"}，价差 {price_diff.iloc[0]["diff_pct"]:.1f}%' if len(price_diff) > 0 else '2. 无多供应商物料',
    f'3. 电器元件金额占比 {df_elec["total_inc_tax"].sum()/total_amount*100:.1f}%，是核心采购品类',
    f'4. Top 10 供应商占比: {supplier_rank.head(10)["金额占比%"].sum():.1f}%',
    f'5. 采购金额最高的物料: {repeat.iloc[0]["物料名称"]}，总额 {repeat.iloc[0]["采购总额"]:,.2f} 元' if len(repeat) > 0 else '5. N/A',
]
for i, f in enumerate(findings):
    ws1.cell(row=findings_row + i, column=1, value=f).font = Font(bold=(i==0), size=11)

# ========== Sheet 2: 供应商价格对比 ==========
if len(df_price_compare) > 0:
    ws2 = wb.create_sheet('供应商价格对比')
    start = write_sheet(ws2, df_price_compare.head(2000), '同一物料不同供应商价格对比')
    # 条件格式
    diff_col_idx = list(df_price_compare.columns).index('价差%')
    for i in range(start + 1, start + 1 + min(len(df_price_compare), 2000)):
        row_data = df_price_compare.iloc[i - start - 1]
        try:
            if float(row_data['价差%']) > 30:
                for j in range(1, len(df_price_compare.columns) + 1):
                    ws2.cell(row=i, column=j).fill = red_fill
        except:
            pass

# ========== Sheet 3: 重复采购统计 ==========
ws3 = wb.create_sheet('重复采购统计')
write_sheet(ws3, repeat.head(2000), '元器件重复采购统计（按采购金额降序）')

# ========== Sheet 4: 电器元件分析 ==========
ws4 = wb.create_sheet('电器元件分析')

# 子标题1: 电器元件采购汇总
ws4.merge_cells('A1:I1')
ws4.cell(row=1, column=1, value='电器元件采购汇总').font = Font(bold=True, size=14, color='2F5496')
start1 = 3
for j, col in enumerate(elec_repeat.columns, 1):
    cell = ws4.cell(row=start1, column=j, value=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
for i, (_, row) in enumerate(elec_repeat.head(1000).iterrows(), start1 + 1):
    for j, val in enumerate(row, 1):
        cell = ws4.cell(row=i, column=j)
        if isinstance(val, float):
            cell.value = round(val, 2)
            cell.number_format = '#,##0.00'
        else:
            cell.value = val
        cell.border = thin_border

# 子标题2: 电器元件供应商价格对比
gap_row = start1 + 1 + min(len(elec_repeat), 1000) + 2
ws4.merge_cells(start_row=gap_row, start_column=1, end_row=gap_row, end_column=9)
ws4.cell(row=gap_row, column=1, value='电器元件多供应商价格对比').font = Font(bold=True, size=14, color='2F5496')
start2 = gap_row + 2
if len(df_elec_compare) > 0:
    for j, col in enumerate(df_elec_compare.columns, 1):
        cell = ws4.cell(row=start2, column=j, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    for i, (_, row) in enumerate(df_elec_compare.head(500).iterrows(), start2 + 1):
        for j, val in enumerate(row, 1):
            cell = ws4.cell(row=i, column=j)
            if isinstance(val, float):
                cell.value = round(val, 2)
            else:
                cell.value = val
            cell.border = thin_border
        try:
            if float(row['价差%']) > 30:
                for j in range(1, len(df_elec_compare.columns) + 1):
                    ws4.cell(row=i, column=j).fill = red_fill
        except:
            pass

# ========== Sheet 5: 价格趋势 ==========
ws5 = wb.create_sheet('价格趋势')
# 汇总表
trend_summary = trend_change.head(500)
ws5.merge_cells('A1:F1')
ws5.cell(row=1, column=1, value='物料价格趋势汇总（采购≥3次，按价格变化%排序）').font = Font(bold=True, size=14, color='2F5496')
start = 3
for j, col in enumerate(trend_summary.columns, 1):
    cell = ws5.cell(row=start, column=j, value=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
for i, (_, row) in enumerate(trend_summary.iterrows(), start + 1):
    for j, val in enumerate(row, 1):
        cell = ws5.cell(row=i, column=j)
        if isinstance(val, float):
            cell.value = round(val, 2)
            cell.number_format = '#,##0.00'
        else:
            cell.value = val
        cell.border = thin_border
    try:
        v = float(row['价格变化%'])
        if v > 20:
            for j in range(1, len(trend_summary.columns) + 1):
                ws5.cell(row=i, column=j).fill = red_fill
        elif v < -20:
            for j in range(1, len(trend_summary.columns) + 1):
                ws5.cell(row=i, column=j).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    except:
        pass

# ========== Sheet 6: 供应商排名 ==========
ws6 = wb.create_sheet('供应商排名')
write_sheet(ws6, supplier_rank.head(200), '供应商采购金额排名')

# 保存
wb.save(OUTPUT)
print(f"\n✅ 报告已保存: {OUTPUT}")
print(f"   Sheet数: {len(wb.sheetnames)} - {wb.sheetnames}")
