#!/usr/bin/env python3
"""
IN3 物料重复检查工具（stable CLI version）

用法：
  # 分析最新导出的物料主数据，输出可疑重复 Excel
  python in3_dup_check.py analyze <输入文件.xlsx> [--output <输出文件.xlsx>]

  # 对比两次检查结果，生成变更报告
  python in3_dup_check.py compare <旧文件.xlsx> <新文件.xlsx> [--output <报告.txt>]

  # 自动查找：用最新的物料主数据文件进行分析
  python in3_dup_check.py auto [--dir <目录>]

示例：
  python in3_dup_check.py analyze IN3数据/物料主数据导出结果-2026-05-21.xlsx
  python in3_dup_check.py analyze IN3数据/物料主数据导出结果-2026-05-21.xlsx -o IN3数据/可疑重复物料-2026-05-21.xlsx
  python in3_dup_check.py compare IN3数据/可疑重复物料-2026-05-19-v3.xlsx IN3数据/可疑重复物料-2026-05-21-v3.xlsx
  python in3_dup_check.py auto
"""

import argparse
import os
import re
import sys
import time
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─── 常量 ───────────────────────────────────────────────────────────────────

# IN3 导出 Excel 的列映射
COL = {
    'code': 2, 'name': 6, 'desc': 7, 'cat': 9, 'subcat': 11,
    'manufacturer': 13, 'source': 14, 'lead_time': 18, 'unit': 22,
    'price': 61, 'creator': 70, 'create_date': 72,
    'modifier': 73, 'modify_date': 75, 'color': 16,
}

# 排除的物料类别
EXCLUDE_CATS = {'成品柜', '外购成套'}

# Excel 输出字段
FIELDS = [
    ('标记', 'label'), ('物料编号', 'code'), ('物料名称', 'name'),
    ('物料描述', 'desc'), ('物料类别', 'cat'), ('物料子类别', 'subcat'),
    ('制造商', 'manufacturer'), ('物料来源', 'source'), ('提前期', 'lead_time'),
    ('主计量单位', 'unit'), ('标准价格', 'price'), ('创建人', 'creator'),
    ('创建日期', 'create_date'), ('最近修改人', 'modifier'),
    ('最近修改日期', 'modify_date'), ('备注', 'remark'),
]
DIFF_FIELDS = {'name', 'desc', 'cat', 'subcat', 'manufacturer', 'source', 'lead_time', 'unit', 'price'}

# ─── 工具函数 ───────────────────────────────────────────────────────────────

def norm(s):
    """基本标准化：统一标点、大小写"""
    s = s.lower().strip()
    s = re.sub(r'\s+', ' ', s)
    s = (s.replace('（', '(').replace('）', ')').replace('，', ',')
          .replace('：', ':').replace('×', 'x').replace('—', '-')
          .replace('φ', 'Φ').replace('ø', 'Φ'))
    return s


def ultra(s):
    """深度标准化：去除所有空白和标点（保留数字间小数点）"""
    s = norm(s)
    s = re.sub(r'(?<=\d)\.(?=\d)', '<<DOT>>', s)
    s = re.sub(r'[\s\-_/\\(),;:.，：；、。{}\[\]<>]+', '', s)
    s = s.replace('<<DOT>>', '.')
    return s


def sim(a, b):
    """文本相似度"""
    return SequenceMatcher(None, norm(a), norm(b)).ratio()


def gv(ws, row, col):
    """安全读取单元格值"""
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ''


# ─── 物料加载 ───────────────────────────────────────────────────────────────

def load_materials(input_path):
    """从 IN3 导出的 Excel 加载物料数据"""
    print(f"📂 加载文件: {input_path}")
    t0 = time.time()

    if not os.path.exists(input_path):
        print(f"❌ 文件不存在: {input_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb['物料主数据']
    print(f"   Sheet: {ws.max_row} 行 × {ws.max_column} 列")

    materials = []
    for r in range(2, ws.max_row + 1):
        cat = gv(ws, r, COL['cat'])
        if cat in EXCLUDE_CATS:
            continue
        materials.append({
            'code': gv(ws, r, COL['code']),
            'name': gv(ws, r, COL['name']),
            'desc': gv(ws, r, COL['desc']),
            'cat': cat,
            'subcat': gv(ws, r, COL['subcat']),
            'manufacturer': gv(ws, r, COL['manufacturer']),
            'source': gv(ws, r, COL['source']),
            'unit': gv(ws, r, COL['unit']),
            'lead_time': gv(ws, r, COL['lead_time']),
            'price': gv(ws, r, COL['price']),
            'creator': gv(ws, r, COL['creator']),
            'create_date': gv(ws, r, COL['create_date']),
            'modifier': gv(ws, r, COL['modifier']),
            'modify_date': gv(ws, r, COL['modify_date']),
            'color': gv(ws, r, COL['color']),
        })
    wb.close()
    print(f"✅ 加载 {len(materials)} 条物料，耗时 {time.time()-t0:.1f}s")
    return materials


# ─── 非重复判断规则（42条） ────────────────────────────────────────────────

def is_nondup(a, b):
    """
    判断两个物料是否明显非重复。
    返回 (is_nondup: bool, reason: str)
    """
    da, db = a['desc'], b['desc']
    na, nb = a['name'], b['name']
    ta, tb = da + ' ' + na, db + ' ' + nb
    ma, mb = a['manufacturer'], b['manufacturer']

    # --- 规则 #0a: 工作服/服装类直接跳过不算重复 ---
    clothing_kw = ['工作服', '保安服', '厂服', '制服']
    if any(k in na for k in clothing_kw) or any(k in nb for k in clothing_kw):
        return True, '工作服/服装类不算重复'

    # --- 规则 #0b: 明显不同的工具/器具不算重复 ---
    # 油漆刷 ≠ 铲刀/毛刷，电烙铁 ≠ 熔锡炉
    tool_conflicts = [
        ('油漆刷', ['铲刀', '毛刷', '滚筒']),
        ('铲刀', ['油漆刷', '毛刷', '滚筒']),
        ('电烙铁', ['熔锡炉', '热风枪']),
        ('熔锡炉', ['电烙铁', '热风枪']),
    ]
    for tool_a, conflicts_b in tool_conflicts:
        if tool_a in na:
            for cf in conflicts_b:
                if cf in nb:
                    return True, f'不同工具: {na} ≠ {nb}'
        if tool_a in nb:
            for cf in conflicts_b:
                if cf in na:
                    return True, f'不同工具: {na} ≠ {nb}'

    # --- 规则 #0: 甲供件不算重复 ---
    for m in [a, b]:
        if '甲供' in m['cat'] or '甲供' in m['subcat']:
            return True, '甲供件不算重复'

    # --- 规则 #1: 一个有制造商一个没有 ---
    if (ma and not mb) or (not ma and mb):
        return True, '一个有制造商一个没有'

    # --- 规则 #2: 制造商不同 ---
    if ma and mb and ma != mb:
        aliases = [
            ('宁波三爱', '三爱'), ('海越电气', '海越湖北'), ('天灵', '宁波天灵'),
            ('长沙威胜', '威胜'), ('上海良信电器股份有限公司', '良信'),
            ('上海良信电器', '良信'),
        ]
        if not any((x in ma and y in mb) or (y in ma and x in mb) for x, y in aliases):
            return True, f'制造商不同: {ma} vs {mb}'

    # --- 规则 #3: 内 vs 外 ---
    iw = ['内六角', '内门', '内侧', '内部', '内层', '内牙', '内螺纹']
    ow = ['外六角', '外门', '外侧', '外部', '外层', '外牙', '外螺纹']
    for i in iw:
        for o in ow:
            if (i in ta and o in tb) or (o in ta and i in tb):
                return True, f'内vs外: {i} vs {o}'

    # --- 规则 #4: 螺丝头部形状不同 ---
    head_types = [
        '十字', '一字', '内六角', '外六角', '梅花', '半圆头', '沉头',
        '平头', '圆头', '盘头', '六角头', '方头', '圆柱头', '扁圆头',
        '大扁头', '伞头',
    ]
    ht_a = set(h for h in head_types if h in ta)
    ht_b = set(h for h in head_types if h in tb)
    if ht_a and ht_b and ht_a != ht_b:
        return True, f'螺丝头部形状不同: {ht_a} vs {ht_b}'

    # --- 规则 #5: 分闸 vs 合闸 ---
    if ('分闸' in ta and '合闸' in tb) or ('合闸' in ta and '分闸' in tb):
        return True, '分闸 vs 合闸: 不同功能'
    if ('分励' in ta and '合闸' in tb) or ('合闸' in ta and '分励' in tb):
        return True, '分励 vs 合闸: 不同功能'

    # --- 规则 #6: 静 vs 动 ---
    if ('静插件' in ta and '动插件' in tb) or ('动插件' in ta and '静插件' in tb):
        return True, '静插件 vs 动插件: 不同部件'
    if ('静触头' in ta and '动触头' in tb) or ('动触头' in ta and '静触头' in tb):
        return True, '静触头 vs 动触头: 不同部件'
    if ('静' in ta and '动' in tb and '插件' in ta) or ('动' in ta and '静' in tb and '插件' in tb):
        return True, '静vs动: 不同部件'

    # --- 规则 #7: 形状不同 ---
    shapes = [('方形', '圆形'), ('方型', '圆型'), ('U型', '对接型'),
              ('U型', '直型'), ('平型', '立式')]
    for s1, s2 in shapes:
        if (s1 in ta and s2 in tb) or (s2 in ta and s1 in tb):
            return True, f'形状不同: {s1} vs {s2}'

    # --- 规则 #8: 安装方式 ---
    if ('明装' in ta and '暗装' in tb) or ('暗装' in ta and '明装' in tb):
        return True, '明装 vs 暗装'
    if ('明装' in ta and '嵌入式' in tb) or ('嵌入式' in ta and '明装' in tb):
        return True, '明装 vs 嵌入式'
    if ('暗装' in ta and '嵌入式' in tb) or ('嵌入式' in ta and '暗装' in tb):
        return True, '暗装 vs 嵌入式'
    install_kw = ['暗装', '户外', '户外型', '户外架', '壁挂', '落地', '嵌墙', '挂墙', '明装', '嵌入式']
    inst_a = set(k for k in install_kw if k in na)
    inst_b = set(k for k in install_kw if k in nb)
    # 两边都有安装方式但不同 → 非重复
    if inst_a and inst_b and inst_a != inst_b:
        return True, f'安装方式不同: {inst_a} vs {inst_b}'
    # 一边有安装方式，另一边没有 → 非重复（如「非标箱（暗装）」vs「非标箱」）
    if inst_a != inst_b:
        # 去掉安装方式后名称一致 = 安装方式是唯一差异
        def strip_install(name, inst_set):
            for k in inst_set:
                name = name.replace(k, '').replace('（', '').replace('）', '').strip()
            return name
        base_a = strip_install(na, inst_a)
        base_b = strip_install(nb, inst_b)
        if base_a == base_b:
            if inst_a:
                return True, f'有安装方式 vs 无安装方式: {inst_a}'
            else:
                return True, f'有安装方式 vs 无安装方式: {inst_b}'

    # --- 规则 #9: 方向不同 ---
    dirs = [('左操', '右操'), ('平进平出', '平进侧出'), ('上进', '下进'),
            ('上进下出', '下进上出'), ('立式', '卧式'), ('侧进', '上进'), ('侧出', '下出')]
    for d1, d2 in dirs:
        if (d1 in ta and d2 in tb) or (d2 in ta and d1 in tb):
            return True, f'方向不同: {d1} vs {d2}'

    # --- 规则 #10: 断路器分断能力 ---
    fp = r'(?:CDM|NDM|CM|NM)\d*-?\d+([FCDHLMNS])'
    fa, fb = re.search(fp, da, re.I), re.search(fp, db, re.I)
    if fa and fb and fa.group(1).upper() != fb.group(1).upper():
        return True, f'分断能力不同: {fa.group(1)} vs {fb.group(1)}'

    # --- 规则 #11: 脱扣曲线 B/C/D ---
    ca = re.search(r'(?:^|[\s/\-+])([BCDbcd])(\d+)', da)
    cb = re.search(r'(?:^|[\s/\-+])([BCDbcd])(\d+)', db)
    if ca and cb and ca.group(1).upper() != cb.group(1).upper():
        return True, f'脱扣曲线不同: {ca.group(1)} vs {cb.group(1)}'

    # --- 规则 #12: 极数 ---
    pa = set(re.findall(r'(\d)\s*[Pp]', da))
    pb = set(re.findall(r'(\d)\s*[Pp]', db))
    if pa and pb and pa != pb:
        return True, f'极数不同: {pa} vs {pb}'

    # --- 规则 #13: 颜色 ---
    if a.get('color') and b.get('color') and a['color'] != b['color']:
        return True, f'颜色不同: {a["color"]} vs {b["color"]}'

    # --- 规则 #14: 漏电类型 ---
    if ('A型' in da and 'AC型' in db) or ('AC型' in da and 'A型' in db):
        return True, '漏电类型不同'

    # --- 规则 #15: 附件差异 ---
    atts = ['失压', '分励', '合闸', '辅助', '报警', '门框', '电磁锁', 'RS485', '通讯', '温度指示', '带电显示', '加热器']
    aa = set(x for x in atts if x in da)
    ab = set(x for x in atts if x in db)
    diff = aa ^ ab
    if diff:
        return True, f'附件不同: {",".join(sorted(diff))}'

    # --- 规则 #16: 电线类型 ---
    wires = ['BV', 'BVR', 'YJV', 'ZR-YJV', 'RVV', 'RV', 'RVB', 'NH-YJV']
    wa = set(w for w in wires if w in da)
    wb2 = set(w for w in wires if w in db)
    if wa and wb2 and wa != wb2:
        return True, f'电线类型不同: {wa} vs {wb2}'

    # --- 规则 #17: 螺纹方向 ---
    if ('正牙' in ta and '反牙' in tb) or ('反牙' in ta and '正牙' in tb):
        return True, '螺纹方向不同'

    # --- 规则 #18: 脱扣方式 ---
    if ('TMD' in da.upper() and 'MA' in db.upper()) or ('MA' in da.upper() and 'TMD' in db.upper()):
        return True, '脱扣方式不同'

    # --- 规则 #19: 脱扣单元 ---
    if ('LSI' in da and 'TMA' in db) or ('TMA' in da and 'LSI' in db):
        return True, '脱扣单元不同'

    # --- 规则 #20: 材质 ---
    mat_pairs = [('铜', '铝'), ('304', '316'), ('紫铜', '黄铜')]
    for m1, m2 in mat_pairs:
        if (m1 in ta and m2 in tb) or (m2 in ta and m1 in tb):
            return True, f'材质不同: {m1} vs {m2}'

    # --- 规则 #21-27: 数值参数不同 ---
    # 尺寸 (数字x数字)
    dims_a = re.findall(r'(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)(?:\s*x\s*(\d+(?:\.\d+)?))?', norm(da))
    dims_b = re.findall(r'(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)(?:\s*x\s*(\d+(?:\.\d+)?))?', norm(db))
    if dims_a and dims_b and dims_a != dims_b:
        return True, f'尺寸不同: {dims_a} vs {dims_b}'

    # 母线尺寸
    bus_a = re.findall(r'(?:TMY|TMR|LMY)[\-]?(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)', norm(da), re.I)
    bus_b = re.findall(r'(?:TMY|TMR|LMY)[\-]?(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)', norm(db), re.I)
    if bus_a and bus_b and bus_a != bus_b:
        return True, f'母线尺寸不同: {bus_a} vs {bus_b}'

    # 宽度
    width_a = re.findall(r'(?:宽|宽度|W)[=:\s]*(\d+(?:\.\d+)?)\s*(?:mm|MM)?', da)
    width_b = re.findall(r'(?:宽|宽度|W)[=:\s]*(\d+(?:\.\d+)?)\s*(?:mm|MM)?', da)
    if width_a and width_b and width_a != width_b:
        return True, f'宽度不同: {width_a} vs {width_b}'

    # 直径
    dia_a = re.findall(r'(?:Φ|φ|直径)[=:\s]*(\d+(?:\.\d+)?)', da, re.I)
    dia_b = re.findall(r'(?:Φ|φ|直径)[=:\s]*(\d+(?:\.\d+)?)', db, re.I)
    if dia_a and dia_b and dia_a != dia_b:
        return True, f'直径不同: {dia_a} vs {dia_b}'

    # 电流
    cur_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Aa](?:\s|$|[^a-zA-Z0-9])', da)))
    cur_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Aa](?:\s|$|[^a-zA-Z0-9])', db)))
    if cur_a and cur_b and cur_a != cur_b:
        return True, f'电流不同: {cur_a}A vs {cur_b}A'

    # 电压
    vol_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Vv](?:\s|$|[^a-zA-Z0-9])', da)))
    vol_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Vv](?:\s|$|[^a-zA-Z0-9])', db)))
    if vol_a and vol_b and vol_a != vol_b:
        return True, f'电压不同: {vol_a}V vs {vol_b}V'

    # 功率
    pow_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Kk][Ww]', da)))
    pow_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Kk][Ww]', db)))
    if pow_a and pow_b and pow_a != pow_b:
        return True, f'功率不同: {pow_a}KW vs {pow_b}KW'

    # 截面积
    area_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*(?:mm²|mm2|平方)', da)))
    area_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*(?:mm²|mm2|平方)', db)))
    if area_a and area_b and area_a != area_b:
        return True, f'截面积不同: {area_a} vs {area_b}'

    # 电线/导线: 描述中的独立数字 = 线径
    if a['subcat'] in ('导线', '电线') or b['subcat'] in ('导线', '电线'):
        nums_a = sorted(set(re.findall(r'(?:^|\s)(\d+(?:\.\d+)?)(?:\s|$)', da)))
        nums_b = sorted(set(re.findall(r'(?:^|\s)(\d+(?:\.\d+)?)(?:\s|$)', db)))
        if nums_a and nums_b and nums_a != nums_b:
            return True, f'线径/截面积不同: {nums_a} vs {nums_b}'

    # --- 规则 #28: 名称中的数字不同 ---
    name_nums_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)', a['name'])))
    name_nums_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)', b['name'])))
    if name_nums_a and name_nums_b and name_nums_a != name_nums_b:
        return True, f'名称数字不同: {a["name"]} vs {b["name"]}'

    # --- 规则 #29: 柜型不同 ---
    cabinet_models = [
        'GGD', 'GCK', 'GCS', 'MNS', 'MCS', 'KYN28', 'KYN61', 'XGN', 'HXGN',
        'DFX', 'YBM', 'ZBW', 'ZGS', 'KYN', 'XGN', 'HXGN',
    ]
    found_cab_a = set(cm for cm in cabinet_models if cm.lower() in ta.lower())
    found_cab_b = set(cm for cm in cabinet_models if cm.lower() in tb.lower())
    if found_cab_a and found_cab_b and found_cab_a != found_cab_b:
        return True, f'柜型不同: {found_cab_a} vs {found_cab_b}'

    # --- 规则 #30-34: 型号 token 比较 ---
    models_a = re.findall(r'[a-zA-Z]+\d+[a-zA-Z0-9\-/]*', da)
    models_b = re.findall(r'[a-zA-Z]+\d+[a-zA-Z0-9\-/]*', db)
    if models_a and models_b:
        main_a = max(models_a, key=len)
        main_b = max(models_b, key=len)
        ma_low = main_a.lower().replace('-', '').replace('/', '')
        mb_low = main_b.lower().replace('-', '').replace('/', '')
        if ma_low != mb_low:
            pre_a = re.match(r'^([a-zA-Z]+)', main_a)
            pre_b = re.match(r'^([a-zA-Z]+)', main_b)
            if pre_a and pre_b:
                pa_s, pb_s = pre_a.group(1).lower(), pre_b.group(1).lower()
                rest_a, rest_b = main_a[len(pre_a.group(1)):], main_b[len(pre_b.group(1)):]
                if pa_s == pb_s:
                    nums_a = re.findall(r'\d+', rest_a)
                    nums_b = re.findall(r'\d+', rest_b)
                    if nums_a != nums_b:
                        return True, f'型号数字不同: {main_a} vs {main_b}'
                    lets_a = re.findall(r'[a-zA-Z]+', rest_a)
                    lets_b = re.findall(r'[a-zA-Z]+', rest_b)
                    if lets_a and lets_b and lets_a != lets_b:
                        return True, f'型号字母不同: {main_a} vs {main_b}'
                else:
                    base_a = re.sub(r'\d+', '', pa_s)
                    base_b = re.sub(r'\d+', '', pb_s)
                    if base_a != base_b:
                        return True, f'型号系列不同: {main_a} vs {main_b}'

    return False, ''


def desc_numbers_match(a, b):
    """检查两个物料描述+名称中的数字集合是否一致"""
    nums_a = sorted(set(re.findall(r'\d+(?:\.\d+)?', a['desc'] + ' ' + a['name'])))
    nums_b = sorted(set(re.findall(r'\d+(?:\.\d+)?', b['desc'] + ' ' + b['name'])))
    return nums_a == nums_b


# ─── 配对 + 分类 ────────────────────────────────────────────────────────────

def find_pairs(materials):
    """找出候选重复配对"""
    groups = defaultdict(list)
    for m in materials:
        key = (m['cat'], m['subcat'], norm(m['name']))
        groups[key].append(m)

    desc_groups = defaultdict(list)
    for m in materials:
        key = (m['cat'], m['subcat'], ultra(m['desc']))
        desc_groups[key].append(m)

    print(f"📊 同名分组: {len(groups)} 组, 描述分组: {len(desc_groups)} 组")

    seen = set()
    pairs = []

    # 同名组内比较
    for key, grp in groups.items():
        if len(grp) < 2:
            continue
        desc_map = defaultdict(list)
        for m in grp:
            desc_map[ultra(m['desc'])].append(m)
        for dk, items in desc_map.items():
            if len(items) < 2:
                continue
            for i in range(len(items)):
                for j in range(i + 1, len(items)):
                    a, b = items[i], items[j]
                    if a['code'] == b['code']:
                        continue
                    if not desc_numbers_match(a, b):
                        continue
                    pk = tuple(sorted([a['code'], b['code']]))
                    if pk not in seen:
                        seen.add(pk)
                        pairs.append((a, b))

    # 跨名称：描述完全一致但名称不同
    for key, grp in desc_groups.items():
        if len(grp) < 2:
            continue
        for i in range(len(grp)):
            for j in range(i + 1, len(grp)):
                a, b = grp[i], grp[j]
                if a['code'] == b['code']:
                    continue
                if a['name'] == b['name']:
                    continue
                if not desc_numbers_match(a, b):
                    continue
                pk = tuple(sorted([a['code'], b['code']]))
                if pk in seen:
                    continue
                seen.add(pk)
                pairs.append((a, b))

    return pairs


def classify_pairs(pairs):
    """将配对分类为 confirmed / review / nondup"""
    confirmed = []
    review = []
    nondup_count = 0

    for a, b in pairs:
        # 器具/工具类 → 先过非重复规则，剩下的待人工确认
        if any(k in a['subcat'] or k in b['subcat'] for k in ['器具', '工具', '工器具']):
            nd, reason = is_nondup(a, b)
            if nd:
                nondup_count += 1
                continue
            review.append((a, b, '器具/工具类需人工确认'))
            continue

        # 非重复规则检查
        nd, reason = is_nondup(a, b)
        if nd:
            nondup_count += 1
            continue

        # ★ 名称不同的配对 ★
        if a['name'] != b['name']:
            norm_na = re.sub(r'[\s\-_/\\(),;:.，：；、。{}\[\]<>]+', '', a['name'].lower())
            norm_nb = re.sub(r'[\s\-_/\\(),;:.，：；、。{}\[\]<>]+', '', b['name'].lower())

            if norm_na == norm_nb:
                confirmed.append((a, b, '命名不规范（名称仅符号差异）'))
                continue

            # 中文词语交集
            na_words = set(re.findall(r'[\u4e00-\u9fff]{2,}', a['name']))
            nb_words = set(re.findall(r'[\u4e00-\u9fff]{2,}', b['name']))
            common = na_words & nb_words
            if not common and len(na_words) > 0 and len(nb_words) > 0:
                nondup_count += 1
                continue

            # 材质差异
            mat_pairs_chk = [
                ('不锈钢', '镀彩'), ('不锈钢', '镀锌'), ('不锈钢', '镀蓝锌'),
                ('镀铝锌', '镀锌'), ('紫铜', '黄铜'), ('铜', '铝'),
                ('不锈钢', '发黑'), ('白锌', '镀彩'),
            ]
            if any((m1 in a['name'] and m2 in b['name']) or (m2 in a['name'] and m1 in b['name'])
                   for m1, m2 in mat_pairs_chk):
                nondup_count += 1
                continue

            # 类型/功能差异
            type_pairs_chk = [
                ('熔芯', '底座'), ('三通', '活接'), ('三通', '直通'), ('定向', '万向'),
                ('手动', '电动'), ('燕尾', '防爆'), ('滚轮', '脚轮'),
                ('轻轨', '重轨'), ('盲板', '凸面'), ('管夹', '垫片'),
                ('车轮', '手轮'), ('驱动大头', '驱动链条'), ('驱动器', '顶升'),
                ('大门', '加热包'), ('绝缘靴', '绝缘手套'),
                ('电度表', '电能表'), ('电压', '电流'), ('两孔', '三孔'),
                ('简易版', '全功能'),
            ]
            if any((t1 in a['name'] and t2 in b['name']) or (t2 in a['name'] and t1 in b['name'])
                   for t1, t2 in type_pairs_chk):
                nondup_count += 1
                continue

            # 形状差异
            shape_pairs_chk = [('方形', '圆形'), ('方形', '管装'), ('U型', '对接')]
            if any((s1 in a['name'] and s2 in b['name']) or (s2 in a['name'] and s1 in b['name'])
                   for s1, s2 in shape_pairs_chk):
                nondup_count += 1
                continue

            review.append((a, b, '名称不同但描述一致'))
            continue

        # 名称相同，检查描述
        if ultra(a['desc']) == ultra(b['desc']):
            confirmed.append((a, b, '命名不规范（仅符号/格式差异）'))
            continue

        if a['desc'] == b['desc']:
            confirmed.append((a, b, '描述完全相同'))
            continue
        if sim(a['desc'], b['desc']) >= 0.9:
            confirmed.append((a, b, '描述高度相似'))
            continue

        review.append((a, b, '需人工确认'))

    return confirmed, review, nondup_count


# ─── Excel 输出 ─────────────────────────────────────────────────────────────

BLUE = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
ORANGE = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
RED = Font(color='FF0000', size=10)
BLACK = Font(size=10)
thin = Side(style='thin', color='D0D0D0')
BORDER = Border(top=thin, left=thin, right=thin, bottom=thin)
WRAP = Alignment(vertical='center', wrap_text=True)
CENTER = Alignment(vertical='center')


def make_sheet(wb_out, title, data):
    """生成一个 Excel sheet"""
    ws_out = wb_out.create_sheet(title)
    for ci, (fname, _) in enumerate(FIELDS, 1):
        c = ws_out.cell(row=1, column=ci, value=fname)
        c.font = Font(bold=True, size=10)
        c.alignment = CENTER
        c.border = BORDER

    ri = 2
    for idx, (a, b, reason) in enumerate(data, 1):
        for label, mat, fill in [(f'A-{idx}', a, BLUE), (f'B-{idx}', b, ORANGE)]:
            for ci, (_, key) in enumerate(FIELDS, 1):
                if key == 'label':
                    val = label
                elif key == 'remark':
                    val = reason
                else:
                    val = mat.get(key, '')
                    if isinstance(val, float) and val == int(val):
                        val = int(val)
                cell = ws_out.cell(row=ri, column=ci, value=val)
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = WRAP if key == 'desc' else CENTER
                if key in DIFF_FIELDS:
                    va = str(a.get(key, '') or '').strip()
                    vb = str(b.get(key, '') or '').strip()
                    cell.font = RED if va != vb else BLACK
                else:
                    cell.font = BLACK
            ri += 1

    widths = [8, 16, 18, 50, 12, 14, 20, 10, 8, 8, 10, 10, 14, 10, 14, 40]
    for i, w in enumerate(widths):
        ws_out.column_dimensions[get_column_letter(i + 1)].width = w
    ws_out.freeze_panes = 'C2'
    if ri > 2:
        ws_out.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}{ri - 1}"
    return ws_out


def generate_excel(confirmed, review, output_path):
    """生成结果 Excel"""
    print(f"\n📝 生成 Excel: {output_path}")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    make_sheet(wb_out, '确认重复', confirmed)
    make_sheet(wb_out, '待人工确认', review)
    wb_out.save(output_path)
    print(f"✅ 已保存")


# ─── 对比报告 ───────────────────────────────────────────────────────────────

def read_pairs_from_sheet(file_path, sheet_name):
    """从结果 Excel 中读取配对"""
    if not os.path.exists(file_path):
        print(f"❌ 文件不存在: {file_path}")
        sys.exit(1)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]

    pairs = []
    current_pair = []
    for row in range(2, ws.max_row + 1):
        mark = str(ws.cell(row=row, column=1).value or '').strip()
        code = str(ws.cell(row=row, column=2).value or '').strip()
        if mark.startswith('A-') or mark.startswith('B-'):
            if current_pair and len(current_pair) >= 2:
                pairs.append(current_pair)
                current_pair = []
            current_pair.append({
                'mark': mark, 'code': code,
                'name': str(ws.cell(row=row, column=3).value or '').strip(),
                'desc': str(ws.cell(row=row, column=4).value or '').strip(),
                'cat': str(ws.cell(row=row, column=5).value or '').strip(),
            })
        else:
            continue
    if current_pair and len(current_pair) >= 2:
        pairs.append(current_pair)
    wb.close()
    return pairs


def get_pair_key(pair):
    """物料对的唯一标识（按物料编号排序）"""
    return tuple(sorted([pair[0]['code'], pair[1]['code']]))


def compare_reports(old_path, new_path, output_path=None):
    """对比两次检查结果，生成变更报告"""
    print(f"📊 对比物料重复检查结果")
    print(f"   旧文件: {old_path}")
    print(f"   新文件: {new_path}")

    sheets_to_check = ['确认重复', '待人工确认']
    report_lines = []

    # 提取日期
    def extract_date(path):
        m = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(path))
        return m.group(1) if m else 'unknown'

    old_date = extract_date(old_path)
    new_date = extract_date(new_path)

    report_lines.append(f"📊 物料重复检查对比报告")
    report_lines.append(f"📅 {old_date} → {new_date}")
    report_lines.append("=" * 60)

    total_old = 0
    total_new = 0
    total_resolved = 0
    total_new_added = 0
    total_unresolved = 0

    for sheet_name in sheets_to_check:
        old_pairs = read_pairs_from_sheet(old_path, sheet_name)
        new_pairs = read_pairs_from_sheet(new_path, sheet_name)

        old_dict = {get_pair_key(p): p for p in old_pairs}
        new_dict = {get_pair_key(p): p for p in new_pairs}

        resolved = [(k, v) for k, v in old_dict.items() if k not in new_dict]
        new_added = [(k, v) for k, v in new_dict.items() if k not in old_dict]
        unresolved = [(k, old_dict[k], new_dict[k]) for k in old_dict if k in new_dict]

        total_old += len(old_pairs)
        total_new += len(new_pairs)
        total_resolved += len(resolved)
        total_new_added += len(new_added)
        total_unresolved += len(unresolved)

        report_lines.append(f"\n### {sheet_name} ###")
        report_lines.append(f"  上次: {len(old_pairs)} 对 | 本次: {len(new_pairs)} 对")
        report_lines.append(f"  ✅ 已解决: {len(resolved)} 对")
        if resolved:
            for k, p in resolved[:10]:
                report_lines.append(f"    - {p[0]['code']} ({p[0]['name']}) vs {p[1]['code']} ({p[1]['name']})")
            if len(resolved) > 10:
                report_lines.append(f"    ... 还有 {len(resolved) - 10} 对")
        report_lines.append(f"  🆕 新增: {len(new_added)} 对")
        if new_added:
            for k, p in new_added[:10]:
                report_lines.append(f"    - {p[0]['code']} ({p[0]['name']}) vs {p[1]['code']} ({p[1]['name']})")
            if len(new_added) > 10:
                report_lines.append(f"    ... 还有 {len(new_added) - 10} 对")
        report_lines.append(f"  ⚠️ 未解决: {len(unresolved)} 对")

    report_lines.append(f"\n{'=' * 60}")
    report_lines.append(f"📈 总计:")
    report_lines.append(f"  上次: {total_old} 对 → 本次: {total_new} 对")
    report_lines.append(f"  ✅ 已解决 {total_resolved} | 🆕 新增 {total_new_added} | ⚠️ 未解决 {total_unresolved}")

    report = '\n'.join(report_lines)
    print(report)

    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(report)
        print(f"\n📄 报告已保存: {output_path}")

    return report


# ─── 自动模式 ───────────────────────────────────────────────────────────────

def find_latest_export(data_dir='IN3数据'):
    """找到最新的物料主数据导出文件"""
    pattern = re.compile(r'物料主数据导出结果.*\.xlsx$')
    candidates = []
    for f in os.listdir(data_dir):
        if pattern.match(f) and 'v2' not in f.lower():
            path = os.path.join(data_dir, f)
            candidates.append((os.path.getmtime(path), path, f))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def find_latest_result(data_dir='IN3数据'):
    """找到最新的可疑重复物料结果文件"""
    pattern = re.compile(r'可疑重复物料-\d{4}-\d{2}-\d{2}-v3\.xlsx$')
    candidates = []
    for f in os.listdir(data_dir):
        if pattern.match(f):
            path = os.path.join(data_dir, f)
            candidates.append((os.path.getmtime(path), path, f))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


# ─── CLI 入口 ───────────────────────────────────────────────────────────────

def cmd_analyze(args):
    """分析命令"""
    materials = load_materials(args.input)
    pairs = find_pairs(materials)
    print(f"📊 候选配对: {len(pairs)} 对")

    confirmed, review, nondup_count = classify_pairs(pairs)

    print(f"\n📈 分类结果:")
    print(f"  ❌ 非重复（已排除）: {nondup_count}")
    print(f"  ✅ 确认重复: {len(confirmed)}")
    print(f"  ❓ 待人工确认: {len(review)}")

    # 自动生成输出路径
    output = args.output
    if not output:
        m = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(args.input))
        date_str = m.group(1) if m else time.strftime('%Y-%m-%d')
        output = os.path.join(os.path.dirname(args.input) or '.', f'可疑重复物料-{date_str}-v3.xlsx')
        print(f"   自动输出路径: {output}")

    generate_excel(confirmed, review, output)
    return output


def cmd_compare(args):
    """对比命令"""
    output = args.output
    if not output:
        m = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(args.new))
        date_str = m.group(1) if m else time.strftime('%Y-%m-%d')
        output = os.path.join(os.path.dirname(args.new) or '.', f'对比报告-{date_str}.txt')
    compare_reports(args.old, args.new, output)
    return output


def cmd_auto(args):
    """自动模式：找最新文件，分析+对比"""
    data_dir = args.dir or 'IN3数据'
    print(f"🤖 自动模式，数据目录: {data_dir}")

    latest_export = find_latest_export(data_dir)
    if not latest_export:
        print(f"❌ 在 {data_dir} 中找不到物料主数据导出文件")
        sys.exit(1)

    print(f"📦 最新导出文件: {os.path.basename(latest_export)}")

    # 先分析
    output = cmd_analyze(argparse.Namespace(input=latest_export, output=None))

    # 如果有上次的结果，自动对比
    prev_result = find_latest_result(data_dir)
    if prev_result and prev_result != output:
        print(f"\n--- 自动对比 ---")
        old_date = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(prev_result))
        new_date = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(output))
        report_path = os.path.join(data_dir, f'对比报告-{new_date.group(1) if new_date else "latest"}.txt')
        compare_reports(prev_result, output, report_path)
    else:
        print("\nℹ️ 没有找到上次的结果文件，跳过对比")

    return output


def main():
    parser = argparse.ArgumentParser(
        description='IN3 物料重复检查工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
子命令:
  analyze   分析物料主数据，输出可疑重复 Excel
  compare   对比两次检查结果，生成变更报告
  auto      自动找最新文件，分析并对比
        """,
    )
    sub = parser.add_subparsers(dest='command', required=True)

    # analyze
    p_analyze = sub.add_parser('analyze', help='分析物料主数据')
    p_analyze.add_argument('input', help='IN3 导出的物料主数据 Excel 文件路径')
    p_analyze.add_argument('-o', '--output', help='输出文件路径（默认自动生成）')

    # compare
    p_compare = sub.add_parser('compare', help='对比两次检查结果')
    p_compare.add_argument('old', help='上次的结果 Excel 文件')
    p_compare.add_argument('new', help='本次的结果 Excel 文件')
    p_compare.add_argument('-o', '--output', help='对比报告输出路径（默认自动生成）')

    # auto
    p_auto = sub.add_parser('auto', help='自动分析+对比')
    p_auto.add_argument('--dir', default='IN3数据', help='数据目录（默认 IN3数据）')

    args = parser.parse_args()

    if args.command == 'analyze':
        cmd_analyze(args)
    elif args.command == 'compare':
        cmd_compare(args)
    elif args.command == 'auto':
        cmd_auto(args)


if __name__ == '__main__':
    main()
