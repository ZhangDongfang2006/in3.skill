#!/usr/bin/env python3
"""
简化版每日采购价格异常检查
"""

import openpyxl
from datetime import datetime, date
from collections import defaultdict
import os

DATA_DIR = 'IN3数据'

def load_purchase_data(filepath, factory):
    """加载采购数据"""
    records = []
    
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        return records
    
    wb = openpyxl.load_workbook(filepath, read_only=False)
    ws = wb.active
    
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else '')
    
    # 查找列索引
    col_map = {}
    for i, header in enumerate(headers):
        if '采购订单号' in header:
            col_map['po'] = i
        elif '物料编号' in header:
            col_map['material_id'] = i
        elif '物料名称' in header:
            col_map['material_name'] = i
        elif '含税单价' in header:
            col_map['price'] = i
        elif '采购数量' in header:
            col_map['qty'] = i
        elif '供应商名称' in header:
            col_map['supplier'] = i
        elif '采购申请人' in header:
            col_map['buyer'] = i
        elif '下单时间' in header or '下单日期' in header:
            col_map['order_date'] = i
    
    print(f"{factory} 列映射: {col_map}")
    
    # 读取数据
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 10:
            continue
            
        def get_col(key):
            idx = col_map.get(key)
            return row[idx] if idx is not None and idx < len(row) else None
        
        po = str(get_col('po') or '')
        if not po.startswith('PO'):
            continue
            
        material_id = str(get_col('material_id') or '')
        if not material_id:
            continue
            
        order_date = get_col('order_date')
        if order_date and hasattr(order_date, 'strftime'):
            order_date_str = order_date.strftime('%Y-%m-%d')
        else:
            order_date_str = str(order_date or '')
        
        # 只选择目标日期的记录
        if order_date_str != '2026-06-18':
            continue
            
        try:
            price = float(get_col('price') or 0)
            qty = float(get_col('qty') or 0)
        except (ValueError, TypeError):
            continue
            
        record = {
            'factory': factory,
            'po': po,
            'material_id': material_id,
            'material_name': str(get_col('material_name') or ''),
            'price': price,
            'qty': qty,
            'supplier': str(get_col('supplier') or ''),
            'buyer': str(get_col('buyer') or ''),
            'order_date': order_date_str,
        }
        records.append(record)
    
    wb.close()
    return records

def analyze_price_history(records):
    """分析历史价格"""
    # 计算历史均价
    material_prices = defaultdict(list)
    for record in records:
        if record['price'] > 0:
            material_prices[record['material_id']].append(record['price'])
    
    # 计算历史均价
    historical_avg = {}
    for material_id, prices in material_prices.items():
        if len(prices) > 1:  # 至少2个记录才有历史对比
            historical_avg[material_id] = sum(prices) / len(prices)
    
    # 找今天的记录
    today_records = [r for r in records if r['order_date'] == '2026-06-18']
    print(f"今日记录数: {len(today_records)}")
    
    # 检查异常
    anomalies = []
    for record in today_records:
        material_id = record['material_id']
        today_price = record['price']
        
        if material_id in historical_avg:
            avg_price = historical_avg[material_id]
            if avg_price > 0:
                deviation = (today_price - avg_price) / avg_price * 100
                if deviation >= 20:  # 偏差20%以上且是涨价
                    anomalies.append({
                        'material_id': material_id,
                        'material_name': record['material_name'],
                        'today_price': today_price,
                        'historical_avg': avg_price,
                        'deviation': deviation,
                        'po': record['po'],
                        'supplier': record['supplier'],
                        'buyer': record['buyer'],
                        'severity': '🔴高' if deviation >= 50 else '🟡中',
                    })
    
    return anomalies, today_records, historical_avg

def main():
    target_date = date(2026, 6, 18)
    print(f"=== 每日采购价格异常检查 ===")
    print(f"检查日期: {target_date}")
    
    # 加载数据
    nb_file = os.path.join(DATA_DIR, '采购订单明细-宁波-20260618.xlsx')
    hb_file = os.path.join(DATA_DIR, '采购订单明细-湖北-20260618.xlsx')
    
    print("\n加载数据...")
    nb_records = load_purchase_data(nb_file, '宁波')
    hb_records = load_purchase_data(hb_file, '湖北')
    
    all_records = nb_records + hb_records
    print(f"\n总记录数: {len(all_records)}")
    print(f"宁波: {len(nb_records)} 条")
    print(f"湖北: {len(hb_records)} 条")
    
    if not all_records:
        print("❌ 无数据")
        return
    
    # 分析
    anomalies, today_records, historical_avg = analyze_price_history(all_records)
    
    if not anomalies:
        print(f"\n✅ {target_date} 无价格异常")
        return
    
    print(f"\n🚨 发现 {len(anomalies)} 条价格异常:")
    
    high = sum(1 for a in anomalies if '高' in a['severity'])
    mid = sum(1 for a in anomalies if '中' in a['severity'])
    print(f"异常分布: 🔴高 {high}, 🟡中 {mid}")
    
    for anomaly in anomalies:
        print(f"  {anomaly['severity']} {anomaly['po']} | {anomaly['material_name']} | ¥{anomaly['today_price']:.2f} vs 均价¥{anomaly['historical_avg']:.2f} ({anomaly['deviation']:.1f}%) | {anomaly['supplier'][:15]}")
    
    print(f"\n[SUMMARY] {target_date}: {len(anomalies)} 条异常 (🔴{high} 🟡{mid})")

if __name__ == '__main__':
    main()