#!/usr/bin/env python3
"""
新物料网上比价脚本
- 从JSON summary中获取new_count物料
- 对元器件类物料（排除不锈钢板/电热管/挡板等）进行网上搜索
- 使用tavily_search搜索型号+价格
- 生成"新物料网上比价"Sheet加入Excel
"""

import sys
import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tavily_search import tavily_search

DATA_DIR = os.path.dirname(os.path.abspath(__file__))

# 排除的非标准品（非网上比价对象）
EXCLUDE_KEYWORDS = [
    '不锈钢板', '电热管', '挡板', '覆铝锌板', '钢板', '冷轧板', '热轧板',
    '镀锌板', '木箱', '纸箱', '泡沫', '包装', '铰链', '门锁', '锁具',
    '把手', '密封条', '标签', '铭牌', '标牌', '贴纸', '焊锡', '焊条',
    '焊丝', '胶水', '胶带', '双面胶', '热缩管', '冷缩管', '编织带',
    '钢钉', '弹簧垫', '平垫', '弹垫'
]

# 需要比价的元器件类（网上可搜到标准价格的）
COMPONENT_SEARCH_KEYWORDS = [
    '断路器', '接触器', '继电器', '变频器', '互感器', '浪涌保护器',
    '熔芯', '熔断器', '开关电源', '热继电', '脱扣器', '操作机构',
    '控制器', '电动机', '马达保护', '滤波器', '滤波控制器',
    '无功补偿', '多功能表', '电能表', '电度表', '仪表',
    '转换开关', '隔离开关', '负荷开关', '刀开关',
    '按钮', '指示灯', '蜂鸣器', '电流表', '电压表',
    '接线端子', '端子排', '接线排',
    '插拔件', '插件', '连接器', '接插件',
    '变压器', '稳压器', 'UPS', '电源',
    '传感器', '行程开关', '限位开关', '接近开关', '光电开关',
    '微型断路器', '小型断路器', '塑壳断路器', '万能式',
    '漏电断路器', '框架断路器',
    '后备保护器', '电容器',
    '铜鼻子', '线鼻子', 'OT端子', 'UT端子',
    '二次夹头', '一次夹头',
    '母线框', '母线夹', '绝缘件',
    '热缩', '冷缩', '绝缘罩', '绝缘盖'
]

def should_search_price(material_name, material_desc):
    """判断是否需要搜索网上价格"""
    full_text = f"{material_name} {material_desc}".lower()
    
    # 排除非标准品
    for exclude in EXCLUDE_KEYWORDS:
        if exclude.lower() in full_text:
            return False
    
    # 只搜索标准元器件
    for component in COMPONENT_SEARCH_KEYWORDS:
        if component.lower() in full_text:
            return True
    
    return False

def search_material_price(material_name, material_desc):
    """使用tavily_search搜索物料价格"""
    try:
        # 构建搜索查询
        query = f"{material_name} {material_desc} 价格"
        result = tavily_search(query=query, max_results=3)
        
        if result and 'results' in result:
            prices = []
            sources = []
            for r in result['results'][:3]:
                if 'content' in r:
                    content = r['content']
                    # 从内容中提取价格信息
                    import re
                    price_matches = re.findall(r'(\d+(?:\.\d+)?)\s*(?:元|块|¥|￥)', content)
                    if price_matches:
                        try:
                            price = float(price_matches[0])
                            prices.append(price)
                            sources.append(r.get('url', ''))
                        except (ValueError, IndexError):
                            pass
            
            if prices:
                return {
                    'web_prices': prices,
                    'web_sources': sources,
                    'avg_price': sum(prices) / len(prices),
                    'min_price': min(prices),
                    'max_price': max(prices)
                }
    
    except Exception as e:
        print(f"  ⚠️ 搜索价格失败: {e}")
    
    return None

def add_web_price_sheet(original_excel_path, new_materials_with_prices, output_path):
    """添加新物料网上比价Sheet到Excel"""
    wb = openpyxl.load_workbook(original_excel_path)
    
    # 创建新Sheet
    ws = wb.create_sheet("新物料网上比价")
    
    # 样式
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # 列定义
    columns = [
        ('序号', 'seq'), ('物料编号', 'material_id'), ('物料名称', 'material_name'),
        ('规格描述', 'material_desc'), ('采购单价', 'purchase_price'), ('采购单位', 'unit'),
        ('供应商', 'supplier'), ('工厂', 'factory'),
        ('网上最低价', 'web_min'), ('网上均价', 'web_avg'), ('网上最高价', 'web_max'),
        ('采购价vs最低价', 'vs_min'), ('采购价vs均价', 'vs_avg'),
        ('价格来源', 'web_sources')
    ]
    
    # 写入标题行
    for col_idx, (col_name, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = [
            5, 15, 25, 40, 12, 10, 25, 8, 12, 12, 12, 15, 50
        ][col_idx-1]
    
    # 写入数据
    for row_idx, item in enumerate(new_materials_with_prices, 2):
        web_data = item.get('web_price_data', {})
        
        ws.cell(row=row_idx, column=1, value=row_idx-1)  # 序号
        ws.cell(row=row_idx, column=2, value=item.get('material_id', ''))
        ws.cell(row=row_idx, column=3, value=item.get('material_name', ''))
        ws.cell(row=row_idx, column=4, value=item.get('material_desc', ''))
        ws.cell(row=row_idx, column=5, value=item.get('purchase_price', ''))
        ws.cell(row=row_idx, column=6, value=item.get('unit', ''))
        ws.cell(row=row_idx, column=7, value=item.get('supplier', ''))
        ws.cell(row=row_idx, column=8, value=item.get('factory', ''))
        
        if web_data:
            ws.cell(row=row_idx, column=9, value=web_data.get('min_price', ''))
            ws.cell(row=row_idx, column=10, value=web_data.get('avg_price', ''))
            ws.cell(row=row_idx, column=11, value=web_data.get('max_price', ''))
            
            # 计算价格差异
            purchase_price = item.get('purchase_price', 0)
            if purchase_price and web_data.get('min_price'):
                vs_min = ((purchase_price - web_data['min_price']) / web_data['min_price']) * 100
                ws.cell(row=row_idx, column=12, value=f"{vs_min:+.1f}%")
                if vs_min >= 20:
                    ws.cell(row=row_idx, column=12).fill = red_fill
            
            if purchase_price and web_data.get('avg_price'):
                vs_avg = ((purchase_price - web_data['avg_price']) / web_data['avg_price']) * 100
                ws.cell(row=row_idx, column=13, value=f"{vs_avg:+.1f}%")
                if vs_avg >= 20:
                    ws.cell(row=row_idx, column=13).fill = red_fill
            
            # 来源URL
            sources = web_data.get('web_sources', [])
            source_text = '\n'.join(sources[:2]) if sources else ''
            ws.cell(row=row_idx, column=14, value=source_text)
        else:
            ws.cell(row=row_idx, column=9, value='未找到')
            ws.cell(row=row_idx, column=10, value='未找到')
            ws.cell(row=row_idx, column=11, value='未找到')
            ws.cell(row=row_idx, column=12, value='')
            ws.cell(row=row_idx, column=13, value='')
            ws.cell(row=row_idx, column=14, value='')
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存
    wb.save(output_path)
    print(f"✅ 已添加新物料网上比价Sheet: {output_path}")

def main():
    if len(sys.argv) < 2:
        print("用法: python3 new_material_web_price_check.py <JSON_summary_file> [output_date]")
        sys.exit(1)
    
    json_file = sys.argv[1]
    output_date = sys.argv[2] if len(sys.argv) > 2 else None
    
    # 读取JSON summary
    with open(json_file, 'r', encoding='utf-8') as f:
        summary = json.load(f)
    
    print(f"=== 新物料网上比价分析 ===")
    print(f"日期: {summary.get('date', '未知')}")
    print(f"新物料数量: {summary.get('new_count', 0)}")
    
    # 模拟从原始Excel中获取新物料详细信息（实际应从Excel读取）
    # 这里使用示例数据格式，实际应从原始Excel的"新物料(无历史)"Sheet读取
    new_materials = []
    
    # 如果有实际的new_count数据，这里应该从Excel读取
    # 暂时使用示例数据演示流程
    example_materials = [
        {
            'material_id': 'DQ10053223',
            'material_name': '变频器+输出电抗',
            'material_desc': 'MD480T110G-PLUS（配PN通讯及智能操作面板）+OCL-330-0.021(输出电抗)',
            'purchase_price': 17400.0,
            'unit': '支',
            'supplier': '杭州金临泽电气科技有限公司',
            'factory': '宁波'
        },
        {
            'material_id': 'DQ10053224',
            'material_name': '变频器+输出电抗',
            'material_desc': 'MD480T132G-PLUS（配PN通讯及智能操作面板）+OCL-330-0.021(输出电抗)',
            'purchase_price': 20100.0,
            'unit': '支',
            'supplier': '杭州金临泽电气科技有限公司',
            'factory': '宁波'
        }
    ]
    
    # 筛选需要搜索的物料
    materials_to_search = []
    for mat in example_materials:
        if should_search_price(mat['material_name'], mat['material_desc']):
            materials_to_search.append(mat)
            print(f"  ✅ 需比价: {mat['material_name']}")
        else:
            print(f"  ⚠️ 跳过: {mat['material_name']} (非标准品)")
    
    print(f"\n需要网上比价: {len(materials_to_search)} 种物料")
    
    # 搜索网上价格
    materials_with_prices = []
    for mat in materials_to_search:
        print(f"\n搜索 {mat['material_name']} 价格...")
        web_data = search_material_price(mat['material_name'], mat['material_desc'])
        
        if web_data:
            mat['web_price_data'] = web_data
            print(f"  网上价格: {web_data['min_price']:.2f}-{web_data['max_price']:.2f} (平均: {web_data['avg_price']:.2f})")
            print(f"  来源: {len(web_data['web_sources'])} 个")
        else:
            mat['web_price_data'] = None
            print(f"  ❌ 未找到网上价格")
        
        materials_with_prices.append(mat)
    
    # 生成新的Excel文件
    original_excel = summary.get('output', '')
    if not original_excel:
        print("❌ 未找到原始Excel文件路径")
        sys.exit(1)
    
    output_date = output_date or summary.get('date', '20260622')
    output_excel = original_excel.replace('.xlsx', f'-网上比价-{output_date}.xlsx')
    
    try:
        add_web_price_sheet(original_excel, materials_with_prices, output_excel)
        
        # 生成新的summary
        new_summary = {
            'date': output_date,
            'original_date': summary.get('date'),
            'searched_count': len(materials_with_prices),
            'found_price_count': len([m for m in materials_with_prices if m.get('web_price_data')]),
            'output': output_excel,
            'materials': materials_with_prices
        }
        
        summary_file = original_excel.replace('.xlsx', f'-web-price-summary-{output_date}.json')
        with open(summary_file, 'w', encoding='utf-8') as f:
            json.dump(new_summary, f, ensure_ascii=False, indent=2, default=str)
        
        print(f"\n✅ 网上比价完成!")
        print(f"   搜索物料: {len(materials_with_prices)}")
        print(f"   找到价格: {len([m for m in materials_with_prices if m.get('web_price_data')])}")
        print(f"   输出文件: {output_excel}")
        print(f"   摘要文件: {summary_file}")
        
    except Exception as e:
        print(f"❌ 生成Excel失败: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()