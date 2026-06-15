#!/usr/bin/env python3
"""Generate IN3 可疑重复物料_已验证_v4.xlsx with proper formatting."""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "/Users/zhangdongfang/.openclaw/workspace-in3bot/IN3数据"
ORIGINAL_XLSX = f"{BASE}/物料主数据导出结果-20260905161742956.xlsx"
OUTPUT = f"{BASE}/可疑重复物料_已验证_v4.xlsx"

# --- 1. Load JSON data ---
confirmed_dup = json.load(open(f"{BASE}/confirmed_dup_v3.json"))
likely_dup = json.load(open(f"{BASE}/likely_dup_v3.json"))
not_dup = json.load(open(f"{BASE}/not_dup_v3.json"))
pending = json.load(open(f"{BASE}/pending_v3.json"))
reclassified = json.load(open(f"{BASE}/pending_reclassified.json"))

print(f"confirmed_dup: {len(confirmed_dup)}, likely_dup: {len(likely_dup)}")
print(f"not_dup: {len(not_dup)}, pending: {len(pending)}")
print(f"reclassified: not_dup={len(reclassified['not_dup'])}, confirmed={len(reclassified['confirmed_dup'])}, still_pending={len(reclassified['still_pending'])}")

# --- 2. Build sheet data ---
# Sheet 1: 确认重复 = confirmed_dup + likely_dup + reclassified.confirmed_dup
sheet1_pairs = confirmed_dup + likely_dup + reclassified["confirmed_dup"]
# Sheet 2: 非重复 = not_dup + reclassified.not_dup
sheet2_pairs = not_dup + reclassified["not_dup"]
# Sheet 3: 待人工确认 = reclassified.still_pending
sheet3_pairs = reclassified["still_pending"]

print(f"Sheet1 (确认重复): {len(sheet1_pairs)} pairs")
print(f"Sheet2 (非重复): {len(sheet2_pairs)} pairs")
print(f"Sheet3 (待人工确认): {len(sheet3_pairs)} pairs")

# --- 3. Load original Excel for extra fields ---
print("Loading original Excel for extra fields...")
wb_orig = openpyxl.load_workbook(ORIGINAL_XLSX, read_only=True, data_only=True)
ws_orig = wb_orig["物料主数据"]

# Find column indices
header_row = next(ws_orig.iter_rows(min_row=1, max_row=1, values_only=True))
col_map = {}
for i, h in enumerate(header_row):
    if h:
        col_map[str(h).strip()] = i

print(f"Total columns in original: {len(header_row)}")

# Build lookup: material_id -> dict of extra fields
# Key columns we need beyond what JSON has:
# 创建人, 创建日期, 最近修改人, 最近修改日期, 提前期, 标准价格
extra_fields_map = {}
needed_cols = {
    "创建人": col_map.get("创建人"),
    "创建日期": col_map.get("创建日期"),
    "最近修改人": col_map.get("最近修改人"),
    "最近修改日期": col_map.get("最近修改日期"),
    "提前期": col_map.get("提前期"),
    "标准价格": col_map.get("标准价格"),
    "*物料编号": col_map.get("*物料编号"),
}
print(f"Column mapping: {needed_cols}")

id_col = needed_cols["*物料编号"]
if id_col is None:
    # Try to find it
    for k, v in col_map.items():
        if "物料编号" in k:
            id_col = v
            print(f"Found 物料编号 at col {v}: {k}")
            break

count = 0
for row in ws_orig.iter_rows(min_row=2, values_only=True):
    mat_id = row[id_col] if id_col is not None and id_col < len(row) else None
    if not mat_id:
        continue
    extra = {}
    for field, ci in needed_cols.items():
        if ci is not None and ci < len(row) and field != "*物料编号":
            extra[field] = row[ci]
    extra_fields_map[str(mat_id).strip()] = extra
    count += 1
    if count % 10000 == 0:
        print(f"  Loaded {count} rows...")

wb_orig.close()
print(f"Loaded {len(extra_fields_map)} material records from original Excel")

# --- 4. Define output columns ---
COLS = ["标记", "物料编号", "物料名称", "物料描述", "物料类别", "物料子类别",
        "制造商", "物料来源", "主计量单位", "提前期", "标准价格",
        "创建人", "创建日期", "最近修改人", "最近修改日期", "备注"]

# Columns where differences should be red
RED_DIFF_COLS = {"物料名称", "物料描述", "物料类别", "物料子类别",
                 "制造商", "物料来源", "主计量单位", "提前期", "标准价格"}

# --- 5. Styles ---
fill_blue = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
fill_orange = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
font_red = Font(color="FF0000")
font_normal = Font()
font_header = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

# --- 6. Helper function to build rows for a pair ---
def make_rows(pair, pair_idx):
    """Return list of 2 row dicts for a pair."""
    a = pair["a"]
    b = pair["b"]
    reason = pair.get("reason", "")
    
    rows = []
    for side, mat, label in [("A", a, f"A-{pair_idx}"), ("B", b, f"B-{pair_idx}")]:
        mid = mat["id"]
        extra = extra_fields_map.get(mid, {})
        row = {
            "标记": label,
            "物料编号": mid,
            "物料名称": mat.get("name", ""),
            "物料描述": mat.get("desc", ""),
            "物料类别": mat.get("category", ""),
            "物料子类别": mat.get("subcategory", ""),
            "制造商": mat.get("manufacturer", ""),
            "物料来源": mat.get("source", ""),
            "主计量单位": mat.get("unit", ""),
            "提前期": extra.get("提前期", ""),
            "标准价格": extra.get("标准价格", ""),
            "创建人": extra.get("创建人", ""),
            "创建日期": extra.get("创建日期", ""),
            "最近修改人": extra.get("最近修改人", ""),
            "最近修改日期": extra.get("最近修改日期", ""),
            "备注": reason if side == "A" else "",
        }
        rows.append(row)
    return rows

# --- 7. Write Excel ---
print("Writing Excel...")
wb = openpyxl.Workbook()

for sheet_idx, (sheet_name, pairs) in enumerate([
    ("确认重复", sheet1_pairs),
    ("非重复", sheet2_pairs),
    ("待人工确认", sheet3_pairs),
]):
    print(f"  Processing {sheet_name}: {len(pairs)} pairs...")
    
    if sheet_idx == 0:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    # Header
    for col_idx, col_name in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    row_num = 2
    for pair_idx, pair in enumerate(pairs, 1):
        pair_rows = make_rows(pair, pair_idx)
        for i, row_data in enumerate(pair_rows):
            is_a = (i == 0)
            fill = fill_blue if is_a else fill_orange
            
            for col_idx, col_name in enumerate(COLS, 1):
                val = row_data[col_name]
                # Handle date values
                if hasattr(val, 'isoformat'):
                    val = val.isoformat()
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = thin_border
                cell.fill = fill
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            row_num += 1
    
    # Mark differences in red - process pair by pair
    # Each pair occupies 2 rows (row_num starts at 2, pair 1 = rows 2-3, etc.)
    for pair_idx in range(len(pairs)):
        row_a = 2 + pair_idx * 2
        row_b = row_a + 1
        
        for col_idx, col_name in enumerate(COLS, 1):
            if col_name not in RED_DIFF_COLS:
                continue
            val_a = ws.cell(row=row_a, column=col_idx).value
            val_b = ws.cell(row=row_b, column=col_idx).value
            if str(val_a) != str(val_b):
                ws.cell(row=row_a, column=col_idx).font = font_red
                ws.cell(row=row_b, column=col_idx).font = font_red
    
    # Column widths
    for col_idx, col_name in enumerate(COLS, 1):
        if col_name == "物料描述":
            ws.column_dimensions[get_column_letter(col_idx)].width = 50
        elif col_name == "备注":
            ws.column_dimensions[get_column_letter(col_idx)].width = 40
        elif col_name == "标记":
            ws.column_dimensions[get_column_letter(col_idx)].width = 8
        elif col_name == "物料编号":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
        elif col_name == "物料名称":
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
    
    # Freeze panes: freeze first 2 columns + first 2 rows
    ws.freeze_panes = "C3"
    
    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{row_num - 1}"
    
    print(f"  {sheet_name}: {row_num - 2} rows written")

wb.save(OUTPUT)
print(f"\nDone! Saved to {OUTPUT}")
