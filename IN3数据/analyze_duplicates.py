#!/usr/bin/env python3
"""IN3 物料重复检查分析脚本 - 2026-05-12"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re
import time

INPUT_FILE = '物料主数据导出结果-20260512.xlsx'
OUTPUT_FILE = '可疑重复物料-2026-05-12.xlsx'

# Column mapping (1-indexed)
COL = {
    'id': 1, 'code': 2, 'name': 6, 'desc': 7,
    'cat_code': 8, 'cat': 9, 'subcat_code': 10, 'subcat': 11,
    'manufacturer': 13, 'source': 14, 'lead_time': 18,
    'unit': 22, 'price': 61,
    'creator': 70, 'create_date': 72,
    'modifier': 73, 'modify_date': 75,
    'color': 16
}

print("Loading workbook...")
start = time.time()
wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb['物料主数据']
print(f"Loaded in {time.time()-start:.1f}s, {ws.max_row} rows")

# Read all materials
materials = []
exclude_cats = {'成品柜', '外购成套'}

for row_idx in range(2, ws.max_row + 1):
    cat = ws.cell(row=row_idx, column=COL['cat']).value or ''
    subcat = ws.cell(row=row_idx, column=COL['subcat']).value or ''
    
    # Skip excluded categories
    if cat in exclude_cats:
        continue
    
    code = ws.cell(row=row_idx, column=COL['code']).value or ''
    name = ws.cell(row=row_idx, column=COL['name']).value or ''
    desc = ws.cell(row=row_idx, column=COL['desc']).value or ''
    manufacturer = ws.cell(row=row_idx, column=COL['manufacturer']).value or ''
    source = ws.cell(row=row_idx, column=COL['source']).value or ''
    unit = ws.cell(row=row_idx, column=COL['unit']).value or ''
    lead_time = ws.cell(row=row_idx, column=COL['lead_time']).value or ''
    price = ws.cell(row=row_idx, column=COL['price']).value or ''
    creator = ws.cell(row=row_idx, column=COL['creator']).value or ''
    create_date = ws.cell(row=row_idx, column=COL['create_date']).value or ''
    modifier = ws.cell(row=row_idx, column=COL['modifier']).value or ''
    modify_date = ws.cell(row=row_idx, column=COL['modify_date']).value or ''
    color = ws.cell(row=row_idx, column=COL['color']).value or ''
    
    materials.append({
        'row': row_idx,
        'code': str(code).strip(),
        'name': str(name).strip(),
        'desc': str(desc).strip(),
        'cat': str(cat).strip(),
        'subcat': str(subcat).strip(),
        'manufacturer': str(manufacturer).strip(),
        'source': str(source).strip(),
        'unit': str(unit).strip(),
        'lead_time': str(lead_time).strip(),
        'price': str(price).strip(),
        'creator': str(creator).strip(),
        'create_date': str(create_date).strip(),
        'modifier': str(modifier).strip(),
        'modify_date': str(modify_date).strip(),
        'color': str(color).strip(),
    })

print(f"Total materials (excl. 成品柜/外购成套): {len(materials)}")

# Group by (cat + subcat) for comparison
groups = defaultdict(list)
for m in materials:
    key = (m['cat'], m['subcat'])
    groups[key].append(m)

print(f"Groups: {len(groups)}")

# Normalize description for comparison
def normalize(s):
    """Normalize string for comparison"""
    s = s.lower().strip()
    # Remove extra spaces
    s = re.sub(r'\s+', ' ', s)
    # Normalize common variations
    s = s.replace('（', '(').replace('）', ')')
    s = s.replace('，', ',').replace('：', ':')
    s = s.replace('×', 'x').replace('—', '-')
    return s

# Exclusion rules
def apply_exclusion_rules(a, b):
    """
    Apply 27 exclusion rules. Returns (is_excluded, reason) or (False, '')
    """
    desc_a = a['desc']
    desc_b = b['desc']
    name_a = a['name']
    name_b = b['name']
    mfg_a = a['manufacturer']
    mfg_b = b['manufacturer']
    cat_a = a['cat']
    cat_b = b['cat']
    subcat_a = a['subcat']
    subcat_b = b['subcat']
    
    # Rule 26: 成品柜/外购成套 - already excluded at load time
    
    # Rule 1: 甲供件
    if '甲供' in cat_a or '甲供' in cat_b or '甲供' in subcat_a or '甲供' in subcat_b:
        return True, '甲供件不算重复'
    
    # Rule 2: One has manufacturer, one doesn't
    if (mfg_a and not mfg_b) or (not mfg_a and mfg_b):
        return True, '一个有制造商一个没有'
    
    # Rule 3: Manufacturers completely different
    if mfg_a and mfg_b and mfg_a != mfg_b:
        # Check if they might be same company different names
        # Common aliases
        mfg_pairs = [
            ('宁波三爱', '三爱'), ('海越电气', '海越湖北'), ('天灵', '宁波天灵'),
            ('宁波天灵', '天灵'), ('长沙威胜', '威胜'),
        ]
        is_alias = False
        for x, y in mfg_pairs:
            if (x in mfg_a and y in mfg_b) or (y in mfg_a and x in mfg_b):
                is_alias = True
                break
            if (x in mfg_b and y in mfg_a) or (y in mfg_b and x in mfg_a):
                is_alias = True
                break
        if not is_alias:
            return True, f'制造商不同: {mfg_a} vs {mfg_b}'
    
    # Rule 4: Direction differences
    directions = ['左操', '右操', '平进平出', '平进侧出', '上进', '下进', 
                  '上进下出', '下进上出', '侧进', '侧出', '上出', '下出',
                  '水平', '垂直', '立式', '卧式']
    for d in directions:
        if d in desc_a and d not in desc_b and any(dd in desc_b for dd in directions if dd != d):
            # Only if there's a conflicting direction
            conflict = False
            for d2 in directions:
                if d2 != d and d2 in desc_b and d2 not in desc_a:
                    conflict = True
                    break
            if conflict:
                return True, f'方向不同: {d} vs others'
    
    # Rule 5: Breaker frame letter suffix (F/C/D/H/L/M/N/S)
    frame_pattern = r'(CDM\d+|NDM\d+|CM\d+|NM\d+)-?(\d+)([FCDHLMNS])'
    match_a = re.search(frame_pattern, desc_a, re.IGNORECASE)
    match_b = re.search(frame_pattern, desc_b, re.IGNORECASE)
    if match_a and match_b:
        if match_a.group(2) == match_b.group(2) and match_a.group(3).upper() != match_b.group(3).upper():
            return True, f'断路器分断能力等级不同: {match_a.group(3)} vs {match_b.group(3)}'
    
    # Rule 6: MCB trip curve B/C/D
    curve_pattern = r'[/-]([BCD])[一二三四五六七八九十\d]*\d{1,3}[/A]'
    ca = re.search(curve_pattern, desc_a)
    cb = re.search(curve_pattern, desc_b)
    if ca and cb and ca.group(1) != cb.group(1):
        return True, f'脱扣曲线不同: {ca.group(1)} vs {cb.group(1)}'
    
    # Rule 7: Pole count 3P/4P/1P etc
    pole_pattern = r'(\d)[Pp]'
    poles_a = set(re.findall(pole_pattern, desc_a))
    poles_b = set(re.findall(pole_pattern, desc_b))
    if poles_a and poles_b and poles_a != poles_b:
        return True, f'极数不同: {poles_a} vs {poles_b}'
    
    # Rule 8: Color different
    color_a = a.get('color', '')
    color_b = b.get('color', '')
    if color_a and color_b and color_a != color_b:
        return True, f'颜色不同: {color_a} vs {color_b}'
    
    # Rule 9: Leakage type A vs AC
    if ('A型' in desc_a and 'AC型' in desc_b) or ('AC型' in desc_a and 'A型' in desc_b):
        return True, '漏电保护类型不同'
    
    # Rule 10: Attachment differences
    attachments = ['失压', '辅助', '报警', '门框', '电磁锁', '分励', '合闸', 
                   'RS485', '通讯', '温度指示', '带电显示']
    attach_a = [x for x in attachments if x in desc_a]
    attach_b = [x for x in attachments if x in desc_b]
    if set(attach_a) != set(attach_b) and (attach_a or attach_b):
        diff = set(attach_a) ^ set(attach_b)
        if diff:
            return True, f'附件不同: {diff}'
    
    # Rule 11: CT differences
    if '互感器' in name_a or '互感器' in name_b:
        # Compare window size, accuracy
        acc_pattern = r'(\d+(?:\.\d+)?[Ss]?(?:级)?)'
        # Check for different specifications
        pass  # Will be handled by general desc comparison
    
    # Rule 12: Wire type BV vs BVR etc
    wire_types = ['BV', 'BVR', 'YJV', 'ZR-YJV', 'RVV', 'RV', 'RVB']
    wire_a = [w for w in wire_types if w in desc_a]
    wire_b = [w for w in wire_types if w in desc_b]
    if wire_a and wire_b and set(wire_a) != set(wire_b):
        return True, f'电线类型不同: {wire_a} vs {wire_b}'
    
    # Rule 13: Protection type
    prot_types = ['变压器保护', '线路保护', '电容器保护', '电动机保护']
    prot_a = [p for p in prot_types if p in desc_a or p in name_a]
    prot_b = [p for p in prot_types if p in desc_b or p in name_b]
    if prot_a and prot_b and set(prot_a) != set(prot_b):
        return True, f'保护类型不同'
    
    # Rule 14: Thread direction
    if ('正牙' in desc_a and '反牙' in desc_b) or ('反牙' in desc_a and '正牙' in desc_b):
        return True, '螺纹方向不同'
    
    # Rule 15: Version
    ver_pattern = r'[Vv](\d+[A-Za-z]?)'
    ver_a = re.findall(ver_pattern, desc_a)
    ver_b = re.findall(ver_pattern, desc_b)
    if ver_a and ver_b and ver_a != ver_b:
        return True, f'版本不同'
    
    # Rule 16: Trip method TMD vs MA
    if ('TMD' in desc_a and 'MA' in desc_b) or ('MA' in desc_a and 'TMD' in desc_b):
        return True, '脱扣方式不同(TMD vs MA)'
    
    # Rule 17: Trip unit LSI vs TMA
    if ('LSI' in desc_a and 'TMA' in desc_b) or ('TMA' in desc_a and 'LSI' in desc_b):
        return True, '脱扣单元不同(LSI vs TMA)'
    
    # Rule 18: Accessories different (safety lock vs phase barrier etc)
    acc2 = ['安全挂锁', '相间隔板', '挂锁', '隔板']
    acc2_a = [x for x in acc2 if x in desc_a]
    acc2_b = [x for x in acc2 if x in desc_b]
    if acc2_a and acc2_b and set(acc2_a) != set(acc2_b):
        return True, f'配件不同'
    
    # Rule 19: Overvoltage protector model TBP-A vs TBP-B
    if 'TBP' in desc_a and 'TBP' in desc_b:
        tbp_a = re.search(r'TBP[-]?([A-Z])', desc_a)
        tbp_b = re.search(r'TBP[-]?([A-Z])', desc_b)
        if tbp_a and tbp_b and tbp_a.group(1) != tbp_b.group(1):
            return True, f'过电压保护器型号不同: TBP-{tbp_a.group(1)} vs TBP-{tbp_b.group(1)}'
    
    # Rule 20: Model series CM3E vs CM3
    series_patterns = [
        (r'(CM3E)', r'(CM3[^E])'),  # CM3E electronic vs CM3 standard
        (r'(NDM3)', r'(CDM3)'),  # Different series
    ]
    for p1, p2 in series_patterns:
        if re.search(p1, desc_a) and re.search(p2, desc_b):
            return True, '型号系列不同'
        if re.search(p2, desc_a) and re.search(p1, desc_b):
            return True, '型号系列不同'
    
    # Rule 21: RV vs BVR
    if ('RV' in desc_a and 'BVR' in desc_b) or ('BVR' in desc_a and 'RV' in desc_b):
        return True, 'RV vs BVR 电线类型不同'
    
    # Rule 22: Material differences
    mat_keywords = [('铜', '铝'), ('紫铜', '黄铜'), ('304', '316')]
    for m1, m2 in mat_keywords:
        if (m1 in desc_a and m2 in desc_b) or (m2 in desc_a and m1 in desc_b):
            return True, f'材质不同: {m1} vs {m2}'
    
    # Rule 23: Size number differences in model
    # Compare numeric values that likely represent dimensions
    # Extract key numbers from descriptions
    
    # Rule 24: Current/voltage parameter differences
    # Extract current values (e.g., 63A, 100A, 250A)
    curr_pattern = r'(\d+(?:\.\d+)?)\s*[Aa]'
    currs_a = re.findall(curr_pattern, desc_a)
    currs_b = re.findall(curr_pattern, desc_b)
    if currs_a and currs_b:
        # Get the main current (usually the first or largest)
        try:
            main_a = float(currs_a[0])
            main_b = float(currs_b[0])
            if abs(main_a - main_b) > 1 and main_a > 0 and main_b > 0:
                # Check if these are clearly different current ratings
                if main_a / main_b > 1.1 or main_b / main_a > 1.1:
                    return True, f'额定电流不同: {currs_a[0]}A vs {currs_b[0]}A'
        except:
            pass
    
    # Rule 25: Model encoding position differences
    # If descriptions share a base model but differ in specific positions
    # This is handled by the general model comparison
    
    # Rule 27: EPS power differences
    eps_pattern = r'EPS\s*(\d+(?:\.\d+)?)\s*KW'
    eps_a = re.search(eps_pattern, desc_a, re.IGNORECASE)
    eps_b = re.search(eps_pattern, desc_b, re.IGNORECASE)
    if eps_a and eps_b and eps_a.group(1) != eps_b.group(1):
        return True, f'EPS功率不同: {eps_a.group(1)}KW vs {eps_b.group(1)}KW'
    
    return False, ''


def is_confirmed_duplicate(a, b):
    """Check if a pair is confirmed duplicate (same product, different formatting)"""
    desc_a = normalize(a['desc'])
    desc_b = normalize(b['desc'])
    
    # If descriptions are identical after normalization
    if desc_a == desc_b:
        return True
    
    # 1N vs 1P+N (Schneider convention)
    desc_a_1n = desc_a.replace('1n', '1p+n').replace('1N', '1P+N')
    desc_b_1n = desc_b.replace('1n', '1p+n').replace('1N', '1P+N')
    if desc_a_1n == desc_b_1n:
        return True
    
    # Minor formatting differences only
    # Remove all spaces, punctuation for comparison
    def ultra_normalize(s):
        s = re.sub(r'[\s\-_/\\(),;:.\[\]{}]', '', s)
        s = s.replace('（', '').replace('）', '').replace('，', '').replace('：', '')
        s = s.replace('×', 'x').replace('—', '').replace('"', '').replace("'", '')
        s = s.replace('全角', '').replace('半角', '')
        return s
    
    ua = ultra_normalize(desc_a)
    ub = ultra_normalize(desc_b)
    
    if ua == ub:
        return True
    
    return False


def find_suspicious_pairs(materials):
    """Find pairs of materials with similar descriptions within same (cat, subcat) group"""
    suspicious = []
    
    # Group by (cat, subcat)
    groups = defaultdict(list)
    for m in materials:
        key = (m['cat'], m['subcat'])
        groups[key].append(m)
    
    for key, group in groups.items():
        if len(group) < 2:
            continue
        
        # Build description-based groups
        # First pass: exact name matches
        name_groups = defaultdict(list)
        for m in group:
            name_groups[m['name']].append(m)
        
        for name, items in name_groups.items():
            if len(items) < 2:
                continue
            # Compare all pairs within same name
            for i in range(len(items)):
                for j in range(i+1, len(items)):
                    a, b = items[i], items[j]
                    # Skip if same description (unless different code)
                    if a['desc'] == b['desc'] and a['code'] != b['code']:
                        suspicious.append((a, b))
                    elif a['desc'] != b['desc']:
                        # Check similarity
                        norm_a = normalize(a['desc'])
                        norm_b = normalize(b['desc'])
                        # Simple similarity check
                        if norm_a == norm_b or is_confirmed_duplicate(a, b):
                            suspicious.append((a, b))
                        else:
                            # Check if descriptions are very similar
                            # (one is substring of other, or share common prefix)
                            if len(norm_a) > 5 and len(norm_b) > 5:
                                # Calculate simple character overlap
                                common = sum(1 for c in norm_a if c in norm_b)
                                ratio = common / max(len(norm_a), len(norm_b))
                                if ratio > 0.85:
                                    suspicious.append((a, b))
        
        # Second pass: similar names across different name values
        # Group by normalized name
        norm_groups = defaultdict(list)
        for m in group:
            nn = normalize(m['name'])
            norm_groups[nn].append(m)
        
        # Check across normalized name groups that differ slightly
        norm_keys = list(norm_groups.keys())
        for i in range(len(norm_keys)):
            for j in range(i+1, len(norm_keys)):
                k1, k2 = norm_keys[i], norm_keys[j]
                # Check if names are very similar
                common = sum(1 for c in k1 if c in k2)
                ratio = common / max(len(k1), len(k2)) if max(len(k1), len(k2)) > 0 else 0
                if ratio > 0.8 and k1 != k2:
                    # Cross-compare descriptions
                    for ma in norm_groups[k1]:
                        for mb in norm_groups[k2]:
                            norm_desc_a = normalize(ma['desc'])
                            norm_desc_b = normalize(mb['desc'])
                            if norm_desc_a == norm_desc_b:
                                suspicious.append((ma, mb))
    
    return suspicious


print("Finding suspicious pairs...")
suspicious = find_suspicious_pairs(materials)
print(f"Found {len(suspicious)} suspicious pairs")

# Deduplicate pairs (same pair might be found multiple ways)
seen_pairs = set()
unique_suspicious = []
for a, b in suspicious:
    pair_key = tuple(sorted([a['code'], b['code']]))
    if pair_key not in seen_pairs:
        seen_pairs.add(pair_key)
        unique_suspicious.append((a, b))

print(f"Unique suspicious pairs: {len(unique_suspicious)}")

# Classify pairs
confirmed = []  # Confirmed duplicates
non_duplicate = []  # Non-duplicates (ruled out)
needs_review = []  # Needs manual review

for a, b in unique_suspicious:
    is_excluded, reason = apply_exclusion_rules(a, b)
    if is_excluded:
        non_duplicate.append((a, b, reason))
    elif is_confirmed_duplicate(a, b):
        # Check again - even confirmed dup might have different descriptions
        if a['desc'] == b['desc']:
            confirmed.append((a, b, '描述完全相同'))
        else:
            # Descriptions differ slightly but same product
            confirmed.append((a, b, '描述仅格式差异'))
    else:
        # Need more careful analysis
        # Check if descriptions are actually different models
        desc_a = a['desc']
        desc_b = b['desc']
        
        # If names are the same and descriptions are very similar
        if a['name'] == b['name']:
            norm_a = normalize(desc_a)
            norm_b = normalize(desc_b)
            common = sum(1 for c in norm_a if c in norm_b)
            ratio = common / max(len(norm_a), len(norm_b)) if max(len(norm_a), len(norm_b)) > 0 else 0
            
            if ratio > 0.9:
                needs_review.append((a, b, '描述高度相似，需人工确认'))
            elif ratio > 0.85:
                needs_review.append((a, b, '描述较相似，需人工确认'))
            else:
                # Probably not duplicate - descriptions quite different
                non_duplicate.append((a, b, '描述差异较大'))
        else:
            needs_review.append((a, b, '名称不同但相似，需人工确认'))

print(f"\nResults:")
print(f"  Confirmed duplicates: {len(confirmed)}")
print(f"  Non-duplicates: {len(non_duplicate)}")
print(f"  Needs review: {len(needs_review)}")
print(f"  Total: {len(confirmed) + len(non_duplicate) + len(needs_review)}")

# Generate Excel
print("\nGenerating Excel report...")

out_wb = openpyxl.Workbook()

# Define styles
blue_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
orange_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
red_font = Font(color='FF0000')
normal_font = Font()
header_font = Font(bold=True)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(bold=True, color='FFFFFF')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Data columns to show in Excel
EXCEL_COLS = [
    ('序号', 'label'),
    ('物料编号', 'code'),
    ('物料名称', 'name'),
    ('物料描述', 'desc'),
    ('物料类别', 'cat'),
    ('物料子类别', 'subcat'),
    ('制造商', 'manufacturer'),
    ('物料来源', 'source'),
    ('主计量单位', 'unit'),
    ('提前期', 'lead_time'),
    ('标准价格', 'price'),
    ('创建人', 'creator'),
    ('创建日期', 'create_date'),
    ('最近修改人', 'modifier'),
    ('最近修改日期', 'modify_date'),
    ('备注', 'remark'),
]

# Columns to highlight red if different
DIFF_COLS = {'name', 'desc', 'cat', 'subcat', 'manufacturer', 'source', 'unit', 'lead_time', 'price'}
# Columns to never highlight red
NO_DIFF_COLS = {'label', 'code', 'creator', 'create_date', 'modifier', 'modify_date', 'remark'}


def write_sheet(ws, data, sheet_name):
    """Write a sheet with alternating A/B rows"""
    # Headers
    for col_idx, (col_name, _) in enumerate(EXCEL_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    for pair_num, (a, b, reason) in enumerate(data, 1):
        for is_b, mat in enumerate([a, b]):
            fill = orange_fill if is_b else blue_fill
            label = f"{'B' if is_b else 'A'}-{pair_num}"
            
            for col_idx, (col_name, col_key) in enumerate(EXCEL_COLS, 1):
                if col_key == 'label':
                    value = label
                elif col_key == 'remark':
                    value = reason
                else:
                    value = mat.get(col_key, '')
                    if isinstance(value, float) and value == int(value):
                        value = int(value)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=(col_key == 'desc'))
                
                # Highlight differences in red
                if col_key in DIFF_COLS:
                    val_a = str(a.get(col_key, '') or '').strip()
                    val_b = str(b.get(col_key, '') or '').strip()
                    if val_a != val_b:
                        cell.font = red_font
                    else:
                        cell.font = normal_font
                else:
                    cell.font = normal_font
            
            row_idx += 1
    
    # Set column widths
    col_widths = {
        '序号': 8, '物料编号': 16, '物料名称': 18, '物料描述': 60,
        '物料类别': 12, '物料子类别': 14, '制造商': 20,
        '物料来源': 10, '主计量单位': 10, '提前期': 8, '标准价格': 12,
        '创建人': 10, '创建日期': 14, '最近修改人': 12, '最近修改日期': 14,
        '备注': 45
    }
    for col_idx, (col_name, _) in enumerate(EXCEL_COLS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 12)
    
    # Freeze panes (freeze first 2 rows and first 2 columns)
    ws.freeze_panes = 'C2'
    
    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLS))}{row_idx - 1}"


# Create sheets
if confirmed:
    ws1 = out_wb.active
    ws1.title = '确认重复'
    write_sheet(ws1, confirmed, '确认重复')
else:
    ws1 = out_wb.active
    ws1.title = '确认重复'
    for col_idx, (col_name, _) in enumerate(EXCEL_COLS, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font_white
        cell.fill = header_fill

ws2 = out_wb.create_sheet('非重复')
write_sheet(ws2, non_duplicate, '非重复')

ws3 = out_wb.create_sheet('待人工确认')
write_sheet(ws3, needs_review, '待人工确认')

# Save
out_wb.save(OUTPUT_FILE)
print(f"\nExcel saved to: {OUTPUT_FILE}")
print(f"Done! Total time: {time.time()-start:.1f}s")
