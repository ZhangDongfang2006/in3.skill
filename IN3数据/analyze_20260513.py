#!/usr/bin/env python3
"""IN3 物料重复检查 v3 - 严格匹配，只输出确认重复+待人工确认"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from difflib import SequenceMatcher
import re, time

INPUT = '/Users/zhangdongfang/.openclaw/workspace-in3bot/IN3数据/物料主数据导出结果-20260513.xlsx'
OUTPUT = '/Users/zhangdongfang/.openclaw/workspace-in3bot/IN3数据/可疑重复物料-2026-05-13.xlsx'

COL = {'code':2, 'name':6, 'desc':7, 'cat':9, 'subcat':11, 'manufacturer':13,
       'source':14, 'lead_time':18, 'unit':22, 'price':61,
       'creator':70, 'create_date':72, 'modifier':73, 'modify_date':75, 'color':16}

print("Loading...")
t0 = time.time()
wb = openpyxl.load_workbook(INPUT, data_only=True)
ws = wb['物料主数据']
print(f"  Sheet has {ws.max_row} rows, {ws.max_column} cols")

def gv(row, col):
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ''

exclude_cats = {'成品柜', '外购成套'}
materials = []
for r in range(2, ws.max_row + 1):
    cat = gv(r, COL['cat'])
    if cat in exclude_cats:
        continue
    materials.append({
        'code': gv(r, COL['code']), 'name': gv(r, COL['name']),
        'desc': gv(r, COL['desc']), 'cat': cat,
        'subcat': gv(r, COL['subcat']), 'manufacturer': gv(r, COL['manufacturer']),
        'source': gv(r, COL['source']), 'unit': gv(r, COL['unit']),
        'lead_time': gv(r, COL['lead_time']), 'price': gv(r, COL['price']),
        'creator': gv(r, COL['creator']), 'create_date': gv(r, COL['create_date']),
        'modifier': gv(r, COL['modifier']), 'modify_date': gv(r, COL['modify_date']),
        'color': gv(r, COL['color']),
    })
wb.close()
print(f"Loaded {len(materials)} materials in {time.time()-t0:.1f}s")

def norm(s):
    s = s.lower().strip()
    s = re.sub(r'\s+', ' ', s)
    s = s.replace('（','(').replace('）',')').replace('，',',').replace('：',':')
    s = s.replace('×','x').replace('—','-').replace('φ','Φ').replace('ø','Φ')
    return s

def ultra(s):
    s = norm(s)
    # 保留数字之间的小数点（如 2.5），去除其他标点
    s = re.sub(r'(?<=\d)\.(?=\d)', '<<DOT>>', s)  # 临时保护小数点
    s = re.sub(r'[\s\-_/\\(),;:.，：；、。{}\[\]<>]+', '', s)
    s = s.replace('<<DOT>>', '.')
    return s

def sim(a, b):
    return SequenceMatcher(None, norm(a), norm(b)).ratio()

def is_nondup(a, b):
    """判断是否明显非重复。返回 (is_nondup, reason)"""
    da, db = a['desc'], b['desc']
    na, nb = a['name'], b['name']
    ta, tb = da + ' ' + na, db + ' ' + nb
    ma, mb = a['manufacturer'], b['manufacturer']

    # 甲供件
    for m in [a, b]:
        if '甲供' in m['cat'] or '甲供' in m['subcat']:
            return True, '甲供件不算重复'

    # 一个有制造商一个没有
    if (ma and not mb) or (not ma and mb):
        return True, '一个有制造商一个没有'

    # 制造商不同
    if ma and mb and ma != mb:
        aliases = [('宁波三爱','三爱'),('海越电气','海越湖北'),('天灵','宁波天灵'),
                   ('长沙威胜','威胜'),('上海良信电器股份有限公司','良信'),('上海良信电器','良信')]
        if not any((x in ma and y in mb) or (y in ma and x in mb) for x,y in aliases):
            return True, f'制造商不同: {ma} vs {mb}'

    # 内vs外
    iw = ['内六角','内门','内侧','内部','内层','内牙','内螺纹']
    ow = ['外六角','外门','外侧','外部','外层','外牙','外螺纹']
    for i in iw:
        for o in ow:
            if (i in ta and o in tb) or (o in ta and i in tb):
                return True, f'内vs外: {i} vs {o}'

    # 螺丝头部形状不同
    head_types = ['十字','一字','内六角','外六角','梅花','半圆头','沉头','平头','圆头','盘头',
                  '六角头','方头','圆柱头','扁圆头','大扁头','伞头']
    ht_a = set(h for h in head_types if h in ta)
    ht_b = set(h for h in head_types if h in tb)
    if ht_a and ht_b and ht_a != ht_b:
        return True, f'螺丝头部形状不同: {ht_a} vs {ht_b}'

    # 分闸vs合闸
    if ('分闸' in ta and '合闸' in tb) or ('合闸' in ta and '分闸' in tb):
        return True, '分闸 vs 合闸: 不同功能'
    if ('分励' in ta and '合闸' in tb) or ('合闸' in ta and '分励' in tb):
        return True, '分励 vs 合闸: 不同功能'

    # 静vs动
    if ('静插件' in ta and '动插件' in tb) or ('动插件' in ta and '静插件' in tb):
        return True, '静插件 vs 动插件: 不同部件'
    if ('静触头' in ta and '动触头' in tb) or ('动触头' in ta and '静触头' in tb):
        return True, '静触头 vs 动触头: 不同部件'
    if ('静' in ta and '动' in tb and '插件' in ta) or ('动' in ta and '静' in tb and '插件' in tb):
        return True, '静vs动: 不同部件'

    # 形状
    shapes = [('方形','圆形'),('方型','圆型'),('U型','对接型'),('U型','直型'),('平型','立式')]
    for s1,s2 in shapes:
        if (s1 in ta and s2 in tb) or (s2 in ta and s1 in tb):
            return True, f'形状不同: {s1} vs {s2}'

    # 安装方式: 明装vs暗装
    if ('明装' in ta and '暗装' in tb) or ('暗装' in ta and '明装' in tb):
        return True, '明装 vs 暗装: 不同安装方式'
    if ('明装' in ta and '嵌入式' in tb) or ('嵌入式' in ta and '明装' in tb):
        return True, '明装 vs 嵌入式: 不同安装方式'
    if ('暗装' in ta and '嵌入式' in tb) or ('嵌入式' in ta and '暗装' in tb):
        return True, '暗装 vs 嵌入式: 不同安装方式'
    # 暗装 vs 无标注（但另一方名称含户外/壁挂/落地等不同安装方式）
    install_keywords = ['暗装','户外','户外型','户外架','壁挂','落地','嵌墙','挂墙','明装','嵌入式']
    inst_a = set(k for k in install_keywords if k in ta)
    inst_b = set(k for k in install_keywords if k in tb)
    # 两者都有安装方式但不同
    if inst_a and inst_b and inst_a != inst_b:
        return True, f'安装方式不同: {inst_a} vs {inst_b}'
    # 一方有安装方式关键词，另一方名称完全相同但无安装方式 = 需要人工确认（不自动排除）

    # 方向
    dirs = [('左操','右操'),('平进平出','平进侧出'),('上进','下进'),('上进下出','下进上出'),
            ('立式','卧式'),('侧进','上进'),('侧出','下出')]
    for d1,d2 in dirs:
        if (d1 in ta and d2 in tb) or (d2 in ta and d1 in tb):
            return True, f'方向不同: {d1} vs {d2}'

    # 断路器分断能力
    fp = r'(?:CDM|NDM|CM|NM)\d*-?\d+([FCDHLMNS])'
    fa, fb = re.search(fp, da, re.I), re.search(fp, db, re.I)
    if fa and fb and fa.group(1).upper() != fb.group(1).upper():
        return True, f'分断能力不同: {fa.group(1)} vs {fb.group(1)}'

    # 脱扣曲线 B/C/D
    ca = re.search(r'(?:^|[\s/\-+])([BCDbcd])(\d+)', da)
    cb = re.search(r'(?:^|[\s/\-+])([BCDbcd])(\d+)', db)
    if ca and cb and ca.group(1).upper() != cb.group(1).upper():
        return True, f'脱扣曲线不同: {ca.group(1)} vs {cb.group(1)}'

    # 极数
    pa = set(re.findall(r'(\d)\s*[Pp]', da))
    pb = set(re.findall(r'(\d)\s*[Pp]', db))
    if pa and pb and pa != pb:
        return True, f'极数不同: {pa} vs {pb}'

    # 颜色
    if a.get('color') and b.get('color') and a['color'] != b['color']:
        return True, f'颜色不同: {a["color"]} vs {b["color"]}'

    # 漏电类型
    if ('A型' in da and 'AC型' in db) or ('AC型' in da and 'A型' in db):
        return True, '漏电类型不同'

    # 附件差异
    atts = ['失压','分励','合闸','辅助','报警','门框','电磁锁','RS485','通讯','温度指示','带电显示','加热器']
    aa = set(x for x in atts if x in da)
    ab = set(x for x in atts if x in db)
    diff = aa ^ ab
    if diff:
        return True, f'附件不同: {",".join(sorted(diff))}'

    # 电线类型
    wires = ['BV','BVR','YJV','ZR-YJV','RVV','RV','RVB','NH-YJV']
    wa = set(w for w in wires if w in da)
    wb2 = set(w for w in wires if w in db)
    if wa and wb2 and wa != wb2:
        return True, f'电线类型不同: {wa} vs {wb2}'

    # 螺纹方向
    if ('正牙' in ta and '反牙' in tb) or ('反牙' in ta and '正牙' in tb):
        return True, '螺纹方向不同'

    # 脱扣方式
    if ('TMD' in da.upper() and 'MA' in db.upper()) or ('MA' in da.upper() and 'TMD' in db.upper()):
        return True, '脱扣方式不同'

    # 脱扣单元
    if ('LSI' in da and 'TMA' in db) or ('TMA' in da and 'LSI' in db):
        return True, '脱扣单元不同'

    # 材质
    mat_pairs = [('铜','铝'),('304','316'),('紫铜','黄铜')]
    for m1,m2 in mat_pairs:
        if (m1 in ta and m2 in tb) or (m2 in ta and m1 in tb):
            return True, f'材质不同: {m1} vs {m2}'

    # **核心：提取描述中的关键参数数字，不同则非重复**
    # 提取尺寸: 数字x数字 模式
    dims_a = re.findall(r'(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)(?:\s*x\s*(\d+(?:\.\d+)?))?', norm(da))
    dims_b = re.findall(r'(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)(?:\s*x\s*(\d+(?:\.\d+)?))?', norm(db))
    if dims_a and dims_b and dims_a != dims_b:
        return True, f'尺寸不同: {dims_a} vs {dims_b}'
    
    # 母线/铜排尺寸: TMY-50x5 之类
    bus_a = re.findall(r'(?:TMY|TMR|LMY)[\-]?(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)', norm(da), re.I)
    bus_b = re.findall(r'(?:TMY|TMR|LMY)[\-]?(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)', norm(db), re.I)
    if bus_a and bus_b and bus_a != bus_b:
        return True, f'母线尺寸不同: {bus_a} vs {bus_b}'
    
    # 提取独立宽度/长度数值: 宽XXXmm、L=XXX
    width_a = re.findall(r'(?:宽|宽度|W)[=:\s]*(\d+(?:\.\d+)?)\s*(?:mm|MM)?', da)
    width_b = re.findall(r'(?:宽|宽度|W)[=:\s]*(\d+(?:\.\d+)?)\s*(?:mm|MM)?', db)
    if width_a and width_b and width_a != width_b:
        return True, f'宽度不同: {width_a} vs {width_b}'

    # 提取直径: Φ数字 或 直径数字
    dia_a = re.findall(r'(?:Φ|φ|直径)[=:\s]*(\d+(?:\.\d+)?)', da, re.I)
    dia_b = re.findall(r'(?:Φ|φ|直径)[=:\s]*(\d+(?:\.\d+)?)', db, re.I)
    if dia_a and dia_b and dia_a != dia_b:
        return True, f'直径不同: {dia_a} vs {dia_b}'

    # 提取电流: 数字A
    cur_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Aa](?:\s|$|[^a-zA-Z0-9])', da)))
    cur_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Aa](?:\s|$|[^a-zA-Z0-9])', db)))
    if cur_a and cur_b and cur_a != cur_b:
        return True, f'电流不同: {cur_a}A vs {cur_b}A'

    # 提取电压: 数字V
    vol_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Vv](?:\s|$|[^a-zA-Z0-9])', da)))
    vol_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Vv](?:\s|$|[^a-zA-Z0-9])', db)))
    if vol_a and vol_b and vol_a != vol_b:
        return True, f'电压不同: {vol_a}V vs {vol_b}V'

    # 提取功率: 数字KW
    pow_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Kk][Ww]', da)))
    pow_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Kk][Ww]', db)))
    if pow_a and pow_b and pow_a != pow_b:
        return True, f'功率不同: {pow_a}KW vs {pow_b}KW'

    # 提取线径/截面积: 描述中的独立数字（如 RV 2.5、BV 4、6mm²）
    # 先找 XXmm² 或 XX平方 模式
    area_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*(?:mm²|mm2|平方)', da)))
    area_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*(?:mm²|mm2|平方)', db)))
    if area_a and area_b and area_a != area_b:
        return True, f'截面积不同: {area_a} vs {area_b}'
    
    # 对于电线/导线类，描述中的数字往往代表截面积（如 RV 2.5 黑 vs RV 25 黑）
    if a['subcat'] in ('导线','电线') or b['subcat'] in ('导线','电线'):
        nums_a = sorted(set(re.findall(r'(?:^|\s)(\d+(?:\.\d+)?)(?:\s|$)', da)))
        nums_b = sorted(set(re.findall(r'(?:^|\s)(\d+(?:\.\d+)?)(?:\s|$)', db)))
        if nums_a and nums_b and nums_a != nums_b:
            return True, f'线径/截面积不同: {nums_a} vs {nums_b}'

    # **核心：型号比较**
    # 先比较名称中的数字（如 24芯 vs 48芯）
    name_nums_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)', a['name'])))
    name_nums_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)', b['name'])))
    if name_nums_a and name_nums_b and name_nums_a != name_nums_b:
        return True, f'名称数字不同: {a["name"]} vs {b["name"]}'

    # 柜型型号不同 = 非重复
    cabinet_models = ['GGD','GCK','GCS','MNS','MCS','KYN28','KYN61','XGN','HXGN',
                      'DFX','YBM','ZBW','ZGS','KYN','XGN','HXGN']
    found_cab_a = set()
    found_cab_b = set()
    for cm in cabinet_models:
        if cm.lower() in ta.lower():
            found_cab_a.add(cm)
        if cm.lower() in tb.lower():
            found_cab_b.add(cm)
    if found_cab_a and found_cab_b and found_cab_a != found_cab_b:
        return True, f'柜型不同: {found_cab_a} vs {found_cab_b}'

    # 提取型号 token (字母+数字组合)
    models_a = re.findall(r'[a-zA-Z]+\d+[a-zA-Z0-9\-/]*', da)
    models_b = re.findall(r'[a-zA-Z]+\d+[a-zA-Z0-9\-/]*', db)
    if models_a and models_b:
        main_a = max(models_a, key=len)
        main_b = max(models_b, key=len)
        # 标准化比较
        ma_low = main_a.lower().replace('-','').replace('/','')
        mb_low = main_b.lower().replace('-','').replace('/','')
        if ma_low != mb_low:
            # 提取前缀（字母部分）
            pre_a = re.match(r'^([a-zA-Z]+)', main_a)
            pre_b = re.match(r'^([a-zA-Z]+)', main_b)
            if pre_a and pre_b:
                pa_s = pre_a.group(1).lower()
                pb_s = pre_b.group(1).lower()
                rest_a = main_a[len(pre_a.group(1)):]
                rest_b = main_b[len(pre_b.group(1)):]
                # 同系列：前缀相同
                if pa_s == pb_s:
                    # 比较剩余部分的数字
                    nums_a = re.findall(r'\d+', rest_a)
                    nums_b = re.findall(r'\d+', rest_b)
                    if nums_a != nums_b:
                        return True, f'型号数字不同: {main_a} vs {main_b}'
                    # 比较剩余字母
                    lets_a = re.findall(r'[a-zA-Z]+', rest_a)
                    lets_b = re.findall(r'[a-zA-Z]+', rest_b)
                    if lets_a and lets_b and lets_a != lets_b:
                        return True, f'型号字母不同: {main_a} vs {main_b}'
                else:
                    # 不同系列
                    base_a = re.sub(r'\d+', '', pa_s)
                    base_b = re.sub(r'\d+', '', pb_s)
                    if base_a != base_b:
                        return True, f'型号系列不同: {main_a} vs {main_b}'

    return False, ''

def is_naming_mismatch(a, b):
    """仅符号/格式差异 = 命名不规范"""
    da, db = a['desc'], b['desc']
    if ultra(da) == ultra(db):
        return True
    return False

def desc_numbers_match(a, b):
    """检查两个描述+名称中的数字集合是否一致（防止标准化后不同规格变一样）"""
    nums_a = sorted(set(re.findall(r'\d+(?:\.\d+)?', a['desc'] + ' ' + a['name'])))
    nums_b = sorted(set(re.findall(r'\d+(?:\.\d+)?', b['desc'] + ' ' + b['name'])))
    return nums_a == nums_b

# 分组: 按类别+子类别+标准化名称
groups = defaultdict(list)
for m in materials:
    key = (m['cat'], m['subcat'], norm(m['name']))
    groups[key].append(m)

# 跨名称: 按类别+子类别+标准化描述
desc_groups = defaultdict(list)
for m in materials:
    key = (m['cat'], m['subcat'], ultra(m['desc']))
    desc_groups[key].append(m)

print(f"Groups by name: {len(groups)}")
print(f"Groups by desc: {len(desc_groups)}")

# 找配对
seen = set()
pairs = []

# 同名组内比较
for key, grp in groups.items():
    if len(grp) < 2:
        continue
    # 同名+同描述 → 直接候选
    desc_map = defaultdict(list)
    for m in grp:
        desc_map[ultra(m['desc'])].append(m)
    
    for dk, items in desc_map.items():
        if len(items) < 2:
            continue
        for i in range(len(items)):
            for j in range(i+1, len(items)):
                a, b = items[i], items[j]
                if a['code'] == b['code']:
                    continue
                # 即使 ultra 描述相同，还要检查数字集合
                if not desc_numbers_match(a, b):
                    continue
                pk = tuple(sorted([a['code'], b['code']]))
                if pk not in seen:
                    seen.add(pk)
                    pairs.append((a, b))
    
    # 同名+不同描述但高度相似
    # 先按 ultra desc 分组，只有相同的才比较
    # 不再做两两低阈值比较

# 跨名称：描述完全一致但名称不同
for key, grp in desc_groups.items():
    if len(grp) < 2:
        continue
    for i in range(len(grp)):
        for j in range(i+1, len(grp)):
            a, b = grp[i], grp[j]
            if a['code'] == b['code']:
                continue
            if a['name'] == b['name']:
                continue  # 同名已在上面处理
            # 检查数字集合
            if not desc_numbers_match(a, b):
                continue
            pk = tuple(sorted([a['code'], b['code']]))
            if pk in seen:
                continue
            seen.add(pk)
            pairs.append((a, b))

print(f"Total pairs to classify: {len(pairs)}")

# 分类
confirmed = []  # 确认重复
review = []     # 待人工确认
nondup_count = 0

for a, b in pairs:
    # 器具/工具类 → 全部待人工确认
    if any(k in a['subcat'] or k in b['subcat'] for k in ['器具','工具','工器具']):
        review.append((a, b, '器具/工具类需人工确认'))
        continue

    # 先检查是否非重复
    nd, reason = is_nondup(a, b)
    if nd:
        nondup_count += 1
        continue

    # 器具/工具类 → 全部待人工确认
    if any(k in a['subcat'] or k in b['subcat'] for k in ['器具','工具','工器具']):
        review.append((a, b, '器具/工具类需人工确认'))
        continue

    # ★ 名称不同的配对优先处理 ★
    if a['name'] != b['name']:
        # 先看标准化名称是否一致（仅符号差异：空格、-、/等）
        norm_na = re.sub(r'[\s\-_/\\(),;:.，：；、。{}\[\]<>]+', '', a['name'].lower())
        norm_nb = re.sub(r'[\s\-_/\\(),;:.，：；、。{}\[\]<>]+', '', b['name'].lower())
        
        if norm_na == norm_nb:
            # 名称仅符号差异 → 确认重复
            confirmed.append((a, b, '命名不规范（名称仅符号差异）'))
            continue
        
        # 名称有实质性文字差异 → 非重复
        na_words = set(re.findall(r'[\u4e00-\u9fff]{2,}', a['name']))
        nb_words = set(re.findall(r'[\u4e00-\u9fff]{2,}', b['name']))
        common = na_words & nb_words
        if not common and len(na_words) > 0 and len(nb_words) > 0:
            nondup_count += 1
            continue
        
        # 材质差异 → 非重复
        mat_pairs = [('不锈钢','镀彩'),('不锈钢','镀锌'),('不锈钢','镀蓝锌'),
                     ('镀铝锌','镀锌'),('紫铜','黄铜'),('铜','铝'),
                     ('不锈钢','发黑'),('白锌','镀彩')]
        is_mat_diff = any(
            (m1 in a['name'] and m2 in b['name']) or (m2 in a['name'] and m1 in b['name'])
            for m1, m2 in mat_pairs
        )
        if is_mat_diff:
            nondup_count += 1
            continue
        
        # 类型/功能差异 → 非重复
        type_pairs = [('熔芯','底座'),('三通','活接'),('三通','直通'),('定向','万向'),
                      ('手动','电动'),('燕尾','防爆'),('滚轮','脚轮'),
                      ('轻轨','重轨'),('盲板','凸面'),('管夹','垫片'),
                      ('车轮','手轮'),('驱动大头','驱动链条'),('驱动器','顶升'),
                      ('大门','加热包'),('绝缘靴','绝缘手套'),
                      ('电度表','电能表'),('电压','电流'),('两孔','三孔'),
                      ('简易版','全功能')]
        is_type_diff = any(
            (t1 in a['name'] and t2 in b['name']) or (t2 in a['name'] and t1 in b['name'])
            for t1, t2 in type_pairs
        )
        if is_type_diff:
            nondup_count += 1
            continue
        
        # 形状差异 → 非重复
        shape_pairs = [('方形','圆形'),('方形','管装'),('U型','对接')]
        is_shape_diff = any(
            (s1 in a['name'] and s2 in b['name']) or (s2 in a['name'] and s1 in b['name'])
            for s1, s2 in shape_pairs
        )
        if is_shape_diff:
            nondup_count += 1
            continue
        
        # 其余名称不同 → 待人工确认
        review.append((a, b, '名称不同但描述一致'))
        continue

    # 名称相同，检查描述
    if is_naming_mismatch(a, b):
        confirmed.append((a, b, '命名不规范（仅符号/格式差异）'))
        continue

    # 名称相同 + 描述相同
    if a['desc'] == b['desc']:
        confirmed.append((a, b, '描述完全相同'))
        continue
    if sim(a['desc'], b['desc']) >= 0.9:
        confirmed.append((a, b, '描述高度相似'))
        continue

    # 其余待人工确认
    review.append((a, b, '需人工确认'))

print(f"\nResults:")
print(f"  Non-duplicates (excluded): {nondup_count}")
print(f"  Confirmed duplicates: {len(confirmed)}")
print(f"  Needs review: {len(review)}")

# 生成 Excel
print("\nGenerating Excel...")

BLUE = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
ORANGE = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
RED = Font(color='FF0000', size=10)
BLACK = Font(size=10)
thin = Side(style='thin', color='D0D0D0')
BORDER = Border(top=thin, left=thin, right=thin, bottom=thin)
WRAP = Alignment(vertical='center', wrap_text=True)
CENTER = Alignment(vertical='center')

FIELDS = [
    ('标记', 'label'), ('物料编号', 'code'), ('物料名称', 'name'), ('物料描述', 'desc'),
    ('物料类别', 'cat'), ('物料子类别', 'subcat'), ('制造商', 'manufacturer'),
    ('物料来源', 'source'), ('提前期', 'lead_time'), ('主计量单位', 'unit'),
    ('标准价格', 'price'), ('创建人', 'creator'), ('创建日期', 'create_date'),
    ('最近修改人', 'modifier'), ('最近修改日期', 'modify_date'), ('备注', 'remark')
]
DIFF_FIELDS = {'name','desc','cat','subcat','manufacturer','source','lead_time','unit','price'}

def make_sheet(wb_out, title, data):
    ws_out = wb_out.create_sheet(title)
    # Header
    for ci, (fname, _) in enumerate(FIELDS, 1):
        c = ws_out.cell(row=1, column=ci, value=fname)
        c.font = Font(bold=True, size=10)
        c.alignment = CENTER
        c.border = BORDER

    ri = 2
    for idx, (a, b, reason) in enumerate(data, 1):
        for label, mat, fill in [(f'A-{idx}', a, BLUE), (f'B-{idx}', b, ORANGE)]:
            for ci, (_, key) in enumerate(FIELDS, 1):
                if key == 'label':
                    val = label
                elif key == 'remark':
                    val = reason
                else:
                    val = mat.get(key, '')
                    if isinstance(val, float) and val == int(val):
                        val = int(val)
                cell = ws_out.cell(row=ri, column=ci, value=val)
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = WRAP if key == 'desc' else CENTER
                if key in DIFF_FIELDS:
                    va = str(a.get(key,'') or '').strip()
                    vb = str(b.get(key,'') or '').strip()
                    cell.font = RED if va != vb else BLACK
                else:
                    cell.font = BLACK
            ri += 1

    widths = [8,16,18,50,12,14,20,10,8,8,10,10,14,10,14,40]
    for i, w in enumerate(widths):
        ws_out.column_dimensions[get_column_letter(i+1)].width = w
    ws_out.freeze_panes = 'C2'
    if ri > 2:
        ws_out.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}{ri-1}"
    return ws_out

wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)
make_sheet(wb_out, '确认重复', confirmed)
make_sheet(wb_out, '待人工确认', review)
wb_out.save(OUTPUT)
print(f"\nSaved to {OUTPUT}")
print(f"Total time: {time.time()-t0:.1f}s")
