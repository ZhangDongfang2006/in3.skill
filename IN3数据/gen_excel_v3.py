#!/usr/bin/env python3
"""
IN3 物料重复检查 - 生成 Excel 报告 v3
3 sheets: 确认重复+高度可疑、非重复、待人工确认
"""
import json
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print(f"[{time.strftime('%H:%M:%S')}] 生成 Excel 报告...", flush=True)

# 加载分类结果
with open('confirmed_dup_v3.json', 'r') as f:
    confirmed_dup = json.load(f)
with open('likely_dup_v3.json', 'r') as f:
    likely_dup = json.load(f)
with open('not_dup_v3.json', 'r') as f:
    not_dup = json.load(f)
with open('pending_v3.json', 'r') as f:
    pending = json.load(f)
with open('classification_summary_v3.json', 'r') as f:
    summary = json.load(f)

# 合并确认重复和高度可疑为一个 sheet
all_dup = confirmed_dup + likely_dup

print(f"确认重复+高度可疑: {len(all_dup)}", flush=True)
print(f"非重复: {len(not_dup)}", flush=True)
print(f"待人工确认: {len(pending)}", flush=True)

# ====== Excel 生成 ======
OUTPUT_FILE = '可疑重复物料_已验证_v3.xlsx'
wb = Workbook()

BLUE_FILL = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
RED_FONT = Font(color='FF0000')
NORMAL_FONT = Font(color='000000')
HEADER_FONT = Font(color='000000', bold=True)
HEADER_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# 字段定义
FIELDS = [
    ('物料编号', 'id'), ('物料名称', 'name'), ('物料描述', 'desc'),
    ('物料类别', 'category'), ('物料子类别', 'subcategory'),
    ('制造商', 'manufacturer'), ('物料来源', 'source'),
    ('主计量单位', 'unit'),
]

NO_RED_FIELDS = {'物料编号'}

def write_sheet(wb, title, pairs):
    """写入一个 sheet"""
    if len(pairs) == 0:
        ws = wb.create_sheet(title[:31])
        ws.append(['无数据'])
        return
    
    ws = wb.create_sheet(title[:31])
    
    # 标题行
    ws.cell(row=1, column=1, value=f'{title}（共 {len(pairs)} 对）')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(FIELDS)*2+1)
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    
    # 表头行
    headers = []
    for prefix in ['A', 'B']:
        for label, _ in FIELDS:
            headers.append(f'{prefix}-{label}')
    headers.append('差异/备注')
    
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = thin_border
        c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
    
    # 数据行
    for i, p in enumerate(pairs):
        row = 3 + i
        # 交替 A/B 蓝橙背景
        fill_a = BLUE_FILL
        fill_b = ORANGE_FILL
        a_vals = [p['a'].get(k, '') for _, k in FIELDS]
        b_vals = [p['b'].get(k, '') for _, k in FIELDS]
        
        # A 行
        all_vals = a_vals + b_vals + [p['reason']]
        for col, val in enumerate(all_vals, 1):
            c = ws.cell(row=row, column=col, value=val if val else '')
            c.fill = fill_a if col <= len(FIELDS) else (fill_b if col <= len(FIELDS)*2 else PatternFill())
            c.border = thin_border
            c.alignment = Alignment(wrap_text=True, vertical='top')
            
            # 差异标红
            if col <= len(FIELDS):
                label = FIELDS[col-1][0]
                if label not in NO_RED_FIELDS:
                    b_col_idx = col  # B 的对应列 = A列 + len(FIELDS)
                    if a_vals[col-1] and b_vals[col-1] and str(a_vals[col-1]) != str(b_vals[col-1]):
                        c.font = RED_FONT
                    else:
                        c.font = NORMAL_FONT
                else:
                    c.font = NORMAL_FONT
            elif col <= len(FIELDS) * 2:
                a_col_idx = (col - 1) % len(FIELDS)
                label = FIELDS[a_col_idx][0]
                if label not in NO_RED_FIELDS:
                    if a_vals[a_col_idx] and b_vals[a_col_idx] and str(a_vals[a_col_idx]) != str(b_vals[a_col_idx]):
                        c.font = RED_FONT
                    else:
                        c.font = NORMAL_FONT
                else:
                    c.font = NORMAL_FONT
            else:
                c.font = NORMAL_FONT
    
    # 列宽
    for col in range(1, len(headers) + 1):
        field_idx = (col - 1) % len(FIELDS)
        label = FIELDS[field_idx][0]
        if label == '物料描述':
            ws.column_dimensions[get_column_letter(col)].width = 50
        elif label == '物料编号':
            ws.column_dimensions[get_column_letter(col)].width = 16
        elif label in ('物料名称', '制造商'):
            ws.column_dimensions[get_column_letter(col)].width = 25
        elif label == '物料子类别':
            ws.column_dimensions[get_column_letter(col)].width = 18
        elif label == '差异/备注':
            ws.column_dimensions[get_column_letter(col)].width = 40
        else:
            ws.column_dimensions[get_column_letter(col)].width = 14
    
    # 冻结窗格（冻结前2行前2列）
    ws.freeze_panes = 'C3'
    
    # 自动筛选
    ws.auto_filter.ref = f'A2:{get_column_letter(len(headers))}{len(pairs)+2}'

# ====== Sheet 1: 总览 =====
ws = wb.active
ws.title = '总览'
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 18

ws.cell(row=1, column=1, value='IN3 物料重复检查报告').font = Font(bold=True, size=14)
ws.cell(row=3, column=1, value='数据来源')
ws.cell(row=3, column=2, value=f'物料主数据导出 ({summary["timestamp"][:10]})')
ws.cell(row=4, column=1, value='物料总数')
ws.cell(row=4, column=2, value=summary['total_materials'])
ws.cell(row=5, column=1, value='检查范围')
ws.cell(row=5, column=2, value='非成品物料')
ws.cell(row=6, column=1, value='候选配对数')
ws.cell(row=6, column=2, value=summary['total_candidates'])

ws.cell(row=8, column=1, value='分类结果').font = Font(bold=True, size=12)
ws.cell(row=9, column=1, value='✅ 确认重复（名称+描述+制造商均相同）')
ws.cell(row=9, column=2, value=len(confirmed_dup))
ws.cell(row=10, column=1, value='⚠️  高度可疑（描述相似度≥0.95或制造商不同）')
ws.cell(row=10, column=2, value=len(likely_dup))
ws.cell(row=11, column=1, value='❌ 非重复（规则排除）')
ws.cell(row=11, column=2, value=len(not_dup))
ws.cell(row=12, column=1, value='❓ 待人工确认（描述相似度0.85-0.95）')
ws.cell(row=12, column=2, value=len(pending))

ws.cell(row=14, column=1, value='说明').font = Font(bold=True, size=12)
notes = [
    '• 蓝色背景 = A 物料，橙色背景 = B 物料',
    '• 红色字体 = A/B 之间存在差异的字段',
    '• 物料编号始终黑色（不标红）',
    '• 排除规则（20条）：甲供件、制造商一有一无、安装方向、断路器分断能力等级、分断能力kA、脱扣曲线、极数、颜色、漏电保护类型、带附件vs不带、互感器变比/窗口/精度、电线类型、保护类型、螺纹方向、版本号、脱扣方式、脱扣单元、配件差异、过电压保护器型号、型号系列、额定电流、保护等级IP、尺寸',
    '• 建议优先审核「确认重复+高度可疑」sheet',
]
for i, note in enumerate(notes):
    ws.cell(row=15+i, column=1, value=note)

# Sheet 2: 确认重复+高度可疑
write_sheet(wb, '确认重复+高度可疑', all_dup)

# Sheet 3: 非重复
write_sheet(wb, '非重复', not_dup)

# Sheet 4: 待人工确认
write_sheet(wb, '待人工确认', pending)

wb.save(OUTPUT_FILE)
print(f"\n[{time.strftime('%H:%M:%S')}] ✅ Excel 已保存: {OUTPUT_FILE}", flush=True)
print(f"  📊 确认重复: {len(confirmed_dup)} 对", flush=True)
print(f"  ⚠️  高度可疑: {len(likely_dup)} 对", flush=True)
print(f"  ❌ 非重复: {len(not_dup)} 对", flush=True)
print(f"  ❓ 待人工确认: {len(pending)} 对", flush=True)
