#!/usr/bin/env python3
"""
IN3 物料重复检查 v5 - 基于描述相似度筛选 + 规则分类
"""
import polars as pl
from collections import defaultdict
import re, time, sys, math

INPUT_FILE = '/Users/zhangdongfang/.openclaw/workspace-in3bot/IN3数据/物料主数据导出结果-20260905161742956.xlsx'
OUTPUT_FILE = '/Users/zhangdongfang/.openclaw/workspace-in3bot/IN3数据/可疑重复物料_已验证_v5.xlsx'

KEY_COLS = ['物料ID', '*物料编号', '*物料名称', '*物料描述', '物料类别',
            '物料子类别', '制造商', '*物料来源', '*提前期', '主计量单位',
            '标准价格', '创建人', '创建日期', '最近修改人', '最近修改日期',
            '*物料类别编码']

print("Step 1: Loading data...", flush=True)
t0 = time.time()
df = pl.read_excel(INPUT_FILE, sheet_name='物料主数据', columns=KEY_COLS)
print(f"  Loaded {df.height} rows in {time.time()-t0:.1f}s", flush=True)

materials = []
for i in range(df.height):
    row = df.row(i, named=True)
    cc = str(row.get('*物料类别编码') or '')
    cat = str(row.get('物料类别') or '')
    subcat = str(row.get('物料子类别') or '')
    if cc in ('CP', 'WGCT') or '甲供' in cat or '甲供' in subcat:
        continue
    desc = str(row.get('*物料描述') or '').strip()
    if not desc:
        continue
    materials.append({
        'code': str(row.get('*物料编号') or ''),
        'name': str(row.get('*物料名称') or ''),
        'desc': desc,
        'category': cat,
        'subcategory': subcat,
        'manufacturer': str(row.get('制造商') or ''),
        'source': str(row.get('*物料来源') or ''),
        'lead_time': row.get('*提前期'),
        'unit': str(row.get('主计量单位') or ''),
        'price': row.get('标准价格'),
        'creator': str(row.get('创建人') or ''),
        'create_date': row.get('创建日期'),
        'modifier': str(row.get('最近修改人') or ''),
        'modify_date': row.get('最近修改日期'),
    })

print(f"  After exclusion: {len(materials)} materials", flush=True)

# ---- Similarity matching ----
def normalize(s):
    if not s: return ''
    s = re.sub(r'\s+', ' ', str(s).strip().lower())
    # unify common format differences
    for orig, repl in [('(','('),(')',')'),(',',','),('。','.'),(':',':'),(';',';'),
                       ('i','i'),('I','I'),('v','v'),('V','V'),('x','x'),('X','X'),
                       ('l','l'),('L','L'),('+','+'),('P','P'),('N','N')]:
        s = s.replace(orig, repl)
    return s

def jaccard_similarity(a, b):
    """Jaccard similarity of character sets"""
    if not a or not b: return 0.0
    sa, sb = set(a), set(b)
    intersection = sa & sb
    union = sa | sb
    if not union: return 0.0
    return len(intersection) / len(union)

def seq_similarity(a, b):
    """Sequence-based similarity using longest common subsequence ratio"""
    if not a or not b: return 0.0
    # Remove spaces for comparison
    a_ns = a.replace(' ', '')
    b_ns = b.replace(' ', '')
    if a_ns == b_ns: return 1.0
    if not a_ns or not b_ns: return 0.0
    # Use a simpler approach: trigram similarity
    def trigrams(s):
        return set(s[i:i+3] for i in range(len(s)-2))
    ta, tb = trigrams(a_ns), trigrams(b_ns)
    if not ta or not tb: return 0.0
    return len(ta & tb) / max(len(ta), len(tb))

# Pre-compute normalized descriptions and group by normalized form for exact matches
print("\nStep 2: Finding candidate pairs...", flush=True)
t1 = time.time()

norm_descs = [normalize(m['desc']) for m in materials]
# Also compute no-space versions
no_space_descs = [re.sub(r'\s+', '', nd) for nd in norm_descs]

# Group by no-space normalized desc (catches format-only diffs immediately)
exact_groups = defaultdict(list)
for idx, nsd in enumerate(no_space_descs):
    exact_groups[nsd].append(idx)

# These are definitely candidates (exact match after normalization)
candidate_pairs = set()
for nsd, indices in exact_groups.items():
    if len(indices) >= 2:
        for i in range(len(indices)):
            for j in range(i+1, len(indices)):
                candidate_pairs.add((min(indices[i], indices[j]), max(indices[i], indices[j])))

print(f"  Exact normalized matches: {len(candidate_pairs)} pairs", flush=True)

# Now find near-matches using description prefix/structure grouping
# Group by first N characters of normalized desc to limit comparisons
PREFIX_LEN = 10
prefix_groups = defaultdict(list)
for idx, nd in enumerate(norm_descs):
    if len(nd) >= PREFIX_LEN:
        prefix = nd[:PREFIX_LEN]
    else:
        prefix = nd
    prefix_groups[prefix].append(idx)

# Compare within each prefix group using trigram similarity
SIMILARITY_THRESHOLD = 0.75
compared = 0
for prefix, indices in prefix_groups.items():
    if len(indices) < 2:
        continue
    # Sort by length to optimize
    indices.sort(key=lambda i: len(norm_descs[i]))
    for i in range(len(indices)):
        for j in range(i+1, len(indices)):
            a_idx, b_idx = indices[i], indices[j]
            pair = (a_idx, b_idx)
            if pair in candidate_pairs:
                continue
            # Quick length check
            len_a, len_b = len(no_space_descs[a_idx]), len(no_space_descs[b_idx])
            if len_a == 0 or len_b == 0:
                continue
            ratio = min(len_a, len_b) / max(len_a, len_b)
            if ratio < 0.6:
                continue
            compared += 1
            sim = seq_similarity(norm_descs[a_idx], norm_descs[b_idx])
            if sim >= SIMILARITY_THRESHOLD:
                candidate_pairs.add(pair)

print(f"  Compared {compared} near-match candidates", flush=True)
print(f"  Total candidate pairs: {len(candidate_pairs)}", flush=True)
print(f"  Found in {time.time()-t1:.1f}s", flush=True)

# ---- Classification ----
def has_chinese(s):
    return bool(re.search(r'[\u4e00-\u9fff]', s))

def manufacturers_different(mfr_a, mfr_b):
    a, b = str(mfr_a).strip(), str(mfr_b).strip()
    if not a or not b or a == b: return False
    if a in b or b in a: return False
    for length in range(min(len(a), len(b)), 2, -1):
        for i in range(len(a) - length + 1):
            substr = a[i:i+length]
            if substr in b and has_chinese(substr):
                return False
    return True

def extract_numbers(s):
    return re.findall(r'\d+\.?\d*', str(s))

def is_format_only_diff(da, db):
    na, nb = normalize(da), normalize(db)
    if na == nb: return True
    return re.sub(r'\s+', '', na) == re.sub(r'\s+', '', nb)

def check_1n_vs_1pn(da, db):
    da_ns = re.sub(r'\s+', '', str(da))
    db_ns = re.sub(r'\s+', '', str(db))
    da_r = da_ns.replace('1P+N','1N').replace('1p+n','1n').replace('1P+N','1N')
    db_r = db_ns.replace('1P+N','1N').replace('1p+n','1n').replace('1P+N','1N')
    return da_r == db_r and ('1P+N' in da or '1P+N' in db)

def classify_pair(a, b):
    da, db = str(a['desc']).strip(), str(b['desc']).strip()
    ma, mb = str(a['manufacturer']).strip(), str(b['manufacturer']).strip()
    sa, sb = str(a['source']).strip(), str(b['source']).strip()

    # Manufacturer rules
    if bool(ma) != bool(mb):
        return ('not_dup', '制造商一有一无')
    if ma and mb and manufacturers_different(ma, mb):
        return ('not_dup', f'制造商完全不同:{ma} vs {mb}')

    # Source different
    if sa and sb and sa != sb:
        return ('not_dup', f'物料来源不同:{sa} vs {sb}')

    # Format-only diff -> confirmed
    if is_format_only_diff(da, db):
        return ('confirmed', '描述仅格式/表述差异(空格/全半角/标点等)')

    # 1N vs 1P+N
    if check_1n_vs_1pn(da, db):
        return ('confirmed', '1N vs 1P+N(施耐德同一极数不同表述)')

    na, nb = normalize(da), normalize(db)
    nums_a, nums_b = extract_numbers(da), extract_numbers(db)

    # Number differences
    if nums_a != nums_b:
        na_nn = re.sub(r'\d+\.?\d*', 'N', na)
        nb_nn = re.sub(r'\d+\.?\d*', 'N', nb)
        if na_nn == nb_nn and len(nums_a) == len(nums_b):
            diffs = [f"{x} vs {y}" for x, y in zip(nums_a, nums_b) if x != y]
            return ('not_dup', f'型号中数值参数不同:{", ".join(diffs[:3])}')
        else:
            return ('not_dup', '描述结构不同且数值参数不同')

    # Color difference in description (single char colors)
    color_chars = ['白','黄','红','绿','蓝','黑','灰','橙','棕','紫','粉']
    for c in color_chars:
        if c in na and c not in nb:
            # Check if another color char from nb is not in na
            for c2 in color_chars:
                if c2 != c and c2 in nb and c2 not in na:
                    return ('not_dup', f'颜色不同:{c} vs {c2}')

    # Keyword-based rules
    diff_kw = [
        ('左操','右操'),('平进平出','平进侧出'),('上进','下进'),('上进线','下进线'),
        ('红色','灰色'),('红色','棕色'),('灰色','棕色'),('红色','黑色'),('白色','灰色'),
        ('蓝色','黑色'),('红色','白色'),('灰色','白色'),('橙色','灰色'),('绿色','红色'),
        ('3P','4P'),('2P','3P'),('1P','2P'),('1P','3P'),
        ('B型','C型'),('C型','D型'),('B型','D型'),('B曲线','C曲线'),('C曲线','D曲线'),
        ('TMD','MA'),('热磁','电磁'),('TM','MA'),('LSI','TMA'),('LSI','MIC'),('LSI','LSIG'),
        ('A型','AC型'),
        ('铜','铝'),('紫铜','黄铜'),('304','316'),('不锈钢','镀锌'),
        ('BV','BVR'),('BVR','RV'),('ZR-YJV','YJV'),('YJV','VV'),('NH-YJV','YJV'),
        ('WDZ-YJY','YJV'),('WDZN-YJY','YJV'),
        ('正牙','反牙'),('右旋','左旋'),
        ('变压器保护','线路保护'),('电容器保护','线路保护'),('电动机保护','线路保护'),
        ('1A','1B'),('V1','V2'),
        ('带失压','不带失压'),('带辅助','不带辅助'),('带报警','不带报警'),
        ('带门框','不带门框'),('带电磁锁','不带电磁锁'),
        ('带欠压','无欠压'),
        ('CM3E','CM3'),('CDM3','CDM3E'),
        ('抽屉式','固定式'),('不加热缩','加热缩'),
        ('单排','双排'),('A相','B相'),('B相','C相'),('A相','C相'),
        ('30i','30ii'),('i ','ii '),
        ('左伸出','右伸出'),('不带欠压','智能脱扣器欠压'),
        ('SY','SL'),('SY ','SL '),
        ('变频','铝壳'),
        ('电动ac220v','手动配分励ac220v'),('电动ac220v','手动ac220v'),
        ('三开三闭','二开二闭'),('三开三闭','四开四闭'),('三开三闭','五开五闭'),
        ('4开4闭','2开2闭'),('4开4闭','3开3闭'),
    ]
    for kw1, kw2 in diff_kw:
        k1, k2 = kw1.lower(), kw2.lower()
        if (k1 in na and k1 not in nb and k2 in nb) or (k2 in na and k2 not in nb and k1 in nb):
            return ('not_dup', f'{kw1} vs {kw2}')

    # Check single-letter trip curve in model number: e.g. D32 vs C32, D40 vs C40
    # Pattern: [BCD] followed by number (amperage)
    curve_nums_a = re.findall(r'[BCD](\d+)', da)
    curve_nums_b = re.findall(r'[BCD](\d+)', db)
    if curve_nums_a and curve_nums_b:
        # Check if any letter prefix differs with same number
        for ca in re.findall(r'([BCD])\d+', da):
            for cb in re.findall(r'([BCD])\d+', db):
                if ca != cb:
                    return ('not_dup', f'微型断路器脱扣曲线不同：{ca}型 vs {cb}型')

    # Check 3P vs 3N (different pole notations)
    da_poles = re.sub(r'3N', '3P', na)
    db_poles = re.sub(r'3N', '3P', nb)
    if da_poles != db_poles and da_poles.replace('3P','3N') == db_poles:
        return ('not_dup', '3P vs 3N 极数不同')

    # Check -J suffix (with/without accessory)
    if na.rstrip('-j') == nb.rstrip('-j') and na != nb:
        return ('not_dup', '-J后缀差异（配件不同）')

    # Check model suffix like iTR326H vs iTR326A (different trip unit)
    model_a = re.findall(r'([A-Za-z]+\d+[A-Z])(?:\s|$)', da)
    model_b = re.findall(r'([A-Za-z]+\d+[A-Z])(?:\s|$)', db)
    if model_a and model_b:
        for ma_s, mb_s in zip(model_a, model_b):
            if ma_s[:-1] == mb_s[:-1] and ma_s[-1] != mb_s[-1]:
                return ('not_dup', f'脱扣单元后缀不同：{ma_s} vs {mb_s}')

    # Breaker shell letter
    bp = r'([A-Za-z]+\d+)([FCHLMNS])(?:\s|$|[^A-Za-z0-9])'
    ma_match, mb_match = re.findall(bp, da), re.findall(bp, db)
    if ma_match and mb_match:
        for (ba, la), (bb, lb) in zip(ma_match, mb_match):
            if ba.lower() == bb.lower() and la != lb:
                return ('not_dup', f'断路器分断能力等级不同:{ba}{la} vs {bb}{lb}')

    # If normalized descriptions are same -> confirmed
    # But check if name suggests different product
    if na == nb:
        if len(da) < 10:
            name_a, name_b = normalize(str(a['name'])), normalize(str(b['name']))
            if name_a != name_b:
                return ('review', f'描述极短({da})且名称不同：{a["name"]} vs {b["name"]}')
        return ('confirmed', '描述完全相同')

    # Check 0.5S vs 0.5级 (precision class notation)
    na_s = re.sub(r'0\.5s', '0.5级', na)
    nb_s = re.sub(r'0\.5s', '0.5级', nb)
    if na_s == nb_s and na != nb:
        return ('confirmed', '0.5S vs 0.5级（精度等级不同表述）')
    # Check 0.5S vs 0.5 (precision class with/without S)
    na_s2 = re.sub(r'0\.5s\b', '0.5', na)
    nb_s2 = re.sub(r'0\.5s\b', '0.5', nb)
    na_s3 = re.sub(r'(?<!\d)0\.5\b(?!s)', '0.5S', na)
    nb_s3 = re.sub(r'(?<!\d)0\.5\b(?!s)', '0.5S', nb)
    if na_s2 == nb_s2 and na != nb:
        return ('confirmed', '0.5S vs 0.5（精度等级表述差异）')
    if na_s3 == nb_s3 and na != nb:
        return ('confirmed', '0.5S vs 0.5（精度等级表述差异）')

    # Check mm presence difference
    da_nm = na.replace('mm', '')
    db_nm = nb.replace('mm', '')
    if da_nm == db_nm and ('mm' in na) != ('mm' in nb):
        return ('confirmed', '仅单位标注差异（mm有无）')

    # 按图 vs 见图 = confirmed duplicate
    if na.replace('按图','见图') == nb or na == nb.replace('按图','见图'):
        return ('confirmed', '按图 vs 见图（表述差异）')

    # 发黑 vs 无发黑 = surface treatment difference
    if ('发黑' in na and '发黑' not in nb) or ('发黑' in nb and '发黑' not in na):
        # Check if the rest is the same
        na_base = na.replace('发黑','')
        nb_base = nb.replace('发黑','')
        if na_base == nb_base:
            return ('not_dup', '表面处理不同：有发黑 vs 无发黑')

    # AC型 vs 无AC型 = leakage type difference
    if 'ac型' in na and 'ac型' not in nb:
        na_base = na.replace('ac型','')
        if na_base == nb:
            return ('not_dup', '漏电保护类型不同：AC型 vs 无标注')
    if 'ac型' in nb and 'ac型' not in na:
        nb_base = nb.replace('ac型','')
        if nb_base == na:
            return ('not_dup', '漏电保护类型不同：AC型 vs 无标注')

    # 实厚 vs 无实厚
    if na.replace('实厚','') == nb and na != nb:
        return ('not_dup', '规格标注不同：实厚 vs 标称厚度')

    # 配置相间隔板 vs 无
    if na.replace('配置相间隔板','') == nb and na != nb:
        return ('not_dup', '配件差异：带相间隔板 vs 不带')

    # 法兰式 vs 无
    if '法兰式' in na and '法兰式' not in nb:
        if na.replace('法兰式','') == nb:
            return ('not_dup', '连接方式不同：法兰式 vs 非法兰式')

    # 通/孔 vs 无
    if '通/孔' in na and '通/孔' not in nb:
        if na.replace('通/孔','') == nb:
            return ('not_dup', '加工方式不同：通孔 vs 无通孔')

    # Character overlap
    da_ns = da.replace(' ','').replace('\u3000','')
    db_ns = db.replace(' ','').replace('\u3000','')
    sa_set, sb_set = set(da_ns), set(db_ns)
    if sa_set and sb_set:
        overlap = len(sa_set & sb_set) / max(len(sa_set), len(sb_set))
        if overlap < 0.7:
            return ('not_dup', f'描述差异较大(相似度{overlap:.0%})')

    return ('review', '描述有差异但无法自动判断')

print("\nStep 3: Classifying pairs...", flush=True)
t2 = time.time()
confirmed, not_dup, review = [], [], []

for a_idx, b_idx in candidate_pairs:
    a, b = materials[a_idx], materials[b_idx]
    result, reason = classify_pair(a, b)
    {'confirmed': confirmed, 'not_dup': not_dup, 'review': review}[result].append((a, b, reason))

print(f"  Classified in {time.time()-t2:.1f}s", flush=True)
print(f"  Confirmed: {len(confirmed)}", flush=True)
print(f"  Not duplicate: {len(not_dup)}", flush=True)
print(f"  Need review: {len(review)}", flush=True)

# ---- Generate Excel ----
print("\nStep 4: Generating Excel...", flush=True)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_COLS = [
    ('标记','label'),('物料编号','code'),('物料名称','name'),('物料描述','desc'),
    ('物料类别','category'),('物料子类别','subcategory'),('制造商','manufacturer'),
    ('物料来源','source'),('主计量单位','unit'),('提前期','lead_time'),
    ('标准价格','price'),('创建人','creator'),('创建日期','create_date'),
    ('最近修改人','modifier'),('最近修改日期','modify_date'),('备注','reason'),
]
NO_RED = {'标记','物料编号','创建人','创建日期','最近修改人','最近修改日期','备注'}
RED = {'物料名称','物料描述','物料类别','物料子类别','制造商','物料来源','主计量单位','提前期','标准价格'}

blue_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
orange_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
red_font = Font(color='FF0000')
normal_font = Font()
header_font = Font(bold=True)
thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

def write_sheet(ws, pairs):
    if not pairs:
        ws.cell(row=1, column=1, value='无数据')
        return
    vis = []
    for cn, ck in OUTPUT_COLS:
        if ck in ('label','reason'):
            vis.append((cn, ck)); continue
        has_data = False
        for a, b, _ in pairs[:100]:  # check first 100
            va = str(a.get(ck,'') or '').strip()
            vb = str(b.get(ck,'') or '').strip()
            if va or vb:
                has_data = True; break
        if has_data:
            vis.append((cn, ck))

    ws.cell(row=1, column=1, value='序号').font = header_font
    ws.cell(row=1, column=1).border = thin_border
    ws.cell(row=1, column=1).fill = header_fill
    for ci, (cn, _) in enumerate(vis, 2):
        c = ws.cell(row=1, column=ci, value=cn)
        c.font = header_font; c.border = thin_border; c.fill = header_fill

    ri = 2
    for pn, (a, b, reason) in enumerate(pairs, 1):
        for is_a in [True, False]:
            m = a if is_a else b
            ws.cell(row=ri, column=1, value=pn).border = thin_border
            ws.cell(row=ri, column=1).fill = blue_fill if is_a else orange_fill
            ci = 2
            for cn, ck in vis:
                if ck == 'label': val = f'A-{pn}' if is_a else f'B-{pn}'
                elif ck == 'reason': val = reason
                else:
                    val = m.get(ck, ''); val = '' if val is None else val
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = thin_border
                cell.fill = blue_fill if is_a else orange_fill
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                if cn in RED:
                    va = str(a.get(ck,'') or '').strip()
                    vb = str(b.get(ck,'') or '').strip()
                    cell.font = red_font if va != vb else normal_font
                else:
                    cell.font = normal_font
                ci += 1
            ri += 1

    ws.column_dimensions['A'].width = 8
    for ci, (cn, _) in enumerate(vis, 2):
        cl = get_column_letter(ci)
        if cn == '物料描述': ws.column_dimensions[cl].width = 50
        elif cn == '备注': ws.column_dimensions[cl].width = 40
        elif cn in ('物料编号','物料名称'): ws.column_dimensions[cl].width = 18
        elif cn == '制造商': ws.column_dimensions[cl].width = 16
        else: ws.column_dimensions[cl].width = 14
    ws.freeze_panes = f'{get_column_letter(3)}3'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(vis)+1)}{ri-1}'

out_wb = openpyxl.Workbook()
ws1 = out_wb.active; ws1.title = '确认重复'; write_sheet(ws1, confirmed)
ws2 = out_wb.create_sheet('非重复'); write_sheet(ws2, not_dup)
ws3 = out_wb.create_sheet('待人工确认'); write_sheet(ws3, review)
out_wb.save(OUTPUT_FILE)

print(f"\nDone! Total: {time.time()-t0:.1f}s", flush=True)
print(f"=== SUMMARY ===", flush=True)
print(f"物料总数: {len(materials)}", flush=True)
print(f"候选配对数: {len(candidate_pairs)}", flush=True)
print(f"确认重复: {len(confirmed)} 对", flush=True)
print(f"非重复: {len(not_dup)} 对", flush=True)
print(f"待人工确认: {len(review)} 对", flush=True)
