#!/usr/bin/env python3
"""
IN3 物料重复检查 - 生成 Excel 报告 v2
阈值: 描述相似度 >= 0.92 为可疑重复
"""
import json, re, sys
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

print(f"[{time.strftime('%H:%M:%S')}] 生成 Excel 报告...", flush=True)

with open('confirmed_dup_pairs_v2.json', 'r') as f:
    all_pairs = json.load(f)
with open('not_dup_pairs_v2.json', 'r') as f:
    not_dup_pairs = json.load(f)

# 重新分类
confirmed_dup = []     # 名称+描述完全相同，制造商相同
suspected_dup = []     # 名称+描述完全相同，制造商不同 OR 描述相似度>=0.95
needs_review = []      # 描述相似度 0.92-0.95
false_positive = []    # 描述相似度 < 0.92（误报）

for p in all_pairs:
    a, b = p['a'], p['b']
    reason = p['reason']
    
    if '名称描述制造商均相同' in reason:
        confirmed_dup.append(p)
    elif '名称描述相同制造商不同' in reason:
        suspected_dup.append({**p, 'reason': '名称描述相同，制造商不同'})
    elif '描述高度相似' in reason:
        ratio = SequenceMatcher(None, a['desc'], b['desc']).ratio()
        if ratio >= 0.95:
            # 检查差异是否只是数值
            suspected_dup.append({**p, 'reason': f'描述高度相似({ratio:.2f})'})
        elif ratio >= 0.92:
            needs_review.append({**p, 'reason': f'描述部分相似({ratio:.2f})'})
        else:
            false_positive.append({**p, 'reason': f'相似度偏低({ratio:.2f})'})

print(f"确认重复: {len(confirmed_dup)}", flush=True)
print(f"高度可疑: {len(suspected_dup)}", flush=True)
print(f"待人工确认: {len(needs_review)}", flush=True)
print(f"误报（相似度<0.92）: {len(false_positive)}", flush=True)
print(f"非重复（规则排除）: {len(not_dup_pairs)}", flush=True)

# ====== 生成 Excel ======
OUTPUT_FILE = '可疑重复物料_已验证_v2.xlsx'
wb = Workbook()

BLUE_FILL = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
RED_FONT = Font(color='FF0000')
NORMAL_FONT = Font(color='000000')
HEADER_FONT = Font(color='000000', bold=True)
HEADER_FILL = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

NO_RED_FIELDS = {'物料编号', '创建人', '创建日期', '修改人', '修改日期'}
FIELDS = [
    ('物料编号', 'id'), ('物料名称', 'name'), ('物料描述', 'desc'),
    ('物料类别', 'category'), ('物料子类别', 'subcategory'),
    ('制造商', 'manufacturer'), ('物料来源', 'source'),
    ('主计量单位', 'unit'),
]

def write_sheet(wb, title, pairs, category_label):
    if len(pairs) == 0:
        ws = wb.create_sheet(title[:31])
        ws.append([f'{title} - 无数据'])
        return
    ws = wb.create_sheet(title[:31])
    
    ws.append([f'{title}（{len(pairs)} 对）', '', ''])
    ws.merge_cells('A1:C1')
    
    # 表头
    headers = []
    for prefix in ['A', 'B']:
        for label, _ in FIELDS:
            headers.append(f'{prefix}-{label}')
    headers.append('差异/备注')
    
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = thin_border
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    
    for i, p in enumerate(pairs):
        row = 3 + i
        fill = BLUE_FILL if i % 2 == 0 else ORANGE_FILL
        a_vals = [p['a'].get(k, '') for _, k in FIELDS]
        b_vals = [p['b'].get(k, '') for _, k in FIELDS]
        
        all_vals = a_vals + b_vals + [p['reason']]
        for col, val in enumerate(all_vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill
            c.border = thin_border
            c.alignment = Alignment(wrap_text=True)
            
            # A/B 对比标红
            if col <= len(FIELDS):
                label = FIELDS[col-1][0]
                if label not in NO_RED_FIELDS and a_vals[col-1] != b_vals[col-1] and a_vals[col-1] and b_vals[col-1]:
                    c.font = RED_FONT
                else:
                    c.font = NORMAL_FONT
            elif col <= len(FIELDS) * 2:
                label = FIELDS[(col-1) % len(FIELDS)][0]
                a_idx = (col-1) % len(FIELDS)
                if label not in NO_RED_FIELDS and a_vals[a_idx] != b_vals[a_idx] and a_vals[a_idx] and b_vals[a_idx]:
                    c.font = RED_FONT
                else:
                    c.font = NORMAL_FONT
            else:
                c.font = NORMAL_FONT
    
    # 列宽
    for col in range(1, len(headers) + 1):
        label = FIELDS[(col-1) % len(FIELDS)][0]
        if label == '物料描述':
            ws.column_dimensions[get_column_letter(col)].width = 50
        elif label == '物料编号':
            ws.column_dimensions[get_column_letter(col)].width = 16
        elif label in ('物料名称', '制造商', '差异/备注'):
            ws.column_dimensions[get_column_letter(col)].width = 25
        else:
            ws.column_dimensions[get_column_letter(col)].width = 15

# Sheet 1: 总览
ws = wb.active
ws.title = '总览'
ws.append(['IN3 物料重复检查报告'])
ws.merge_cells('A1')
ws['A1'].font = Font(bold=True, size=14)
ws.append([])
ts = time.strftime('%Y-%m-%d %H:%M')
ws.append(['数据来源', f'物料主数据导出 ({ts(\"%Y-%m-%d %H:%M\")})'])
ws.append(['检查范围', '非成品物料（排除成品柜），共 18,049 条'])
ws.append([])
ws.append(['分类', '数量'])
ws.append(['确认重复（名称+描述+制造商均相同）', len(confirmed_dup)])
ws.append(['高度可疑（描述相似度>=0.95）', len(suspected_dup)])
ws.append(['待人工确认（描述相似度0.92-0.95）', len(needs_review)])
ws.append(['非重复（规则排除）', len(not_dup_pairs)])
ws.append(['误报（描述相似度<0.92）', len(false_positive)])
ws.append([])
ws.append(['说明', ''])
ws.append(['', '• 蓝色/橙色背景区分 A/B 物料'])
ws.append(['', '• 红色字体 = A/B 之间存在差异的字段'])
ws.append(['', '• 物料编号、创建人/日期、修改人/日期 始终黑色'])
ws.append(['', '• 排除规则：甲供件、制造商一有一无、安装方向、分断能力、脱扣曲线、极数、颜色、漏电类型、带附件vs不带、互感器变比、电线类型、保护等级、螺纹方向、型号系列、额定电流'])

# 其他 sheets
write_sheet(wb, '确认重复', confirmed_dup, '确认重复')
write_sheet(wb, '高度可疑', suspected_dup, '高度可疑')
write_sheet(wb, '待人工确认', needs_review, '待人工确认')
write_sheet(wb, '非重复示例', not_dup_pairs[:200], '非重复')

wb.save(OUTPUT_FILE)
print(f"\n[{time.strftime('%H:%M:%S')}] ✅ Excel 已保存: {OUTPUT_FILE}", flush=True)
print(f"  📊 确认重复: {len(confirmed_dup)} 对", flush=True)
print(f"  ⚠️  高度可疑: {len(suspected_dup)} 对", flush=True)
print(f"  ❓ 待人工确认: {len(needs_review)} 对", flush=True)
print(f"  ❌ 非重复: {len(not_dup_pairs)} 对", flush=True)
print(f"  🔇 误报: {len(false_positive)} 对", flush=True)
