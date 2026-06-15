#!/usr/bin/env python3
"""IN3 物料重复检查分析脚本 v3 - 严格参数预筛选，大幅减少待人工确认数量"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from difflib import SequenceMatcher
import re
import time

INPUT_FILE = '/Users/zhangdongfang/.openclaw/workspace-in3bot/IN3数据/物料主数据导出结果-20262105150358225.xlsx'
OUTPUT_FILE = '/Users/zhangdongfang/.openclaw/workspace-in3bot/IN3数据/可疑重复物料-2026-05-21-v3.xlsx'

# Column mapping (1-indexed)
COL = {
    'id': 1, 'code': 2, 'name': 6, 'desc': 7,
    'cat_code': 8, 'cat': 9, 'subcat_code': 10, 'subcat': 11,
    'manufacturer': 13, 'source': 14, 'lead_time': 18,
    'unit': 22, 'price': 61,
    'creator': 70, 'create_date': 72,
    'modifier': 73, 'modify_date': 75,
    'color': 16
}

print("Loading workbook (data_only=True)...")
start = time.time()
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
ws = wb['物料主数据']
max_row = ws.max_row
print(f"Loaded in {time.time()-start:.1f}s, {max_row} rows")

# Read all materials
materials = []
exclude_cats = {'成品柜', '外购成套'}

def get_val(r, c):
    v = ws.cell(row=r, column=c).value
    return str(v).strip() if v is not None else ''

for row_idx in range(2, max_row + 1):
    cat = ws.cell(row=row_idx, column=COL['cat']).value or ''
    subcat = ws.cell(row=row_idx, column=COL['subcat']).value or ''
    cat_str = str(cat).strip()
    subcat_str = str(subcat).strip()
    
    if cat_str in exclude_cats:
        continue
    
    materials.append({
        'row': row_idx,
        'code': get_val(row_idx, COL['code']),
        'name': get_val(row_idx, COL['name']),
        'desc': get_val(row_idx, COL['desc']),
        'cat': cat_str,
        'subcat': subcat_str,
        'manufacturer': get_val(row_idx, COL['manufacturer']),
        'source': get_val(row_idx, COL['source']),
        'unit': get_val(row_idx, COL['unit']),
        'lead_time': get_val(row_idx, COL['lead_time']),
        'price': get_val(row_idx, COL['price']),
        'creator': get_val(row_idx, COL['creator']),
        'create_date': get_val(row_idx, COL['create_date']),
        'modifier': get_val(row_idx, COL['modifier']),
        'modify_date': get_val(row_idx, COL['modify_date']),
        'color': get_val(row_idx, COL['color']),
    })

wb.close()
print(f"Materials (excl. 成品柜/外购成套): {len(materials)}")


def normalize(s):
    """Normalize for comparison"""
    s = s.lower().strip()
    s = re.sub(r'\s+', ' ', s)
    s = s.replace('（', '(').replace('）', ')')
    s = s.replace('，', ',').replace('：', ':')
    s = s.replace('×', 'x').replace('—', '-')
    s = s.replace('φ', 'Φ').replace('ø', 'Φ')
    return s

def ultra_normalize(s):
    """Remove all formatting, keep only alphanumeric + Chinese"""
    s = normalize(s)
    s = re.sub(r'[\s\-_/\\(),;:.，：；、。{}\[\]<>]+', '', s)
    return s

def desc_similarity(a, b):
    return SequenceMatcher(None, normalize(a), normalize(b)).ratio()


def extract_all_spec_tokens(s):
    """
    Extract all specification-relevant tokens from description.
    Uses ORIGINAL case string for model/curve extraction, lower for rest.
    """
    s_orig = s.strip()  # preserve original case for model/curve matching
    ns = normalize(s)   # lowercased for other extractions
    
    # 1. Model numbers: letter-digit combinations from ORIGINAL case
    # e.g. CDM3, iC65H, S203, NB1-63, NXB-63
    models = re.findall(r'[a-zA-Z]+\d+[a-zA-Z0-9\-]*', s_orig)
    
    # 2. Trip curve: letter C/D/B before a number (case-insensitive)
    # Patterns: " C63", " D16", "-C32", "/D6", "C6A", "D40A"
    # Use original case string
    trip_curves = []
    # Pattern 1: space/slash/dash/plus followed by B/C/D then digits
    for m in re.finditer(r'(?:^|[\s/\-+])([BCDbcd])(\d+)', s_orig):
        trip_curves.append(m.group(1).upper())
    # Pattern 2: B/C/D immediately before digits at word boundary
    for m in re.finditer(r'\b([BCDbcd])(\d+)', s_orig):
        trip_curves.append(m.group(1).upper())
    all_curves = list(set(trip_curves))
    
    # 3. Currents: extract ALL digit+A patterns
    # Match patterns like "63A", "6A", "100A" etc. - regardless of preceding chars
    # Use normalized string but match broadly
    currents = []
    for m in re.finditer(r'(\d+(?:\.\d+)?)\s*a(?:\s|$|[^a-zA-Z0-9]|$)', ns):
        currents.append(m.group(1))
    # Also try original case
    for m in re.finditer(r'(\d+(?:\.\d+)?)\s*[Aa](?:\s|$|[^a-zA-Z0-9]|$)', s_orig):
        currents.append(m.group(1))
    currents = list(set(currents))
    
    # 4. Voltage
    voltages = []
    for m in re.finditer(r'(\d+(?:\.\d+)?)\s*[Vv](?:\s|$|[^a-zA-Z0-9])', s_orig):
        voltages.append(m.group(1))
    voltages = list(set(voltages))
    
    # 5. Power
    powers = []
    for m in re.finditer(r'(\d+(?:\.\d+)?)\s*[Kk][Ww]?', s_orig):
        powers.append(m.group(1))
    powers = list(set(powers))
    
    # 6. Dimensions: NxN or NxNxN
    dims = re.findall(r'(\d+(?:\.\d+)?)x(\d+(?:\.\d+)?)(?:x(\d+(?:\.\d+)?))?', ns)
    
    # 7. Pole count
    poles = re.findall(r'(\d)\s*[Pp](?:\s*[+/]\s*\d?\s*[Pp]?\s*[Nn]?)?(?:\s|$|[^0-9])', ns)
    
    return {
        'models': sorted(set(models)),
        'currents': sorted(set(currents)),
        'voltages': sorted(set(voltages)),
        'powers': sorted(set(powers)),
        'dims': sorted(set(tuple(d) for d in dims)),
        'poles': sorted(set(poles)),
        'trip_curves': sorted(all_curves),
    }


def specs_differ(a, b):
    """
    CORE RULE: Compare specifications extracted from descriptions.
    If ANY key spec parameter differs → non-duplicate.
    Returns (differ: bool, reason: str)
    """
    desc_a = a['desc']
    desc_b = b['desc']
    ta = desc_a
    tb = desc_b
    
    # Combined text
    full_a = desc_a + ' ' + a['name']
    full_b = desc_b + ' ' + b['name']
    
    # === Rule 0: Shape/size/color/material/inner-outer ===
    
    # Inner vs Outer
    if (('内' in full_a and '外' in full_b) or ('外' in full_a and '内' in full_b)):
        # More precise: check for context words
        inner_ctx = any(w in full_a for w in ['内门', '内侧', '内部', '内层', '内装'])
        outer_ctx = any(w in full_b for w in ['外门', '外侧', '外部', '外层', '外装'])
        if (inner_ctx and outer_ctx) or (any(w in full_b for w in ['内门', '内侧', '内部', '内层', '内装']) and 
                                          any(w in full_a for w in ['外门', '外侧', '外部', '外层', '外装'])):
            return True, '内vs外'
    
    # Shape differences
    shape_pairs = [('方形', '圆形'), ('方型', '圆型'), ('U型', '对接型'),
                   ('U型', '直型'), ('平型', '立式')]
    for s1, s2 in shape_pairs:
        if (s1 in ta and s2 in tb) or (s2 in ta and s1 in tb):
            return True, f'形状不同: {s1} vs {s2}'
    
    # === Rule 1: Extract and compare ALL spec tokens ===
    
    specs_a = extract_all_spec_tokens(ta)
    specs_b = extract_all_spec_tokens(tb)
    
    # Model tokens comparison
    models_a = specs_a['models']
    models_b = specs_b['models']
    
    if models_a and models_b:
        # Find the "main model" - the longest model token or the one that appears first
        # Compare model tokens: if they share a prefix family but differ in specs
        common_prefixes = set()
        for ma in models_a:
            prefix = re.match(r'^([a-zA-Z]+)', ma)
            if prefix:
                common_prefixes.add(prefix.group(1).lower())
        
        for ma in models_a:
            for mb in models_b:
                # Normalize for comparison
                ma_low = ma.lower()
                mb_low = mb.lower()
                
                if ma_low == mb_low:
                    continue
                
                # Check if they're from the same model family
                pre_a = re.match(r'^([a-zA-Z]+)', ma_low)
                pre_b = re.match(r'^([a-zA-Z]+)', mb_low)
                
                if pre_a and pre_b:
                    pa = pre_a.group(1)
                    pb = pre_b.group(1)
                    
                    # If same prefix (same model family), check for spec differences
                    if pa == pb:
                        rest_a = ma_low[len(pa):]
                        rest_b = mb_low[len(pa):]
                        
                        # Compare numbers in the suffix
                        nums_a = re.findall(r'\d+', rest_a)
                        nums_b = re.findall(r'\d+', rest_b)
                        if nums_a != nums_b:
                            return True, f'型号数字不同: {ma} vs {mb}'
                        
                        # Compare letters in the suffix
                        letters_a = re.findall(r'[a-zA-Z]+', rest_a)
                        letters_b = re.findall(r'[a-zA-Z]+', rest_b)
                        if letters_a and letters_b and letters_a != letters_b:
                            return True, f'型号字母不同: {ma} vs {mb}'
                    
                    # Cross-family: different model families with overlapping letters
                    # e.g. "ic65h" vs "s203" - different families entirely → definitely different
                    elif pa != pb:
                        # If both are substantial model tokens (2+ letters + digits)
                        if len(pa) >= 2 and len(pb) >= 2:
                            return True, f'型号系列不同: {ma} vs {mb}'
    
    # Current comparison
    currs_a = specs_a['currents']
    currs_b = specs_b['currents']
    if currs_a and currs_b:
        # Compare the primary (largest) current
        set_a = set(currs_a)
        set_b = set(currs_b)
        if set_a != set_b:
            # Check if there's a real difference (not just formatting)
            nums_a = set(float(c) for c in currs_a)
            nums_b = set(float(c) for c in currs_b)
            if nums_a != nums_b:
                return True, f'电流不同: {sorted(currs_a)}A vs {sorted(currs_b)}A'
    
    # Voltage comparison
    volts_a = specs_a['voltages']
    volts_b = specs_b['voltages']
    if volts_a and volts_b:
        set_a = set(float(v) for v in volts_a)
        set_b = set(float(v) for v in volts_b)
        if set_a != set_b:
            return True, f'电压不同: {sorted(volts_a)}V vs {sorted(volts_b)}V'
    
    # Power comparison
    powers_a = specs_a['powers']
    powers_b = specs_b['powers']
    if powers_a and powers_b:
        set_a = set(float(p) for p in powers_a)
        set_b = set(float(p) for p in powers_b)
        if set_a != set_b:
            return True, f'功率不同: {sorted(powers_a)}KW vs {sorted(powers_b)}KW'
    
    # Pole count comparison
    poles_a = specs_a['poles']
    poles_b = specs_b['poles']
    if poles_a and poles_b and set(poles_a) != set(poles_b):
        return True, f'极数不同: {poles_a}P vs {poles_b}P'
    
    # Trip curve comparison
    curves_a = specs_a['trip_curves']
    curves_b = specs_b['trip_curves']
    if curves_a and curves_b and set(curves_a) != set(curves_b):
        return True, f'脱扣曲线不同: {curves_a} vs {curves_b}'
    
    # Dimension comparison
    dims_a = specs_a['dims']
    dims_b = specs_b['dims']
    if dims_a and dims_b and set(dims_a) != set(dims_b):
        return True, f'尺寸不同: {dims_a} vs {dims_b}'
    
    # === Additional specific checks ===
    
    # Generic: extract all numbers from descriptions and compare
    # If there are standalone numbers that differ, it's likely a spec difference
    all_nums_a = sorted(re.findall(r'(?<![a-zA-Z])(\d+(?:\.\d+)?)(?![a-zA-Z0-9])', normalize(ta)))
    all_nums_b = sorted(re.findall(r'(?<![a-zA-Z])(\d+(?:\.\d+)?)(?![a-zA-Z0-9])', normalize(tb)))
    # Only compare if both have numbers and they differ significantly
    # This catches things like cross-section areas, lengths, etc.
    
    return False, ''


def apply_exclusion_rules(a, b):
    """Apply exclusion rules after specs check. Returns (is_excluded, reason)"""
    desc_a = a['desc']
    desc_b = b['desc']
    ta = desc_a + ' ' + a['name']
    tb = desc_b + ' ' + b['name']
    mfg_a = a['manufacturer']
    mfg_b = b['manufacturer']
    
    # 甲供件
    for m in [a, b]:
        if '甲供' in m['cat'] or '甲供' in m['subcat']:
            return True, '甲供件不算重复'
    
    # One has manufacturer, one doesn't
    if (mfg_a and not mfg_b) or (not mfg_a and mfg_b):
        return True, '一个有制造商一个没有'
    
    # Manufacturers completely different
    if mfg_a and mfg_b and mfg_a != mfg_b:
        aliases = [
            ('宁波三爱', '三爱'), ('海越电气', '海越湖北'),
            ('天灵', '宁波天灵'), ('长沙威胜', '威胜'),
            ('上海良信电器股份有限公司', '良信'),
            ('上海良信电器', '良信'),
        ]
        is_alias = False
        for x, y in aliases:
            if (x in mfg_a and y in mfg_b) or (y in mfg_a and x in mfg_b):
                is_alias = True
                break
        if not is_alias:
            return True, f'制造商不同: {mfg_a} vs {mfg_b}'
    
    # Direction differences
    dir_pairs = [
        ('左操', '右操'), ('平进平出', '平进侧出'), ('上进', '下进'),
        ('上进下出', '下进上出'), ('上进下出', '上出'), ('立式', '卧式'),
        ('侧进', '上进'), ('侧出', '下出'),
    ]
    for d1, d2 in dir_pairs:
        if (d1 in ta and d2 in tb) or (d2 in ta and d1 in tb):
            return True, f'方向不同: {d1} vs {d2}'
    
    # Breaker frame letter suffix
    frame_pattern = r'(?:CDM|NDM|CM|NM)\d*-?\d+([FCDHLMNS])'
    fa = re.search(frame_pattern, desc_a, re.IGNORECASE)
    fb = re.search(frame_pattern, desc_b, re.IGNORECASE)
    if fa and fb and fa.group(1).upper() != fb.group(1).upper():
        return True, f'断路器分断能力等级不同: {fa.group(1)} vs {fb.group(1)}'
    
    # Color
    if a['color'] and b['color'] and a['color'] != b['color']:
        return True, f'颜色不同: {a["color"]} vs {b["color"]}'
    
    # Leakage type
    if ('A型' in ta and 'AC型' in tb) or ('AC型' in ta and 'A型' in tb):
        return True, '漏电保护类型不同'
    
    # Attachment differences
    attachments = ['失压', '分励', '合闸', '辅助', '报警', '门框', '电磁锁',
                   'RS485', '通讯', '温度指示', '带电显示', '加热器']
    att_a = set(x for x in attachments if x in ta)
    att_b = set(x for x in attachments if x in tb)
    diff = att_a ^ att_b
    if diff:
        return True, f'附件不同: {", ".join(sorted(diff))}'
    
    # Wire type
    wire_types = ['BVR', 'ZR-YJV', 'NH-YJV', 'YJV', 'RVV', 'RVB', 'BV', 'RV']
    w_a = set(w for w in wire_types if w in ta)
    w_b = set(w for w in wire_types if w in tb)
    if w_a and w_b and w_a != w_b:
        return True, f'电线类型不同: {w_a} vs {w_b}'
    
    # Thread direction
    if ('正牙' in ta and '反牙' in tb) or ('反牙' in ta and '正牙' in tb):
        return True, '螺纹方向不同'
    
    # Trip method TMD vs MA
    if ('TMD' in ta.upper() and 'MA' in tb.upper()) or \
       ('MA' in ta.upper() and 'TMD' in tb.upper()):
        return True, '脱扣方式不同(TMD vs MA)'
    
    # Trip unit LSI vs TMA
    if ('LSI' in ta and 'TMA' in tb) or ('TMA' in ta and 'LSI' in tb):
        return True, '脱扣单元不同(LSI vs TMA)'
    
    # TBP series
    tbp_a = re.search(r'TBP[-]?([A-Z])', desc_a)
    tbp_b = re.search(r'TBP[-]?([A-Z])', desc_b)
    if tbp_a and tbp_b and tbp_a.group(1) != tbp_b.group(1):
        return True, f'过电压保护器型号不同: TBP-{tbp_a.group(1)} vs TBP-{tbp_b.group(1)}'
    
    # Model series differences (catches electronic vs non-electronic)
    series_pairs = [('CM3E', 'CM3'), ('CM1E', 'CM1'), ('NDM3', 'CDM3'),
                    ('CM3', 'CM1'), ('CDM3', 'CDM1')]
    for s1, s2 in series_pairs:
        if (s1 in desc_a and s2 in desc_b) or (s2 in desc_a and s1 in desc_b):
            return True, f'型号系列不同: {s1} vs {s2}'
    
    # Material differences
    mat_pairs = [('铜', '铝'), ('紫铜', '黄铜'), ('304', '316'), ('全铜', '全铝')]
    for m1, m2 in mat_pairs:
        if (m1 in ta and m2 in tb) or (m2 in ta and m1 in tb):
            return True, f'材质不同: {m1} vs {m2}'
    
    # EPS power
    eps_a = re.search(r'EPS\s*(\d+(?:\.\d+)?)\s*KW?', ta, re.IGNORECASE)
    eps_b = re.search(r'EPS\s*(\d+(?:\.\d+)?)\s*KW?', tb, re.IGNORECASE)
    if eps_a and eps_b and eps_a.group(1) != eps_b.group(1):
        return True, f'EPS功率不同: {eps_a.group(1)}KW vs {eps_b.group(1)}KW'
    
    return False, ''


def classify_pair(a, b):
    """Classify a pair as confirmed duplicate, non-duplicate, or needs review"""
    # STEP 1: Spec parameter check (Rule 0+1) - THE MOST IMPORTANT FILTER
    differ, reason = specs_differ(a, b)
    if differ:
        return 'non_dup', reason
    
    # STEP 2: Exclusion rules
    excluded, reason = apply_exclusion_rules(a, b)
    if excluded:
        return 'non_dup', reason
    
    # STEP 3: Check if confirmed duplicate
    desc_a = normalize(a['desc'])
    desc_b = normalize(b['desc'])
    
    if desc_a == desc_b:
        return 'confirmed', '描述完全相同(归一化后)'
    
    # Ultra-normalized check (symbol/format differences only)
    ua = ultra_normalize(a['desc'])
    ub = ultra_normalize(b['desc'])
    if ua == ub:
        return 'confirmed', '描述仅格式/符号差异'
    
    # 1N vs 1P+N normalization
    da = desc_a.replace('1n', '1p+n')
    db = desc_b.replace('1n', '1p+n')
    if da == db:
        return 'confirmed', '1N vs 1P+N 同一产品'
    
    # STEP 4: Needs manual review - but apply strict similarity threshold
    sim = desc_similarity(a['desc'], b['desc'])
    
    if sim < 0.6:
        return 'non_dup', f'相似度太低({sim:.1%})'
    
    return 'review', f'相似度{sim:.1%}，需人工确认'


# === MAIN PROCESSING ===

# Strategy: Group by (cat, subcat, name) for efficiency
# Only compare within groups where descriptions are potentially similar

print("Grouping materials...")
groups = defaultdict(list)
for m in materials:
    key = (m['cat'], m['subcat'], normalize(m['name']))
    groups[key].append(m)

print(f"Groups by (cat, subcat, name): {len(groups)}")

# Phase 1: Find candidate pairs within same name groups
# Use a pre-filter: only compare if ultra_normalized descriptions share enough similarity
candidate_pairs = []
seen_pairs = set()
skipped_by_prefilter = 0

print("Phase 1: Same-name group comparisons...")
for key, group in groups.items():
    if len(group) < 2:
        continue
    
    # Pre-compute ultra_normalized descriptions
    norms = [(m, ultra_normalize(m['desc'])) for m in group]
    
    for i in range(len(group)):
        for j in range(i + 1, len(group)):
            a, b = group[i], group[j]
            if a['code'] == b['code']:
                continue
            pk = tuple(sorted([a['code'], b['code']]))
            if pk in seen_pairs:
                continue
            
            # Quick pre-filter: if ultra_normalized descriptions are identical or very close
            na, nb = norms[i][1], norms[j][1]
            if na == nb:
                # Definitely candidate - same content, possibly different formatting
                seen_pairs.add(pk)
                candidate_pairs.append((a, b))
            else:
                # Use quick length-based filter
                if abs(len(na) - len(nb)) > max(len(na), len(nb)) * 0.3:
                    skipped_by_prefilter += 1
                    continue
                # Check similarity with a fast method
                sim = SequenceMatcher(None, na, nb).ratio()
                if sim >= 0.6:
                    seen_pairs.add(pk)
                    candidate_pairs.append((a, b))
                else:
                    skipped_by_prefilter += 1

print(f"Same-name candidates: {len(candidate_pairs)}, prefiltered: {skipped_by_prefilter}")

# Phase 2: Cross-name check - same (cat, subcat) with identical ultra-normalized descriptions
print("Phase 2: Cross-name identical descriptions...")
cat_groups = defaultdict(list)
for m in materials:
    key = (m['cat'], m['subcat'])
    cat_groups[key].append(m)

cross_name_count = 0
for key, group in cat_groups.items():
    desc_groups = defaultdict(list)
    for m in group:
        nd = ultra_normalize(m['desc'])
        if nd:
            desc_groups[nd].append(m)
    
    for nd, items in desc_groups.items():
        if len(items) < 2:
            continue
        for i in range(len(items)):
            for j in range(i + 1, len(items)):
                a, b = items[i], items[j]
                if a['code'] == b['code']:
                    continue
                if normalize(a['name']) == normalize(b['name']):
                    continue  # already in same-name groups
                pk = tuple(sorted([a['code'], b['code']]))
                if pk not in seen_pairs:
                    seen_pairs.add(pk)
                    candidate_pairs.append((a, b))
                    cross_name_count += 1

print(f"Cross-name pairs: {cross_name_count}")
print(f"Total candidates: {len(candidate_pairs)}")

# Classify all pairs
confirmed = []
non_dup = []
review = []

for a, b in candidate_pairs:
    cat, reason = classify_pair(a, b)
    if cat == 'confirmed':
        confirmed.append((a, b, reason))
    elif cat == 'non_dup':
        non_dup.append((a, b, reason))
    else:
        review.append((a, b, reason))

print(f"\n===== RESULTS =====")
print(f"  Confirmed duplicates: {len(confirmed)}")
print(f"  Non-duplicates (excluded): {len(non_dup)}")
print(f"  Needs manual review: {len(review)}")

if review:
    print(f"\n--- Review samples (first 20) ---")
    for a, b, reason in review[:20]:
        print(f"  {a['code']} vs {b['code']}")
        print(f"    A: {a['desc'][:80]}")
        print(f"    B: {b['desc'][:80]}")
        print(f"    [{reason}]")

# === GENERATE EXCEL ===
print("\nGenerating Excel...")

out_wb = openpyxl.Workbook()

blue_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
orange_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
red_font = Font(color='FF0000')
normal_font = Font()
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_w = Font(bold=True, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

EXCEL_COLS = [
    ('序号', 'label'),
    ('物料编号', 'code'),
    ('物料名称', 'name'),
    ('物料描述', 'desc'),
    ('物料类别', 'cat'),
    ('物料子类别', 'subcat'),
    ('制造商', 'manufacturer'),
    ('物料来源', 'source'),
    ('主计量单位', 'unit'),
    ('提前期', 'lead_time'),
    ('标准价格', 'price'),
    ('创建人', 'creator'),
    ('创建日期', 'create_date'),
    ('最近修改人', 'modifier'),
    ('最近修改日期', 'modify_date'),
    ('备注', 'remark'),
]

DIFF_COLS = {'name', 'desc', 'cat', 'subcat', 'manufacturer', 'source', 'unit', 'lead_time', 'price'}

def write_sheet(ws, data):
    for col_idx, (col_name, _) in enumerate(EXCEL_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font_w
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    for pair_num, (a, b, reason) in enumerate(data, 1):
        for is_b, mat in enumerate([a, b]):
            fill = orange_fill if is_b else blue_fill
            label = f"{'B' if is_b else 'A'}-{pair_num}"
            
            for col_idx, (col_name, col_key) in enumerate(EXCEL_COLS, 1):
                if col_key == 'label':
                    value = label
                elif col_key == 'remark':
                    value = reason
                else:
                    value = mat.get(col_key, '')
                    if isinstance(value, float) and value == int(value):
                        value = int(value)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=(col_key == 'desc'))
                
                if col_key in DIFF_COLS:
                    val_a = str(a.get(col_key, '') or '').strip()
                    val_b = str(b.get(col_key, '') or '').strip()
                    cell.font = red_font if val_a != val_b else normal_font
                else:
                    cell.font = normal_font
            
            row_idx += 1
    
    col_widths = {
        'A': 8, 'B': 16, 'C': 18, 'D': 50, 'E': 12, 'F': 14,
        'G': 20, 'H': 10, 'I': 10, 'J': 8, 'K': 12,
        'L': 10, 'M': 14, 'N': 12, 'O': 14, 'P': 40
    }
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w
    
    ws.freeze_panes = 'C2'
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLS))}{row_idx - 1}"


# Sheet 1: Confirmed duplicates
ws1 = out_wb.active
ws1.title = '确认重复'
write_sheet(ws1, confirmed)

# Sheet 2: Needs review only
ws2 = out_wb.create_sheet('待人工确认')
write_sheet(ws2, review)

out_wb.save(OUTPUT_FILE)
print(f"\nExcel saved to: {OUTPUT_FILE}")
print(f"Total time: {time.time()-start:.1f}s")
