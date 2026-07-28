#!/usr/bin/env python3
"""
IN3 物料重复检查工具（stable CLI version）

用法：
  # 分析最新导出的物料主数据，输出可疑重复 Excel
  python in3_dup_check.py analyze <输入文件.xlsx> [--output <输出文件.xlsx>]

  # 对比两次检查结果，生成变更报告
  python in3_dup_check.py compare <旧文件.xlsx> <新文件.xlsx> [--output <报告.txt>]

  # 自动查找：用最新的物料主数据文件进行分析
  python in3_dup_check.py auto [--dir <目录>]

示例：
  python in3_dup_check.py analyze IN3数据/物料主数据导出结果-2026-05-21.xlsx
  python in3_dup_check.py analyze IN3数据/物料主数据导出结果-2026-05-21.xlsx -o IN3数据/可疑重复物料-2026-05-21.xlsx
  python in3_dup_check.py compare IN3数据/可疑重复物料-2026-05-19-v3.xlsx IN3数据/可疑重复物料-2026-05-21-v3.xlsx
  python in3_dup_check.py auto
"""

import argparse
import json
import os
import re
import sys
import time
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─── 常量 ───────────────────────────────────────────────────────────────────

# IN3 导出 Excel 的列映射
COL = {
    'code': 2, 'name': 6, 'desc': 7, 'cat': 9, 'subcat': 11,
    'manufacturer': 13, 'source': 14, 'lead_time': 18, 'unit': 22,
    'price': 61, 'creator': 70, 'create_date': 72,
    'modifier': 73, 'modify_date': 75, 'color': 16,
}

# 排除的物料类别
EXCLUDE_CATS = {'成品柜', '外购成套', '管道配件', '气动配件', '外包服务'}

# ─── 名称同义词组（用于跨名称配对） ─────────────────────────────────────────
# 同一组内的名称互为同义词，配对后可进入确认/待人工确认
NAME_SYNONYMS = [
    {'弹垫', '弹簧垫片', '弹簧垫圈'},
    {'电度表', '电能表'},
    {'角钢', '角铁'},
    {'螺母', '螺帽'},
    {'接触器', '交流接触器'},
    {'断路器', '塑壳断路器'},
    {'汇流零线排', '汇流端子'},
    {'终端电能计量表', '电能表', '电度表'},
    {'转换开关', '万能转换开关'},
    {'螺丝', '螺钉'},
]

# 构建反向索引: name → synonym_group_id
_NAME_TO_SYN_GROUP = {}
for _i, _grp in enumerate(NAME_SYNONYMS):
    for _name in _grp:
        _NAME_TO_SYN_GROUP[_name] = _i


def _are_synonyms(name_a, name_b):
    """判断两个物料名称是否为同义词"""
    # 精确匹配
    ga = _NAME_TO_SYN_GROUP.get(name_a)
    gb = _NAME_TO_SYN_GROUP.get(name_b)
    if ga is not None and gb is not None and ga == gb:
        return True
    # 包含匹配：名称A包含名称B 或反之（如"不锈钢弹垫"包含"弹垫"）
    for _grp in NAME_SYNONYMS:
        # 找出组内名称在 A 和 B 中的匹配
        a_match = any(_s in name_a for _s in _grp)
        b_match = any(_s in name_b for _s in _grp)
        # 同组名称修饰差异（如"不锈钢弹垫" vs "不锈钢弹簧垫片"）
        # 需要进一步检查：去掉共同前缀/后缀后是否是同义词
        if a_match and b_match:
            # 验证不是不同组的词同时出现
            # 找A匹配的具体词和B匹配的具体词
            for _wa in _grp:
                if _wa in name_a:
                    for _wb in _grp:
                        if _wb in name_b and _wa != _wb:
                            # 确保前缀一致（如都是"不锈钢"开头）
                            # 去掉同义词部分后的前缀应该相同
                            prefix_a = name_a.replace(_wa, '').strip()
                            prefix_b = name_b.replace(_wb, '').strip()
                            if prefix_a == prefix_b:
                                return True
    return False

# ─── 品牌型号命名规则知识库 ─────────────────────────────────────────────────
# 从 brand_naming_rules/naming_rules.json 加载，用于规则 #46
_BRAND_RULES_PATH = Path(__file__).parent / 'brand_naming_rules' / 'naming_rules.json'
BRAND_NAMING_RULES = {}
if _BRAND_RULES_PATH.exists():
    try:
        with open(_BRAND_RULES_PATH, 'r', encoding='utf-8') as _f:
            BRAND_NAMING_RULES = json.load(_f)
    except (json.JSONDecodeError, OSError):
        pass  # 文件损坏时不阻塞，规则#46降级为不生效

# 提取跨品牌的附件/安装/接线后缀字典（短代号→含义），用于后缀差异检测
# 只提取短代号（1-4字符），因为长代号（如"分励脱扣器"）已有中文规则覆盖
_SUFFIX_CODES = {}  # {'F': {'meaning': '辅助触头', 'type': '附件'}, ...}
_INSTALL_CODES = {}  # {'P': {'meaning': '插入式', 'type': '安装'}, ...}
_CONNECTION_CODES = {}  # {'RC': {'meaning': '板后接线', 'type': '接线'}, ...}

def _extract_suffix_codes():
    """从品牌规则JSON中提取所有短代号后缀（附件/安装/接线/脱扣器）"""
    suffixes = {}
    installs = {}
    connects = {}
    trip_units = {}
    for _brand, _data in BRAND_NAMING_RULES.items():
        if _brand == '说明' or not isinstance(_data, dict):
            continue
        for _series_name, _series_data in _data.get('系列', {}).items():
            if not isinstance(_series_data, dict):
                continue
            # 附件代号
            _att = _series_data.get('附件代号', {})
            if isinstance(_att, dict):
                for _code, _desc in _att.items():
                    _c = str(_code).strip()
                    if 1 <= len(_c) <= 4 and _c.isascii() and _c.isupper():
                        suffixes.setdefault(_c, {'meaning': str(_desc)[:30], 'type': '附件'})
            # 安装方式
            _inst = _series_data.get('安装方式', {})
            if isinstance(_inst, dict):
                for _code, _desc in _inst.items():
                    _c = str(_code).strip()
                    if 1 <= len(_c) <= 4 and _c.isascii() and _c.isupper():
                        installs.setdefault(_c, {'meaning': str(_desc)[:30], 'type': '安装'})
            # 接线方式
            _conn = _series_data.get('接线方式', {})
            if isinstance(_conn, dict):
                for _code, _desc in _conn.items():
                    _c = str(_code).strip()
                    if 1 <= len(_c) <= 4 and _c.isascii() and _c.isupper():
                        connects.setdefault(_c, {'meaning': str(_desc)[:30], 'type': '接线'})
            # 脱扣器类型
            _trip = _series_data.get('脱扣器类型', {})
            if isinstance(_trip, dict):
                for _code, _desc in _trip.items():
                    _c = str(_code).strip().replace('-', '')
                    if 2 <= len(_c) <= 6 and _c.isascii() and _c.isupper():
                        trip_units.setdefault(_c, {'meaning': str(_desc)[:30], 'type': '脱扣器'})
    return suffixes, installs, connects, trip_units

_SUFFIX_CODES, _INSTALL_CODES, _CONNECTION_CODES, _TRIP_UNIT_CODES = _extract_suffix_codes()

# 保留硬编码作为兜底（即使JSON文件丢失也能工作）
_FALLBACK_SUFFIXES = {
    'F': '辅助触头', 'P': '板前接线/插入式', 'R': '板后接线',
    'SD': '信号接点', 'FM': '遥信模块',
    'W': '抽出式', 'D': '抽出式',
    'OF': '辅助触头', 'MX': '分励脱扣器', 'MN': '欠压脱扣器',
    'SDE': '故障信号接点', 'MT': '电动操作',
    'EF': '加长前接线', 'ES': '加长扩展前接线',
    'EL': '漏电保护', 'RC': '板后接线', 'FC': '板前接线',
    'AL': '辅助触头', 'AX': '报警触头', 'SHT': '分励脱扣器',
    'UVT': '欠压脱扣器', 'MG': '电动操作',
    'TMF': '热磁脱扣器(配电)', 'TMD': '热磁脱扣器(可调)',
    'TMA': '热磁脱扣器(可调)', 'MA': '单磁脱扣器',
    'MF': '单磁脱扣器(电机)',
    'LI': '电子脱扣器(长延时+瞬时)',
    'LSI': '电子脱扣器(三段保护)',
    'LSIG': '电子脱扣器(含接地)',
}
# 合并：JSON规则优先，兜底补充
_ALL_SUFFIX_CODES = {**_FALLBACK_SUFFIXES, **{k: v['meaning'] for k, v in _SUFFIX_CODES.items()},
                    **{k: v['meaning'] for k, v in _INSTALL_CODES.items()},
                    **{k: v['meaning'] for k, v in _CONNECTION_CODES.items()},
                    **{k: v['meaning'] for k, v in _TRIP_UNIT_CODES.items()}}

# Excel 输出字段
FIELDS = [
    ('标记', 'label'), ('物料编号', 'code'), ('物料名称', 'name'),
    ('物料描述', 'desc'), ('物料类别', 'cat'), ('物料子类别', 'subcat'),
    ('制造商', 'manufacturer'), ('物料来源', 'source'), ('提前期', 'lead_time'),
    ('主计量单位', 'unit'), ('标准价格', 'price'), ('创建人', 'creator'),
    ('创建日期', 'create_date'), ('最近修改人', 'modifier'),
    ('最近修改日期', 'modify_date'), ('备注', 'remark'),
]
DIFF_FIELDS = {'name', 'desc', 'cat', 'subcat', 'manufacturer', 'source', 'lead_time', 'unit', 'price'}

# ─── 工具函数 ───────────────────────────────────────────────────────────────

def norm(s):
    """基本标准化：统一标点、大小写"""
    s = s.lower().strip()
    s = re.sub(r'\s+', ' ', s)
    s = (s.replace('（', '(').replace('）', ')').replace('，', ',')
          .replace('：', ':').replace('×', 'x').replace('—', '-')
          .replace('φ', 'Φ').replace('ø', 'Φ'))
    return s


def ultra(s):
    """深度标准化：去除所有空白和标点（保留数字间小数点）"""
    s = norm(s)
    s = re.sub(r'(?<=\d)\.(?=\d)', '<<DOT>>', s)
    s = re.sub(r'[\s\-_/\\(),;:.，：；、。{}\[\]<>]+', '', s)
    s = s.replace('<<DOT>>', '.')
    return s


def sim(a, b):
    """文本相似度"""
    return SequenceMatcher(None, norm(a), norm(b)).ratio()


def gv(ws, row, col):
    """安全读取单元格值"""
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ''


# ─── 物料加载 ───────────────────────────────────────────────────────────────

def load_materials(input_path):
    """从 IN3 导出的 Excel 加载物料数据"""
    print(f"📂 加载文件: {input_path}")
    t0 = time.time()

    if not os.path.exists(input_path):
        print(f"❌ 文件不存在: {input_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb['物料主数据']
    print(f"   Sheet: {ws.max_row} 行 × {ws.max_column} 列")

    materials = []
    for r in range(2, ws.max_row + 1):
        cat = gv(ws, r, COL['cat'])
        if cat in EXCLUDE_CATS:
            continue
        materials.append({
            'code': gv(ws, r, COL['code']),
            'name': gv(ws, r, COL['name']),
            'desc': gv(ws, r, COL['desc']),
            'cat': cat,
            'subcat': gv(ws, r, COL['subcat']),
            'manufacturer': gv(ws, r, COL['manufacturer']),
            'source': gv(ws, r, COL['source']),
            'unit': gv(ws, r, COL['unit']),
            'lead_time': gv(ws, r, COL['lead_time']),
            'price': gv(ws, r, COL['price']),
            'creator': gv(ws, r, COL['creator']),
            'create_date': gv(ws, r, COL['create_date']),
            'modifier': gv(ws, r, COL['modifier']),
            'modify_date': gv(ws, r, COL['modify_date']),
            'color': gv(ws, r, COL['color']),
        })
    wb.close()
    print(f"✅ 加载 {len(materials)} 条物料，耗时 {time.time()-t0:.1f}s")
    return materials


# ─── 非重复判断规则（42条） ────────────────────────────────────────────────

def is_nondup(a, b):
    """
    判断两个物料是否明显非重复。
    返回 (is_nondup: bool, reason: str)
    """
    da, db = a['desc'], b['desc']
    na, nb = a['name'], b['name']
    ta, tb = da + ' ' + na, db + ' ' + nb
    ma, mb = a['manufacturer'], b['manufacturer']

    # --- 规则 #0a: 工作服/服装类直接跳过不算重复 ---
    clothing_kw = ['工作服', '保安服', '厂服', '制服']
    if any(k in na for k in clothing_kw) or any(k in nb for k in clothing_kw):
        return True, '工作服/服装类不算重复'

    # --- 规则 #0b: 明显不同的工具/器具不算重复 ---
    # 油漆刷 ≠ 铲刀/毛刷，电烙铁 ≠ 熔锡炉
    tool_conflicts = [
        ('油漆刷', ['铲刀', '毛刷', '滚筒']),
        ('铲刀', ['油漆刷', '毛刷', '滚筒']),
        ('电烙铁', ['熔锡炉', '热风枪']),
        ('熔锡炉', ['电烙铁', '热风枪']),
    ]
    for tool_a, conflicts_b in tool_conflicts:
        if tool_a in na:
            for cf in conflicts_b:
                if cf in nb:
                    return True, f'不同工具: {na} ≠ {nb}'
        if tool_a in nb:
            for cf in conflicts_b:
                if cf in na:
                    return True, f'不同工具: {na} ≠ {nb}'

    # --- 规则 #0: 甲供件/结构件不算重复 ---
    for m in [a, b]:
        if '甲供' in m['cat'] or '甲供' in m['subcat']:
            return True, '甲供件不算重复'
        if m['cat'] == '结构件':
            return True, '结构件不算重复'

    # --- 规则 #1: 一个有制造商一个没有 = 非重复（无例外） ---
    if (ma and not mb) or (not ma and mb):
        return True, '一个有制造商一个没有'

    # --- 规则 #2: 制造商不同 ---
    if ma and mb and ma != mb:
        # 完整的制造商别名分组
        mfr_aliases = [
            ('三菱', '三菱电机'),
            ('ABB', 'ABB'),
            ('良信', 'NDM'), ('良信', '良信电器'), ('良信', '上海良信电器'), ('良信', '上海良信电器股份有限公司'),
            ('正泰', 'CHNT'), ('正泰', '正泰电器'), ('正泰', '正泰电器股份有限公司'), ('正泰', '正泰集团股份有限公司'),
            ('德力西', 'DELIXI'), ('德力西', '德力西电气'), ('德力西', '德力西电气有限公司'),
            ('环宇', '环宇高科'), ('环宇', 'HUYU'),
            ('指明', '指明集团'), ('指明', 'ZM'),
            ('天正', 'TENGEN'), ('天正', '天正电器'), ('天正', '浙江天正电气股份有限公司'),
            ('常熟', '常熟开关'), ('常熟', 'RIY1'),
            ('上海人民', '人民电器'), ('上海人民', 'RMW'), ('上海人民', 'RMM'),
            ('宁波三爱', '三爱'), ('宁波三爱', '宁波/三爱'),
            ('大连一互', '一互'),
            ('大连北方', '北方'),
            ('宁波天灵', '天灵'), ('宁波天灵', '宁波/天灵'),
            ('江阴长江', '长江'),
            ('长江电气', '长江'),
            ('浙江天际', '天际'),
            ('宁波同禾', '同禾'),
            ('浙江三狮', '三狮'),
            ('江北和能', '和能'),
            ('宁波莱堡', '莱堡'), ('宁波莱堡', '宁波莱宝'),
            ('宁波正格', '正格'),
            ('杭州华世', '华世'),
            ('浙江侃诚', '侃诚'),
            ('苏州天业', '天业'),
            ('狮特', '狮特电力'), ('狮特', '浙江狮特电力'),
            ('海越电气', '海越湖北'), ('海越电气', '海越（湖北）'),
            ('长沙威胜', '威胜'),
            ('默飓电气', '默飓'),
        ]
        if not any((x in ma and y in mb) or (y in ma and x in mb) or (x in ma and x in mb) for x, y in mfr_aliases):
            return True, f'制造商不同: {ma} vs {mb}'

    # --- 规则 #3a: 方向/位置互斥词 = 非重复 ---
    dir_pos_pairs = [
        ('左', '右'), ('上', '下'), ('前', '后'), ('内', '外'),
        ('左开', '右开'), ('上开', '下开'), ('上开口', '下开口'),
        ('停电', '送电'), ('正转', '反转'), ('正向', '反向'),
    ]
    for w1, w2 in dir_pos_pairs:
        if (w1 in ta and w2 in tb) or (w2 in ta and w1 in tb):
            return True, f'方向/位置不同: {w1} vs {w2}'

    # --- 规则 #3b: AC vs DC 电压类型不同 = 非重复 ---
    ac_dc_a = set(re.findall(r'(AC|DC)\d+[Vv]', ta))
    ac_dc_b = set(re.findall(r'(AC|DC)\d+[Vv]', tb))
    if ac_dc_a and ac_dc_b:
        ac_a = set(x for x in ac_dc_a if x.startswith('AC'))
        dc_a = set(x for x in ac_dc_a if x.startswith('DC'))
        ac_b = set(x for x in ac_dc_b if x.startswith('AC'))
        dc_b = set(x for x in ac_dc_b if x.startswith('DC'))
        if (ac_a and dc_b) or (dc_a and ac_b):
            return True, f'AC vs DC: {ac_dc_a} vs {ac_dc_b}'

    # --- 规则 #3: 内 vs 外 ---
    iw = ['内六角', '内门', '内侧', '内部', '内层', '内牙', '内螺纹']
    ow = ['外六角', '外门', '外侧', '外部', '外层', '外牙', '外螺纹']
    for i in iw:
        for o in ow:
            if (i in ta and o in tb) or (o in ta and i in tb):
                return True, f'内vs外: {i} vs {o}'

    # --- 规则 #4: 螺丝头部形状不同 ---
    head_types = [
        '十字', '一字', '内六角', '外六角', '梅花', '半圆头', '沉头',
        '平头', '圆头', '盘头', '六角头', '方头', '圆柱头', '扁圆头',
        '大扁头', '伞头',
    ]
    ht_a = set(h for h in head_types if h in ta)
    ht_b = set(h for h in head_types if h in tb)
    if ht_a and ht_b and ht_a != ht_b:
        return True, f'螺丝头部形状不同: {ht_a} vs {ht_b}'

    # --- 规则 #5: 分闸 vs 合闸 ---
    # 注意：分励/合闸可能同时出现在断路器描述中，只有互斥才算不同
    if ('分闸' in ta and '合闸' in tb and '合闸' not in ta and '分闸' not in tb) or \
       ('合闸' in ta and '分闸' in tb and '分闸' not in ta and '合闸' not in tb):
        return True, '分闸 vs 合闸: 不同功能'
    if ('分励' in ta and '合闸' in tb and '合闸' not in ta and '分励' not in tb) or \
       ('合闸' in ta and '分励' in tb and '分励' not in ta and '合闸' not in tb):
        return True, '分励 vs 合闸: 不同功能'

    # --- 规则 #6: 静 vs 动 ---
    if ('静插件' in ta and '动插件' in tb) or ('动插件' in ta and '静插件' in tb):
        return True, '静插件 vs 动插件: 不同部件'
    if ('静触头' in ta and '动触头' in tb) or ('动触头' in ta and '静触头' in tb):
        return True, '静触头 vs 动触头: 不同部件'
    if ('静' in ta and '动' in tb and '插件' in ta) or ('动' in ta and '静' in tb and '插件' in tb):
        return True, '静vs动: 不同部件'

    # --- 规则 #7: 形状不同 ---
    shapes = [('方形', '圆形'), ('方型', '圆型'), ('U型', '对接型'),
              ('U型', '直型'), ('平型', '立式')]
    for s1, s2 in shapes:
        if (s1 in ta and s2 in tb) or (s2 in ta and s1 in tb):
            return True, f'形状不同: {s1} vs {s2}'

    # --- 规则 #8: 安装方式 ---
    if ('明装' in ta and '暗装' in tb) or ('暗装' in ta and '明装' in tb):
        return True, '明装 vs 暗装'
    if ('明装' in ta and '嵌入式' in tb) or ('嵌入式' in ta and '明装' in tb):
        return True, '明装 vs 嵌入式'
    if ('暗装' in ta and '嵌入式' in tb) or ('嵌入式' in ta and '暗装' in tb):
        return True, '暗装 vs 嵌入式'
    install_kw = ['暗装', '户外', '户外型', '户外架', '壁挂', '落地', '嵌墙', '挂墙', '明装', '嵌入式']
    inst_a = set(k for k in install_kw if k in na)
    inst_b = set(k for k in install_kw if k in nb)
    # 两边都有安装方式但不同 → 非重复
    if inst_a and inst_b and inst_a != inst_b:
        return True, f'安装方式不同: {inst_a} vs {inst_b}'
    # 一边有安装方式，另一边没有 → 非重复（如「非标箱（暗装）」vs「非标箱」）
    if inst_a != inst_b:
        # 去掉安装方式后名称一致 = 安装方式是唯一差异
        def strip_install(name, inst_set):
            for k in inst_set:
                name = name.replace(k, '').replace('（', '').replace('）', '').strip()
            return name
        base_a = strip_install(na, inst_a)
        base_b = strip_install(nb, inst_b)
        if base_a == base_b:
            if inst_a:
                return True, f'有安装方式 vs 无安装方式: {inst_a}'
            else:
                return True, f'有安装方式 vs 无安装方式: {inst_b}'

    # --- 规则 #9: 方向不同 ---
    dirs = [('左操', '右操'), ('平进平出', '平进侧出'), ('上进', '下进'),
            ('上进下出', '下进上出'), ('立式', '卧式'), ('侧进', '上进'), ('侧出', '下出'),
            ('定向', '万向'), ('被动', '主动')]
    for d1, d2 in dirs:
        if (d1 in ta and d2 in tb) or (d2 in ta and d1 in tb):
            return True, f'方向不同: {d1} vs {d2}'

    # --- 规则 #10: 断路器分断能力 ---
    fp = r'(?:CDM|NDM|CM|NM)\d*-?\d+([FCDHLMNS])'
    fa, fb = re.search(fp, da, re.I), re.search(fp, db, re.I)
    if fa and fb and fa.group(1).upper() != fb.group(1).upper():
        return True, f'分断能力不同: {fa.group(1)} vs {fb.group(1)}'

    # --- 规则 #11: 脱扣曲线 B/C/D ---
    ca = re.search(r'(?:^|[\s/\-+])([BCDbcd])(\d+)', da)
    cb = re.search(r'(?:^|[\s/\-+])([BCDbcd])(\d+)', db)
    if ca and cb and ca.group(1).upper() != cb.group(1).upper():
        return True, f'脱扣曲线不同: {ca.group(1)} vs {cb.group(1)}'

    # --- 规则 #12: 极数 ---
    pa = set(re.findall(r'(\d)\s*[Pp]', da))
    pb = set(re.findall(r'(\d)\s*[Pp]', db))
    if pa and pb and pa != pb:
        return True, f'极数不同: {pa} vs {pb}'

    # --- 规则 #13: 颜色（字段 + 描述/名称中的颜色词） ---
    if a.get('color') and b.get('color') and a['color'] != b['color']:
        return True, f'颜色不同: {a["color"]} vs {b["color"]}'
    # 名称/描述中的颜色词互斥 = 不同物料
    color_words = ['红色', '绿色', '黄色', '蓝色', '白色', '黑色', '橙色', '灰色',
                   '红', '绿', '黄', '蓝', '白', '黑']
    cw_a = set(c for c in color_words if c in ta)
    cw_b = set(c for c in color_words if c in tb)
    # 只有在双方都有颜色词且不同时才算
    if cw_a and cw_b and cw_a != cw_b:
        return True, f'颜色不同: {cw_a} vs {cw_b}'

    # --- 规则 #14: 漏电类型 ---
    if ('A型' in da and 'AC型' in db) or ('AC型' in da and 'A型' in db):
        return True, '漏电类型不同'

    # --- 规则 #15: 附件差异 ---
    atts = ['失压', '分励', '合闸', '辅助', '报警', '门框', '电磁锁', 'RS485', '通讯', '温度指示', '带电显示', '加热器']
    aa = set(x for x in atts if x in da)
    ab = set(x for x in atts if x in db)
    diff = aa ^ ab
    if diff:
        return True, f'附件不同: {",".join(sorted(diff))}'

    # --- 规则 #16: 电线类型 ---
    wires = ['BV', 'BVR', 'YJV', 'ZR-YJV', 'RVV', 'RV', 'RVB', 'NH-YJV']
    wa = set(w for w in wires if w in da)
    wb2 = set(w for w in wires if w in db)
    if wa and wb2 and wa != wb2:
        return True, f'电线类型不同: {wa} vs {wb2}'

    # --- 规则 #17: 螺纹方向 ---
    if ('正牙' in ta and '反牙' in tb) or ('反牙' in ta and '正牙' in tb):
        return True, '螺纹方向不同'

    # --- 规则 #18: 脱扣方式 ---
    if ('TMD' in da.upper() and 'MA' in db.upper()) or ('MA' in da.upper() and 'TMD' in db.upper()):
        return True, '脱扣方式不同'

    # --- 规则 #19: 脱扣单元 ---
    if ('LSI' in da and 'TMA' in db) or ('TMA' in da and 'LSI' in db):
        return True, '脱扣单元不同'

    # --- 规则 #20: 材质 ---
    mat_pairs = [('铜', '铝'), ('304', '316'), ('紫铜', '黄铜')]
    for m1, m2 in mat_pairs:
        if (m1 in ta and m2 in tb) or (m2 in ta and m1 in tb):
            return True, f'材质不同: {m1} vs {m2}'

    # --- 规则 #20b: 描述中数字集合差异 = 非重复（核心规则！） ---
    # Dongfang 多次强调: 描述中任何数字不一致代表不同规格/尺寸/特征 = 非重复
    # 如 400 vs 4000、70 vs 50、L=285 vs L=277、3*95+2*70 vs 3*95+2*50 等
    desc_nums_a = set(re.findall(r'(\d+(?:\.\d+)?)', da))
    desc_nums_b = set(re.findall(r'(\d+(?:\.\d+)?)', db))
    if desc_nums_a != desc_nums_b:
        only_a = desc_nums_a - desc_nums_b
        only_b = desc_nums_b - desc_nums_a
        if only_a or only_b:
            return True, f'描述数字差异: {only_a or "无"} vs {only_b or "无"}'

    # --- 规则 #21-27: 数值参数不同 ---
    # 尺寸 (数字x数字)
    dims_a = re.findall(r'(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)(?:\s*x\s*(\d+(?:\.\d+)?))?', norm(da))
    dims_b = re.findall(r'(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)(?:\s*x\s*(\d+(?:\.\d+)?))?', norm(db))
    if dims_a and dims_b and dims_a != dims_b:
        return True, f'尺寸不同: {dims_a} vs {dims_b}'

    # 母线尺寸
    bus_a = re.findall(r'(?:TMY|TMR|LMY)[\-]?(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)', norm(da), re.I)
    bus_b = re.findall(r'(?:TMY|TMR|LMY)[\-]?(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)', norm(db), re.I)
    if bus_a and bus_b and bus_a != bus_b:
        return True, f'母线尺寸不同: {bus_a} vs {bus_b}'

    # 宽度
    width_a = re.findall(r'(?:宽|宽度|W)[=:\s]*(\d+(?:\.\d+)?)\s*(?:mm|MM)?', da)
    width_b = re.findall(r'(?:宽|宽度|W)[=:\s]*(\d+(?:\.\d+)?)\s*(?:mm|MM)?', da)
    if width_a and width_b and width_a != width_b:
        return True, f'宽度不同: {width_a} vs {width_b}'

    # 直径
    dia_a = re.findall(r'(?:Φ|φ|直径)[=:\s]*(\d+(?:\.\d+)?)', da, re.I)
    dia_b = re.findall(r'(?:Φ|φ|直径)[=:\s]*(\d+(?:\.\d+)?)', db, re.I)
    if dia_a and dia_b and dia_a != dia_b:
        return True, f'直径不同: {dia_a} vs {dia_b}'

    # 电流
    cur_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Aa](?:\s|$|[^a-zA-Z0-9])', da)))
    cur_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Aa](?:\s|$|[^a-zA-Z0-9])', db)))
    if cur_a and cur_b and cur_a != cur_b:
        return True, f'电流不同: {cur_a}A vs {cur_b}A'

    # 电压
    vol_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Vv](?:\s|$|[^a-zA-Z0-9])', da)))
    vol_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Vv](?:\s|$|[^a-zA-Z0-9])', db)))
    if vol_a and vol_b and vol_a != vol_b:
        return True, f'电压不同: {vol_a}V vs {vol_b}V'

    # 功率
    pow_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Kk][Ww]', da)))
    pow_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*[Kk][Ww]', db)))
    if pow_a and pow_b and pow_a != pow_b:
        return True, f'功率不同: {pow_a}KW vs {pow_b}KW'

    # 截面积
    area_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*(?:mm²|mm2|平方)', da)))
    area_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)\s*(?:mm²|mm2|平方)', db)))
    if area_a and area_b and area_a != area_b:
        return True, f'截面积不同: {area_a} vs {area_b}'

    # 电线/导线: 描述中的独立数字 = 线径
    if a['subcat'] in ('导线', '电线') or b['subcat'] in ('导线', '电线'):
        nums_a = sorted(set(re.findall(r'(?:^|\s)(\d+(?:\.\d+)?)(?:\s|$)', da)))
        nums_b = sorted(set(re.findall(r'(?:^|\s)(\d+(?:\.\d+)?)(?:\s|$)', db)))
        if nums_a and nums_b and nums_a != nums_b:
            return True, f'线径/截面积不同: {nums_a} vs {nums_b}'

    # --- 规则 #28: 名称中的数字不同 = 不同物料 ---
    # Dongfang 多次强调: 型号中不一致的数字代表元器件特征, 有任何不同 = 非重复
    name_nums_a = sorted(set(re.findall(r'(\d+(?:\.\d+)?)', a['name'])))
    name_nums_b = sorted(set(re.findall(r'(\d+(?:\.\d+)?)', b['name'])))
    if name_nums_a and name_nums_b and name_nums_a != name_nums_b:
        return True, f'名称数字不同: {a["name"]} vs {b["name"]}'

    # --- 规则 #29: 柜型不同 ---
    cabinet_models = [
        'GGD', 'GCK', 'GCS', 'MNS', 'MCS', 'KYN28', 'KYN61', 'XGN', 'HXGN',
        'DFX', 'YBM', 'ZBW', 'ZGS', 'KYN', 'XGN', 'HXGN',
    ]
    found_cab_a = set(cm for cm in cabinet_models if cm.lower() in ta.lower())
    found_cab_b = set(cm for cm in cabinet_models if cm.lower() in tb.lower())
    if found_cab_a and found_cab_b and found_cab_a != found_cab_b:
        return True, f'柜型不同: {found_cab_a} vs {found_cab_b}'

    # --- 规则 #30-34: 型号 token 比较 (只比较纯型号, 不比较参数) ---
    # 过滤掉 DC24V, AC220V 等参数型 token, 只保留纯型号如 MY4N-GS, CDM3
    def is_pure_model(token):
        """判断是否是纯型号而非参数(如DC24V是参数不是型号)"""
        # 过滤常见的电压/电流参数模式
        if re.match(r'^(AC|DC)?\d+(V|A|KW|Hz)$', token, re.I):
            return False
        return True

    models_a = [m for m in re.findall(r'[a-zA-Z]+\d+[a-zA-Z0-9\-/]*', da) if is_pure_model(m)]
    models_b = [m for m in re.findall(r'[a-zA-Z]+\d+[a-zA-Z0-9\-/]*', db) if is_pure_model(m)]
    if models_a and models_b:
        main_a = max(models_a, key=len)
        main_b = max(models_b, key=len)
        ma_low = main_a.lower().replace('-', '').replace('/', '')
        mb_low = main_b.lower().replace('-', '').replace('/', '')
        if ma_low != mb_low:
            # 如果描述的 ultra 归一化完全一致，说明只是空格/格式差异，不判为不同
            if ultra(da) == ultra(db):
                pass  # 跳过型号差异检查
            else:
                pre_a = re.match(r'^([a-zA-Z]+)', main_a)
                pre_b = re.match(r'^([a-zA-Z]+)', main_b)
                if pre_a and pre_b:
                    pa_s, pb_s = pre_a.group(1).lower(), pre_b.group(1).lower()
                    rest_a, rest_b = main_a[len(pre_a.group(1)):], main_b[len(pre_b.group(1)):]
                    if pa_s == pb_s:
                        nums_a = re.findall(r'\d+', rest_a)
                        nums_b = re.findall(r'\d+', rest_b)
                        if nums_a != nums_b:
                            return True, f'型号数字不同: {main_a} vs {main_b}'
                        lets_a = re.findall(r'[a-zA-Z]+', rest_a)
                        lets_b = re.findall(r'[a-zA-Z]+', rest_b)
                        # Dongfang 2026-07-25: 型号中任何字母差异(含一方有另一方无) = 非重复
                        # 如 FZRN25-12D vs FZRN25-12 (D=带接地)
                        if lets_a != lets_b:
                            return True, f'型号字母不同: {main_a} vs {main_b}'
                    else:
                        base_a = re.sub(r'\d+', '', pa_s)
                        base_b = re.sub(r'\d+', '', pb_s)
                        if base_a != base_b:
                            return True, f'型号系列不同: {main_a} vs {main_b}'

    # --- 规则 #35: 带附加功能 vs 不带 = 不同物料 ---
    # +底座 vs 无、变频 vs 无标注
    extras = ['底座', '变频', '加热包', '带通讯', '带RS485', '插入式', '导轨式',
              '固定式', '抽出式', '板前', '板后']
    for ex in extras:
        has_a = ex in ta
        has_b = ex in tb
        if has_a != has_b:
            return True, f'带/不带{ex}不同'

    # --- 规则 #35b: 描述中 '+xxx' 后缀差异 = 非重复 ---
    # 如 '+插入式'、'+报警'、'+辅助' 等，一方有另一方没有
    plus_suffixes_a = set(re.findall(r'\+([\u4e00-\u9fffA-Za-z0-9]+)', ta))
    plus_suffixes_b = set(re.findall(r'\+([\u4e00-\u9fffA-Za-z0-9]+)', tb))
    if plus_suffixes_a != plus_suffixes_b:
        diff_plus = plus_suffixes_a ^ plus_suffixes_b
        if diff_plus:
            return True, f'附加后缀不同: +{diff_plus}'

    # --- 规则 #36: 材质/颜色标注 vs 无 = 不同 ---
    mat_labels = ['铜制', '铝制', '铜）', '铝）', '(铜', '(铝', '（铜', '（铝', '(白)', '（白）', '白色']
    for ml in mat_labels:
        if (ml in da and ml not in db) or (ml in db and ml not in da):
            return True, f'材质/颜色标注差异: {ml}'

    # --- 规则 #37: 功能位置标注不同 = 不同物料 ---
    func_positions = [
        ('手动', '就地'), ('自动', '远程'),
    ]
    for f1, f2 in func_positions:
        if (f1 in ta and f2 in tb) or (f2 in ta and f1 in tb):
            return True, f'功能位置不同: {f1} vs {f2}'

    # --- 规则 #38: 双开 vs 单开 = 不同 ---
    if ('双开' in ta and '单开' in tb) or ('单开' in ta and '双开' in tb):
        return True, '双开 vs 单开: 不同'
    # 也检查: 双开双控 vs 双控(非双开)
    if '双开双控' in ta and '双控' in tb and '双开' not in tb:
        return True, '双开双控 vs 双控: 不同'
    if '双开双控' in tb and '双控' in ta and '双开' not in ta:
        return True, '双控 vs 双开双控: 不同'

    # --- 规则 #39: 铜排数量不同 = 不同物料（母线框） ---
    # 检查描述中的"+N排"模式差异
    plus_a = re.findall(r'\+(\d+(?:\.\d+)?(?:\*\d+(?:\.\d+)?)*)', da)
    plus_b = re.findall(r'\+(\d+(?:\.\d+)?(?:\*\d+(?:\.\d+)?)*)', db)
    if plus_a and not plus_b and '母线' in ta:
        return True, f'母线框有额外排 vs 无: +{plus_a}'
    if plus_b and not plus_a and '母线' in tb:
        return True, f'母线框有额外排 vs 无: +{plus_b}'

    # --- 规则 #40: 罗马数字不同 (I vs II vs III) ---
    # 支持 30I, 30II, 30III 等后缀形式
    roman_a = set(re.findall(r'(?:\d|(?<!\w))([IVX]{1,4})(?:$|[^a-zA-Z0-9])', da))
    roman_b = set(re.findall(r'(?:\d|(?<!\w))([IVX]{1,4})(?:$|[^a-zA-Z0-9])', db))
    # 也匹配独立的罗马数字（\b 形式）
    roman_a |= set(re.findall(r'\b([IVX]{1,4})\b', da))
    roman_b |= set(re.findall(r'\b([IVX]{1,4})\b', db))
    # 过滤掉太短或非罗马数字的
    roman_a = {r for r in roman_a if r in ('I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X')}
    roman_b = {r for r in roman_b if r in ('I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X')}
    if roman_a and roman_b and roman_a != roman_b:
        return True, f'罗马数字不同: {roman_a} vs {roman_b}'

    # --- 规则 #41: 铜排半成品不参与重复检查 ---
    if '半成品' in na or '半成品' in nb or '半成品' in da or '半成品' in db:
        return True, '半成品不参与重复检查'

    # --- 规则 #42: 元器件后缀不同 (Z3H, Z1, Z2 等) ---
    # Z3H 等后缀代表不同的辅助触头/附件配置
    suffix_pattern = r'[A-Z]\d[A-Z]?(?:$|[\s/\\-])'
    suffix_a = set(re.findall(suffix_pattern, na + ' ' + da))
    suffix_b = set(re.findall(suffix_pattern, nb + ' ' + db))
    if suffix_a and suffix_b and suffix_a != suffix_b:
        # 只在描述中也有对应后缀时排除
        return True, f'元器件后缀不同: {suffix_a} vs {suffix_b}'

    # --- 规则 #43: 型号中的规格数字不同 ---
    # 如 CH3N-63 D6 (6A) vs CH3N-63 D63 (63A)
    # 只检查紧随字母后面的不同数字(如D6 vs D63)
    spec_num_a = re.findall(r'([a-zA-Z])(\d+)', da)
    spec_num_b = re.findall(r'([a-zA-Z])(\d+)', db)
    if spec_num_a and spec_num_b:
        # 按"字母+数字"配对，找同字母不同数字
        dict_a = {k: v for k, v in spec_num_a}
        dict_b = {k: v for k, v in spec_num_b}
        for letter in set(dict_a.keys()) & set(dict_b.keys()):
            if dict_a[letter] != dict_b[letter] and dict_a[letter] and dict_b[letter]:
                return True, f'规格数字不同: {letter}{dict_a[letter]} vs {letter}{dict_b[letter]}'

    # --- 规则 #44: 描述中的参数值不同（KEY=VALUE 模式） ---
    # 如 L=285 vs L=277、W=400 vs W=4000、H=200 vs H=300
    kv_a = set(re.findall(r'([A-Za-z一二三四五六七八九十]+)[=：:]\s*(\d+(?:\.\d+)?)', da))
    kv_b = set(re.findall(r'([A-Za-z一二三四五六七八九十]+)[=：:]\s*(\d+(?:\.\d+)?)', db))
    if kv_a and kv_b:
        kv_dict_a = {k: v for k, v in kv_a}
        kv_dict_b = {k: v for k, v in kv_b}
        for key in set(kv_dict_a.keys()) & set(kv_dict_b.keys()):
            if kv_dict_a[key] != kv_dict_b[key]:
                return True, f'参数值不同: {key}={kv_dict_a[key]} vs {key}={kv_dict_b[key]}'

    # --- 规则 #45 (已被规则 #46 替代，保留向后兼容) ---
    # 原规则#45的逻辑已升级为规则#46（品牌命名规则知识库驱动）

    # --- 规则 #48: 描述额外功能段差异 (2026-06-26) ---
    # 一方描述比另一方多了额外的功能/配置标注 = 非重复
    # 如 "iC65N C63A带SD接点" vs "iC65N C63A"
    # 如 "3VA1M400 F/3P 配置相间隔板" vs "3VA1M400 F/3P"
    # 如 "QSA-400/3 400A(柜外操作)" vs "QSA-400/3 400A"
    # 方法：较短描述是较长描述的前缀，且差异部分包含中文功能词
    extra_kw = ['带SD', 'SD接点', '带OF', '带MX', '带MN', '配置相间隔板', '相间隔板',
                '柜外操作', '煤改电', '老练试验', '老炼', '切电容', '带防跳',
                '合闸闭锁', '灭弧室', '底部摩擦接地', '6kA', '10kA', '15kA',
                'NA', '带电显示', '加热器', '通讯', '电动操作', '侧装',
                '带报警', '带辅助', '防潮', '防腐', '三防']
    # 检查是否一方是另一方的前缀
    short_desc = da if len(da) < len(db) else db
    long_desc = db if len(da) < len(db) else da
    if long_desc.startswith(short_desc):
        extra_part = long_desc[len(short_desc):].strip().strip('（）() ')
        if extra_part:
            for kw in extra_kw:
                if kw in extra_part:
                    return True, f'额外功能标注差异: {extra_part[:30]}'
            # 如果额外部分包含任何有意义的功能词（长度>=2的中文词）
            if re.search(r'[\u4e00-\u9fff]{2,}', extra_part):
                return True, f'描述额外信息: {extra_part[:30]}'

    # --- 规则 #49: 高压元器件常见差异 (2026-06-26) ---
    # 这些是88对待确认中发现的高频差异模式

    # 49a: 开闭数量不同 (三开三闭 vs 五开五闭 vs 4开4闭 等)
    sw_a = re.findall(r'(\d+)开(\d+)闭', da)
    sw_b = re.findall(r'(\d+)开(\d+)闭', db)
    if sw_a and sw_b and sw_a != sw_b:
        return True, f'开闭数量不同: {sw_a} vs {sw_b}'

    # 49b: 触臂材质 有vs无 (铜触臂/铝触臂/铜触头)
    arm_kw = ['铜触臂', '铝触臂', '铜触头', '铝触头']
    arm_a = [k for k in arm_kw if k in da]
    arm_b = [k for k in arm_kw if k in db]
    if arm_a != arm_b:
        return True, f'触臂材质差异: {arm_a or "无"} vs {arm_b or "无"}'

    # 49c: 局放要求 有vs无 (局放小于3pC 等)
    pd_a = '局放' in da or '局放' in na
    pd_b = '局放' in db or '局放' in nb
    if pd_a != pd_b:
        return True, f'局放要求: 有vs无'

    # 49d: 安装位置互斥 (装中部/装柜后部/装柜前部 vs 左操/右操/正装)
    pos_a = set(k for k in ['装中部', '装柜后部', '装柜前部', '装柜顶部'] if k in da)
    pos_b = set(k for k in ['装中部', '装柜后部', '装柜前部', '装柜顶部'] if k in db)
    if pos_a and pos_b and pos_a != pos_b:
        return True, f'安装位置不同: {pos_a} vs {pos_b}'
    # 一方有安装位置，另一方有操作方向（不是同一个维度）
    if pos_a and not pos_b and any(k in db for k in ['左操', '右操', '正装']):
        return True, f'安装位置vs操作方向: {pos_a} vs 操作方向'
    if pos_b and not pos_a and any(k in da for k in ['左操', '右操', '正装']):
        return True, f'安装位置vs操作方向: {pos_b} vs 操作方向'

    # 49e: 母线侧接地 有vs无
    bus_gnd_a = '母线侧接地' in da
    bus_gnd_b = '母线侧接地' in db
    if bus_gnd_a != bus_gnd_b:
        return True, f'母线侧接地: 有vs无'

    # 49f: 名称明确不同的产品类型（非辅材）
    product_conflicts = [
        ('APF', 'SVG'), ('百叶窗', '通风过滤'), ('变频器', '固态继电器'),
        ('快拧头', '快插头'), ('被动', '主动'), ('定向脚轮', '万向脚轮'),
        ('变光护目镜', '护目镜'),
        ('双电源', '双投开关'), ('双电源', '隔离开关'),
        ('电操附件', '塑壳断路器'), ('分闸锁', '欠电压'), ('欠电压脱扣器', '分闸锁'),
        ('微型断路器', '浪涌保护器'), ('电能表', '电流表'),
        ('消谐', '电压互感器'), ('转换开关', '智能操控装置'),
    ]
    for pa, pb in product_conflicts:
        if (pa in na and pb in nb) or (pa in nb and pb in na):
            return True, f'不同产品类型: {na} vs {nb}'

    # 49g: 电缆阻燃前缀 ZC-/ZA-/ZA- 有vs无
    flame_a = re.search(r'^(Z[ABC]-)', da)
    flame_b = re.search(r'^(Z[ABC]-)', db)
    if flame_a and not flame_b:
        return True, f'阻燃等级: {flame_a.group(1)} vs 无'
    if flame_b and not flame_a:
        return True, f'阻燃等级: {flame_b.group(1)} vs 无'

    # --- 规则 #47: 型号解析器比较 (2026-06-26) ---
    # 基于品牌知识库解析型号，逐字段比较（壳架/分断/极数/安装/接线/脱扣器/附件）
    _pa = _parse_model(da, na)
    _pb = _parse_model(db, nb)
    _r47 = _compare_parsed(_pa, _pb)
    if _r47[0]:
        return True, f'[规则#47] {_r47[1]}'

    # --- 规则 #46: 品牌命名规则知识库驱动后缀检测 (2026-06-26) ---
    # 基于11品牌51系列的实际选型手册数据，检测附件/安装/接线后缀差异
    # 替代旧规则#45的硬编码5个后缀，现在覆盖所有已知的品牌后缀代号
    result = _detect_suffix_diff(da, db, ta, tb)
    if result:
        return True, result

    # --- 规则 #50: 名称语义不同的产品（2026-06-27 Dongfang 反馈） ---
    # 这些名称虽然部分相似但完全不同产品
    semantic_diff_pairs = [
        ('传感器支架', '气缸固定座'), ('传感器支架', '气缸'),
        ('铜排包扣', '绝缘子包扣'), ('铜排包扣', '电缆头包扣'),
        ('绝缘子包扣', '电缆头包扣'),
        ('低压电缆四指套', '低压电缆'),
        ('型材', '三角板'),
        ('油漆', '塑粉'), ('油漆', '自喷漆'), ('塑粉', '自喷漆'),
        ('玻璃', '小母线端子'),
        ('金属波纹管', '钢丝加强软管'),
    ]
    for s1, s2 in semantic_diff_pairs:
        if (s1 in na and s2 in nb) or (s2 in na and s1 in nb):
            return True, f'不同产品: {s1} vs {s2}'

    # --- 规则 #51: 一方有颜色, 另一方无颜色标注 = 非重复（2026-06-27） ---
    # 如 FY1-D 黄 vs FY1-D, RVV 3×1.5 白 vs RVV 3×1.5
    # 仅在连接片、线缆、标识牌等类别下生效
    color_only_words = ['黄', '红', '绿', '蓝', '白', '黑', '橙', '灰']
    cw_a_only = [c for c in color_only_words if c in (na + da) and c not in (nb + db)]
    cw_b_only = [c for c in color_only_words if c in (nb + db) and c not in (na + da)]
    if (cw_a_only or cw_b_only):
        # 检查是否在适用类别
        applicable = any(k in na + nb for k in ['连接片', '护套线', '护线套', '导线', '电缆'])
        if applicable:
            color = (cw_a_only[0] if cw_a_only else cw_b_only[0])
            return True, f'颜色标注差异: {color}'

    # --- 规则 #52: 配套件 vs 主件 = 非重复（2026-06-27） ---
    # 如 行程开关-护套 vs 行程开关
    if ('-护套' in na and '护套' not in nb and '行程开关' in nb) or \
       ('-护套' in nb and '护套' not in na and '行程开关' in na):
        return True, '配套件 vs 主件'

    # --- 规则 #53: Acti9 vs Acti9 Pro = 非重复（2026-06-29） ---
    # Acti9 Pro 是施耐德第五代升级产品线，与 Acti9 型号相同但产品号/参数不同
    # 一方描述含 'Pro' 且另一方不含，且都是施耐德 iC65/iC60/iDPN/C120 等 Acti9 系列产品
    acti9_models = ['iC65N', 'iC65H', 'iC65L', 'iC60N', 'iC60H', 'iC65N-S', 'iC65H-S',
                    'iDPNa', 'iDPNN', 'iDPNH', 'iDPNK2', 'C120H', 'C120L', 'iINT125',
                    'Vigi iC65', 'Vigi iDPN', 'iID', 'iCT', 'iTL', 'RCA', 'ARA']
    has_acti9_a = any(m in ta for m in acti9_models)
    has_acti9_b = any(m in tb for m in acti9_models)
    if has_acti9_a and has_acti9_b:
        pro_a = bool(re.search(r'\bPro\b', ta))
        pro_b = bool(re.search(r'\bPro\b', tb))
        if pro_a != pro_b:
            return True, 'Acti9 vs Acti9 Pro（不同产品线/世代）'

    return False, ''


def _detect_suffix_diff(da, db, ta, tb):
    """
    基于品牌命名规则知识库，检测描述中的附件/安装/接线后缀差异。
    返回非空字符串表示判定为非重复（含原因），返回空字符串表示未检测到差异。

    核心原则：后缀差异 = 一方有某附件/安装/接线代号，另一方没有。
    但必须排除极数标注（3P/4P/3D）、分断能力代号（F/N/H/S）、
    脱扣器代号（TM/MA/TMF）、壳架代号（XT1/XT2 等）等非附件上下文。
    """
    # 统一大小写
    da_u, db_u = da.upper(), db.upper()
    combined_a = da_u + ' ' + ta.upper()
    combined_b = db_u + ' ' + tb.upper()

    # 排除极数上下文中的 P/D：\dP 或 \dD（如 3P, 4P, 3D）不是后缀
    # 但 /3F 中 F 不是极数（极数是 P），所以 /3F 中的 F 是附件后缀
    pole_pattern = re.compile(r'\d[PD](?:[\s/,.\u4e00-\u9fff]$)')

    # 排除分断能力代号上下文：在型号核心部分中的 F/N/H/S/B/C/L/M/R/V/W/E/X

    for code, meaning in _ALL_SUFFIX_CODES.items():
        if len(code) == 1:
            # ── 单字母后缀（F/P/R/W/D等）──
            # 只检测以下安全模式：
            # 1) /数字+字母 末尾（如 /3F、/160/3F）
            # 2) 描述末尾独立字母（如 "400A P"）
            # 3) 空格分隔的独立字母（如 "NSX250N 3P OF"）

            # 模式A: /N+字母 在描述末尾位置 (如 /160/3F 中 F)
            pat_a = r'/\d+' + re.escape(code) + r'(?:\s|$|/|[\u4e00-\u9fff])'
            has_a_a = bool(re.search(pat_a, combined_a))
            has_a_b = bool(re.search(pat_a, combined_b))
            if has_a_a != has_a_b:
                return f'后缀{meaning}({code})有vs无: 不同型号'

            # 模式B: 描述末尾独立字母 (如 "...400A P")
            # 必须排除极数（\dP 在末尾不算）
            pat_b = r'(?:\s|^)' + re.escape(code) + r'\s*$'
            has_b_a = bool(re.search(pat_b, da_u.strip()))
            has_b_b = bool(re.search(pat_b, db_u.strip()))
            if has_b_a != has_b_b:
                # 额外检查：排除极数误触发（如描述末尾的 3P）
                if code == 'P':
                    # 检查匹配位置前面是不是数字（极数）
                    if has_b_a and re.search(r'\dP\s*$', da):
                        continue  # 极数P，跳过
                    if has_b_b and re.search(r'\dP\s*$', db):
                        continue
                if code == 'D':
                    if has_b_a and re.search(r'\dD\s*$', da):
                        continue
                    if has_b_b and re.search(r'\dD\s*$', db):
                        continue
                return f'后缀{meaning}({code})有vs无: 不同型号'

            # 模式C: 空格分隔的独立字母，且不在极数上下文中
            # 只对已知附件后缀 F/R/W/D（不是P，P太容易误触发）
            if code in ('F', 'R', 'W', 'D'):
                pat_c = r'(?:\s|^)' + re.escape(code) + r'(?:\s|$|[,，)）])'
                has_c_a = bool(re.search(pat_c, combined_a))
                has_c_b = bool(re.search(pat_c, combined_b))
                if has_c_a != has_c_b:
                    return f'后缀{meaning}({code})有vs无: 不同型号'

        elif len(code) == 2:
            # ── 双字母后缀（OF/SD/MX/MN/EL/RC/FC/EF/ES/AL/AX 等）──
            pat = r'(?:\s|^|/)' + re.escape(code) + r'(?:\s|$|/|[,，)）])'
            has_a = bool(re.search(pat, combined_a))
            has_b = bool(re.search(pat, combined_b))
            if has_a != has_b:
                return f'{meaning}({code})有vs无: 不同型号'

        elif len(code) >= 3:
            # ── 3+字母后缀（SDE/UVT/SHT/MOE/RHD/IFM/COM 等）──
            pat = r'(?:\s|^|/)' + re.escape(code) + r'(?:\s|$|/|[,，)）])'
            has_a = bool(re.search(pat, combined_a))
            has_b = bool(re.search(pat, combined_b))
            if has_a != has_b:
                return f'{meaning}({code})有vs无: 不同型号'

    return ''


def desc_numbers_match(a, b):
    """检查两个物料描述+名称中的数字集合是否一致"""
    nums_a = sorted(set(re.findall(r'\d+(?:\.\d+)?', a['desc'] + ' ' + a['name'])))
    nums_b = sorted(set(re.findall(r'\d+(?:\.\d+)?', b['desc'] + ' ' + b['name'])))
    return nums_a == nums_b


# ─── 型号解析器（规则 #47）────────────────────────────────────────────────────
# 基于品牌命名规则JSON，解析物料描述中的型号字段
# 两个物料如果解析结果中任一字段不同 = 非重复

def _build_series_index():
    """构建 {系列前缀 → (品牌, 系列名, 系列数据)} 的索引"""
    index = {}
    for brand, data in BRAND_NAMING_RULES.items():
        if brand == '说明' or not isinstance(data, dict):
            continue
        for series_name, series_data in data.get('系列', {}).items():
            if not isinstance(series_data, dict):
                continue
            sn = series_name.upper()
            # 主前缀：系列名中的纯字母部分 (NSX, CDM, NDM, CM, BM, NM, RMM, TGM...)
            for m in re.finditer(r'([A-Z]{2,})', sn):
                key = m.group(1)
                if len(key) >= 2 and key not in ('MAX', 'TMAX', 'EMAX', 'ACTI'):
                    if key not in index or len(key) > len(str(index[key][0])):
                        index[key] = (series_name, brand, series_data)
            
            # 从壳架等级提取子前缀 (如 'Tmax XT' 的壳架 'XT1(125A)' → XT)
            frames = series_data.get('壳架等级', [])
            if isinstance(frames, list):
                for fr in frames:
                    m2 = re.match(r'^([A-Z]{2,})', str(fr).upper())
                    if m2:
                        key = m2.group(1)
                        if key not in index:
                            index[key] = (series_name, brand, series_data)
            
            # 特殊处理
            if 'XT' in sn and 'XT' not in index:
                index['XT'] = (series_name, brand, series_data)
            if 'IC65' in sn:
                for k in ['IC', 'IC65', 'IC60']:
                    if k not in index:
                        index[k] = (series_name, brand, series_data)
            if 'LC' in sn and 'LC1' in sn:
                if 'LC1' not in index:
                    index['LC1'] = (series_name, brand, series_data)
    return index

_SERIES_INDEX = _build_series_index()

def _parse_model(desc, name=''):
    """
    解析物料描述，提取结构化型号信息。
    返回 dict: {series, frame, breaking_cap, poles, install, connection, trip_unit, attachments}
    """
    text = (desc + ' ' + name).upper().strip()
    result = {
        'series': '', 'frame': '', 'breaking_cap': '',
        'poles': '', 'install': '', 'connection': '',
        'trip_unit': '', 'attachments': set()
    }

    # 1. 识别系列（优先匹配最长前缀）
    matched_series = None
    matched_brand = None
    matched_data = None
    matched_len = 0
    for prefix, (sname, brand, sdata) in _SERIES_INDEX.items():
        if text.startswith(prefix) and len(prefix) > matched_len:
            after = text[len(prefix):len(prefix)+1] if len(text) > len(prefix) else ''
            if after == '' or after.isdigit() or after in ('-', '/', ' ', '\t'):
                matched_series = sname
                matched_brand = brand
                matched_data = sdata
                matched_len = len(prefix)
                result['series'] = prefix

    # 对于 CM3/CDM3/NDM3 等，前缀索引只匹配到CM/CDM/NDM
    # 需要把系列编号也纳入，避免壳架数字提取错误
    if result['series'] and matched_data:
        sn = matched_series.upper()
        sn_num = re.match(r'^[A-Z]+(\d+)', sn)
        if sn_num:
            series_num = sn_num.group(1)
            remaining_after = text[len(result['series']):]
            if remaining_after.startswith(series_num):
                result['series'] = result['series'] + series_num

    if not matched_data:
        return result

    # 2. 提取壳架电流
    remaining = text[len(result['series']):].lstrip('- /\t')
    # 壳架数字是剩余文本中第一个2-4位数字（跳过可能的单数字系列编号）
    # 但系列编号已在上面处理（CM3的3已被纳入series），所以直接匹配
    fm = re.match(r'(\d{2,4})', remaining)
    if fm:
        result['frame'] = fm.group(1)
    else:
        # 可能有字母在前 (如 XT2H160 → 剩余 2H160)
        fm2 = re.match(r'[A-Z0-9]*?(\d{2,4})', remaining)
        if fm2:
            result['frame'] = fm2.group(1)
        else:
            all_nums = re.findall(r'\b(\d{2,4})\b', remaining)
            if all_nums:
                result['frame'] = max(all_nums, key=lambda x: (len(x), int(x)))

    # 3. 提取分断能力代号
    if isinstance(matched_data.get('分断能力代号'), dict):
        for code in matched_data['分断能力代号']:
            code_u = code.upper()
            found = False
            # 模式1: 壳架数字后紧跟分断字母 (NSX100F → F, CDM3-100C → C)
            if result['frame']:
                pat = re.escape(result['frame']) + r'\s*([A-Z])'
                m = re.search(pat, text)
                if m and m.group(1) == code_u:
                    result['breaking_cap'] = code
                    found = True
                    break
            # 模式2: 系列前缀的数字部分后紧跟分断字母 (XT2H160 → H, XT5S400 → S)
            # 如 XT2H160，前缀XT，后面是 2H160，H是分断
            m2 = re.search(result['series'] + r'(\d)([A-Z])', text)
            if m2 and m2.group(2) == code_u:
                # 确认这不是壳架数字的一部分
                # 检查这个字母后面是否跟更多数字(壳架电流)
                after_letter = text[m2.end():m2.end()+3] if m2.end() < len(text) else ''
                if re.match(r'\d', after_letter):
                    result['breaking_cap'] = code
                    found = True
                    break
            # 模式3: -数字后紧跟分断字母 (-100S/3300 → S)
            if not found:
                m3 = re.search(r'\-\s*\d+\s*([A-Z])', text)
                if m3 and m3.group(1) == code_u:
                    result['breaking_cap'] = code
                    break

    # 4. 提取极数
    pole_match = re.findall(r'(\d)P', text)
    if pole_match:
        result['poles'] = '/'.join(sorted(set(pole_match)))
    # 也检查 3P3D 等复合形式
    pole3d = re.search(r'(\d)P(\d)D', text)
    if pole3d:
        result['poles'] = f'{pole3d.group(1)}P{pole3d.group(2)}D'

    # 5. 提取安装方式
    if isinstance(matched_data.get('安装方式'), dict):
        for code in matched_data['安装方式']:
            if len(code) == 1 and code.isalpha():
                # 检查描述末尾或空格后的单字母
                if re.search(r'(?:\s|^|/)' + code + r'(?:\s|$)', text):
                    result['install'] = code
                    break

    # 6. 提取接线方式
    if isinstance(matched_data.get('接线方式'), dict):
        for code in matched_data['接线方式']:
            if len(code) >= 1 and code.isalpha():
                if re.search(r'(?:\s|^|/)' + re.escape(code) + r'(?:\s|$|/)', text):
                    result['connection'] = code
                    break

    # 7. 提取脱扣器类型
    if isinstance(matched_data.get('脱扣器类型'), dict):
        # 按长度降序匹配（先匹配长的，如TMF优先于TM）
        for code in sorted(matched_data['脱扣器类型'].keys(), key=lambda x: len(x.replace('-', '')), reverse=True):
            code_clean = code.replace('-', '').upper()
            # 在描述中搜索（空格分隔或行首行尾）
            if re.search(r'(?:\s|^|/)' + re.escape(code_clean) + r'(?:\s|$|/|[,，)）.])', text):
                result['trip_unit'] = code_clean
                break

    # 8. 提取附件代号
    if isinstance(matched_data.get('附件代号'), dict):
        for code in matched_data['附件代号']:
            if len(code) >= 2:
                if re.search(r'\b' + re.escape(code) + r'\b', text):
                    result['attachments'].add(code)

    return result

def _compare_parsed(pa, pb):
    """
    比较两个解析结果，返回 (is_nondup: bool, reason: str)
    只在双方都能识别系列时才比较。
    """
    if not pa['series'] or not pb['series']:
        return False, ''  # 无法解析，交给其他规则

    # 不同系列 → 不比较（交给其他规则）
    if pa['series'] != pb['series']:
        return False, ''

    # 同系列，逐字段比较
    # 壳架不同
    if pa['frame'] and pb['frame'] and pa['frame'] != pb['frame']:
        return True, f'壳架不同: {pa["series"]}{pa["frame"]} vs {pb["series"]}{pb["frame"]}'

    # 分断能力不同
    if pa['breaking_cap'] and pb['breaking_cap'] and pa['breaking_cap'] != pb['breaking_cap']:
        return True, f'分断能力不同: {pa["breaking_cap"]} vs {pb["breaking_cap"]}'

    # 极数不同
    if pa['poles'] and pb['poles'] and pa['poles'] != pb['poles']:
        return True, f'极数不同: {pa["poles"]} vs {pb["poles"]}'

    # 安装方式不同
    if pa['install'] and pb['install'] and pa['install'] != pb['install']:
        return True, f'安装方式不同: {pa["install"]} vs {pb["install"]}'

    # 接线方式不同
    if pa['connection'] and pb['connection'] and pa['connection'] != pb['connection']:
        return True, f'接线方式不同: {pa["connection"]} vs {pb["connection"]}'

    # 脱扣器不同
    if pa['trip_unit'] and pb['trip_unit'] and pa['trip_unit'] != pb['trip_unit']:
        return True, f'脱扣器不同: {pa["trip_unit"]} vs {pb["trip_unit"]}'

    # 附件差异（一方有另一方没有）
    att_diff = pa['attachments'] ^ pb['attachments']
    if att_diff:
        only_a = pa['attachments'] - pb['attachments']
        only_b = pb['attachments'] - pa['attachments']
        if only_a:
            return True, f'附件差异: 一方有{only_a}'
        if only_b:
            return True, f'附件差异: 一方有{only_b}'

    return False, ''


# ─── 语义配对辅助函数 ────────────────────────────────────────────────────────

def _get_name_core(name):
    """从名称中提取核心产品词（去掉编号、数字、修饰词）"""
    s = name
    s = re.sub(r'[/\\].*?(?=\s|$)', '', s)          # 去掉 / 及后面内容
    s = s.replace('+底座', '').replace('+', '')       # 去 +附件
    s = re.sub(r'[a-zA-Z]+\d+[a-zA-Z0-9\-]*', '', s)  # 去字母数字编号
    s = re.sub(r'\d{2,}', '', s)                       # 去2位以上数字
    s = re.sub(r'-?\d*高$', '', s)                     # 去 -42高 后缀
    for p in ['通用', '老式', '新型', '标准', '非标', '含老式']:
        s = s.replace(p, '')
    s = re.sub(r'^[-\d]+', '', s)                      # 去开头 -和数字
    return s.strip()


def _extract_product_numbers(name):
    """名称中产品编号: 3-4位独立数字 + 字母数字组合"""
    nums = set(re.findall(r'\d{3,4}', name))
    alphas = set(re.findall(r'[a-zA-Z]+\d+[a-zA-Z0-9\-]*', name))
    return nums | alphas


def _extract_model_tokens(text):
    """文本中的型号token（>=3字符，归一化后）"""
    # 先把中文数字替换为阿拉伯数字（CXJG一11 → CXJG111）
    text = text.replace('一', '1').replace('二', '2').replace('三', '3').replace('四', '4').replace('五', '5').replace('六', '6').replace('七', '7').replace('八', '8').replace('九', '9').replace('零', '0')
    tokens = re.findall(r'[a-zA-Z]{2,}[\-\.]?[\d]*[\-\.]?[\d]*[\-\.]?[\d]*[a-zA-Z0-9\-]*', text)
    result = set()
    for t in tokens:
        t_norm = t.lower().replace('-', '').replace('.', '')
        if len(t_norm) >= 4:
            result.add(t_norm)
    # 编号模式如 5SQ.GS-4-3
    codes = re.findall(r'[a-zA-Z\d]+\.[a-zA-Z\d]+[\-][\d]+[\-]?[\d]*', text)
    for c in codes:
        result.add(c.lower().replace('.', '').replace('-', ''))
    return result


def _all_significant_numbers(text):
    """文本中所有>=2位数字"""
    return set(re.findall(r'\d{2,}', text))


def _core_names_related(core_a, core_b):
    """两个核心词是否有包含/重叠关系"""
    if not core_a or not core_b:
        return False
    # 直接包含
    if core_a in core_b or core_b in core_a:
        return True
    # 中文字符重叠 >= 60%
    chars_a = set(re.findall(r'[\u4e00-\u9fff]', core_a))
    chars_b = set(re.findall(r'[\u4e00-\u9fff]', core_b))
    if not chars_a or not chars_b:
        return False
    overlap = len(chars_a & chars_b) / min(len(chars_a), len(chars_b))
    return overlap >= 0.6 and len(chars_a & chars_b) >= 2


# ─── 配对 + 分类 ────────────────────────────────────────────────────────────

def find_pairs(materials):
    """找出候选重复配对（语义匹配版）"""
    seen = set()
    pairs = []

    # ---- 索引: 按类别分组 ----
    by_cat_sub = defaultdict(list)
    for m in materials:
        by_cat_sub[(m['cat'], m['subcat'])].append(m)
    by_cat = defaultdict(list)
    for m in materials:
        by_cat[m['cat']].append(m)

    # ---- 层1: 同名 + 描述ultra一致 或 仅微小格式差异 ----
    name_groups = defaultdict(list)
    for m in materials:
        name_groups[(m['cat'], m['subcat'], m['name'])].append(m)

    c1 = 0
    for key, grp in name_groups.items():
        if len(grp) < 2:
            continue
        # 1a: ultra(desc) 完全一致
        desc_map = defaultdict(list)
        for m in grp:
            desc_map[ultra(m['desc'])].append(m)
        for dk, items in desc_map.items():
            if len(items) < 2:
                continue
            for i in range(len(items)):
                for j in range(i + 1, len(items)):
                    pk = tuple(sorted([items[i]['code'], items[j]['code']]))
                    if pk not in seen:
                        seen.add(pk)
                        pairs.append((items[i], items[j]))
                        c1 += 1

        # 1b: ultra微小差异 — 分两种情况:
        #     小组(<=50): 允许差异<=10%
        #     大组(>50): 只允许子串关系
        ultras = [(ultra(m['desc']), m) for m in grp]
        is_small = len(grp) <= 50
        for i in range(len(ultras)):
            for j in range(i + 1, len(ultras)):
                ua, ma = ultras[i]
                ub, mb = ultras[j]
                if ua == ub:
                    continue  # 已在1a处理
                # 子串关系 (始终允许)
                if ua in ub or ub in ua:
                    pk = tuple(sorted([ma['code'], mb['code']]))
                    if pk not in seen:
                        seen.add(pk)
                        pairs.append((ma, mb))
                        c1 += 1
                    continue
                # 相似度 (仅小组)
                if is_small:
                    shorter = min(len(ua), len(ub))
                    if shorter > 0:
                        from difflib import SequenceMatcher as _SM
                        ratio = _SM(None, ua, ub).ratio()
                        if 1 - ratio <= 0.05:
                            pk = tuple(sorted([ma['code'], mb['code']]))
                            if pk not in seen:
                                seen.add(pk)
                                pairs.append((ma, mb))
                                c1 += 1

        # 1c: 大组中捡回 — 同名+核心型号完全一致+ultra子串
        # (不用ultra相似度, 太松)
        if len(grp) > 50:
            main_models = {}
            for m in grp:
                tokens = _extract_model_tokens(m['name'] + ' ' + m['desc'])
                main_t = ''
                for t in tokens:
                    if not re.match(r'^(ac|dc)?\\d+(v|a)$', t, re.I) and len(t) > len(main_t):
                        main_t = t
                if main_t:
                    main_models[m['code']] = (main_t, m)
            mm_groups = defaultdict(list)
            for code, (mm, m) in main_models.items():
                mm_groups[mm].append(m)
            for mm, items in mm_groups.items():
                if len(items) < 2 or len(items) > 20:
                    continue
                mm_ultras = [(ultra(m['desc']), m) for m in items]
                for i in range(len(mm_ultras)):
                    for j in range(i+1, len(mm_ultras)):
                        ua, ma = mm_ultras[i]
                        ub, mb = mm_ultras[j]
                        if ua == ub or ua in ub or ub in ua:
                            pk = tuple(sorted([ma['code'], mb['code']]))
                            if pk not in seen:
                                seen.add(pk)
                                pairs.append((ma, mb))
                                c1 += 1
    print(f"  层1(同名+描述ultra): {c1} 对")

    # ---- 层1.5: 同义词名称 + 描述ultra一致 ----
    c15 = 0
    # 按 (类别, 子类别, ultra(描述)) 建索引
    desc_group_idx = defaultdict(list)
    for m in materials:
        ud = ultra(m['desc'])
        if ud:
            desc_group_idx[(m['cat'], m['subcat'], ud)].append(m)
    for (cat, subcat, ud), grp in desc_group_idx.items():
        if len(grp) < 2:
            continue
        # 在同一描述组内找名称互为同义词的配对
        for i in range(len(grp)):
            for j in range(i + 1, len(grp)):
                a, b = grp[i], grp[j]
                if a['name'] == b['name']:
                    continue  # 同名的已在层1处理
                if not _are_synonyms(a['name'], b['name']):
                    continue
                pk = tuple(sorted([a['code'], b['code']]))
                if pk not in seen:
                    seen.add(pk)
                    pairs.append((a, b))
                    c15 += 1
    print(f"  层1.5(同义词+描述一致): {c15} 对")

    # ---- 层2: 同子类别 + 核心词包含 + 名称编号交集 ----
    c2 = 0
    for (cat, subcat), grp in by_cat_sub.items():
        if len(grp) < 2:
            continue
        # 按 (核心词, 编号) 建索引
        core_num_idx = defaultdict(list)
        for m in grp:
            core = _get_name_core(m['name'])
            nums = _extract_product_numbers(m['name'])
            if core and nums:
                for n in nums:
                    core_num_idx[(core, n)].append(m)

        # 同 (核心词, 编号) 内配对
        for (core, num), items in core_num_idx.items():
            if len(items) < 2:
                continue
            for i in range(len(items)):
                for j in range(i + 1, len(items)):
                    if items[i]['name'] == items[j]['name']:
                        continue  # 同名的走层1
                    pk = tuple(sorted([items[i]['code'], items[j]['code']]))
                    if pk not in seen:
                        seen.add(pk)
                        pairs.append((items[i], items[j]))
                        c2 += 1

        # 跨核心词但有包含关系 + 编号交集
        cores_in_sub = defaultdict(list)
        for m in grp:
            core = _get_name_core(m['name'])
            if core:
                cores_in_sub[core].append(m)
        core_list = list(cores_in_sub.keys())
        for ci in range(len(core_list)):
            for cj in range(ci + 1, len(core_list)):
                ca, cb = core_list[ci], core_list[cj]
                if not _core_names_related(ca, cb):
                    continue
                items_a = cores_in_sub[ca]
                items_b = cores_in_sub[cb]
                for a in items_a:
                    for b in items_b:
                        if a['name'] == b['name']:
                            continue
                        nums_a = _extract_product_numbers(a['name'])
                        nums_b = _extract_product_numbers(b['name'])
                        # 名称编号交集
                        if nums_a & nums_b:
                            pk = tuple(sorted([a['code'], b['code']]))
                            if pk not in seen:
                                seen.add(pk)
                                pairs.append((a, b))
                                c2 += 1
    print(f"  层2(核心词+编号): {c2} 对")

    # ---- 层3: 描述型号token分组（可跨子类别，限同类别） ----
    # 按型号token建索引，只保留出现次数 <=20 的token（太泛的跳过）
    # 核心词不需要关联——完全靠型号token匹配（可跨子类别）
    c3 = 0
    model_idx = defaultdict(list)
    for m in materials:
        tokens = _extract_model_tokens(m['name'] + ' ' + m['desc'])
        for t in tokens:
            model_idx[t].append(m)

    for token, items in model_idx.items():
        if len(items) < 2 or len(items) > 30:
            continue  # 太泛的token跳过
        for i in range(len(items)):
            for j in range(i + 1, len(items)):
                a, b = items[i], items[j]
                if a['cat'] != b['cat']:
                    continue
                if a['name'] == b['name']:
                    continue  # 同名的走层1
                pk = tuple(sorted([a['code'], b['code']]))
                if pk not in seen:
                    seen.add(pk)
                    pairs.append((a, b))
                    c3 += 1
    print(f"  层3(型号token): {c3} 对")

    # ---- 层4: 核心词包含 + 交叉编号匹配 ----
    # 名称编号在对方全部数字中出现
    c4 = 0
    for (cat, subcat), grp in by_cat_sub.items():
        if len(grp) < 2:
            continue
        cores_in_sub = defaultdict(list)
        for m in grp:
            core = _get_name_core(m['name'])
            if core:
                cores_in_sub[core].append(m)
        core_list = list(cores_in_sub.keys())
        for ci in range(len(core_list)):
            for cj in range(ci + 1, len(core_list)):
                ca, cb = core_list[ci], core_list[cj]
                if not _core_names_related(ca, cb):
                    continue
                for a in cores_in_sub[ca]:
                    for b in cores_in_sub[cb]:
                        if a['name'] == b['name']:
                            continue
                        # A的名称编号在B的全部数字中，或反之
                        nums_a = _extract_product_numbers(a['name'])
                        nums_b = _extract_product_numbers(b['name'])
                        all_nums_a = _all_significant_numbers(a['name'] + ' ' + a['desc'])
                        all_nums_b = _all_significant_numbers(b['name'] + ' ' + b['desc'])
                        cross = (nums_a & (nums_b | all_nums_b)) | (nums_b & (nums_a | all_nums_a))
                        # 必须有交叉且不能是太泛的数字(排除220,380等常见电压)
                        generic = {'220', '380', '110', '24', '48', '12', '10', '50', '60', '100', '200'}
                        specific_cross = cross - generic
                        if specific_cross:
                            pk = tuple(sorted([a['code'], b['code']]))
                            if pk not in seen:
                                seen.add(pk)
                                pairs.append((a, b))
                                c4 += 1
    print(f"  层4(核心词+交叉编号): {c4} 对")

    # ---- 层5: 同名 + 名称编号匹配 (覆盖同名但描述差异大的) ----
    c5 = 0
    for key, grp in name_groups.items():
        if len(grp) < 2:
            continue
        for i in range(len(grp)):
            for j in range(i + 1, len(grp)):
                a, b = grp[i], grp[j]
                nums_a = _extract_product_numbers(a['name'])
                nums_b = _extract_product_numbers(b['name'])
                if nums_a & nums_b:  # 名称编号有交集
                    pk = tuple(sorted([a['code'], b['code']]))
                    if pk not in seen:
                        seen.add(pk)
                        pairs.append((a, b))
                        c5 += 1
    print(f"  层5(同名+名称编号): {c5} 对")

    # ---- 层6: 同子类别 + 核心词关联 + 单向编号匹配 ----
    # 一方名称有编号，该编号在对方的全部数字中
    c6 = 0
    for (cat, subcat), grp in by_cat_sub.items():
        if len(grp) < 2:
            continue
        cores_in_sub = defaultdict(list)
        for m in grp:
            core = _get_name_core(m['name'])
            if core:
                cores_in_sub[core].append(m)
        core_list = list(cores_in_sub.keys())
        # 同核心词组内 + 不同名称
        for core in core_list:
            items = cores_in_sub[core]
            if len(items) < 2:
                continue
            for i in range(len(items)):
                for j in range(i + 1, len(items)):
                    a, b = items[i], items[j]
                    if a['name'] == b['name']:
                        continue  # 同名走层1/5
                    # 至少一方有名称编号，且编号在对方全部数字中
                    nums_a = _extract_product_numbers(a['name'])
                    nums_b = _extract_product_numbers(b['name'])
                    all_a = _all_significant_numbers(a['name'] + ' ' + a['desc'])
                    all_b = _all_significant_numbers(b['name'] + ' ' + b['desc'])
                    generic = {'220', '380', '110', '24', '48', '12', '10', '50', '60', '100', '200'}
                    # A的编号在B中
                    cross_a = (nums_a & (nums_b | all_b)) - generic if nums_a else set()
                    cross_b = (nums_b & (nums_a | all_a)) - generic if nums_b else set()
                    if cross_a or cross_b:
                        pk = tuple(sorted([a['code'], b['code']]))
                        if pk not in seen:
                            seen.add(pk)
                            pairs.append((a, b))
                            c6 += 1
    print(f"  层6(同核心词+单向编号): {c6} 对")

    print(f"📊 候选配对合计: {len(pairs)} 对")
    return pairs


def load_processed_pairs():
    """加载已人工确认的配对库"""
    pp_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'processed_pairs.json')
    if not os.path.exists(pp_path):
        return {}
    with open(pp_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data  # key: 'id1|id2' (sorted), value: 'dup' or 'nondup'


def classify_pairs(pairs):
    """将配对分类为 confirmed / review / nondup"""
    confirmed = []
    review = []
    nondup_count = 0
    skipped_processed = 0

    # 加载已处理配对库
    processed = load_processed_pairs()

    for a, b in pairs:
        # 先检查是否已在已处理库中
        pk = '|'.join(sorted([str(a['code']), str(b['code'])]))
        if pk in processed:
            skipped_processed += 1
            continue
        # 器具/工具类 → 先过非重复规则，剩下的待人工确认
        if any(k in a['subcat'] or k in b['subcat'] for k in ['器具', '工具', '工器具']):
            nd, reason = is_nondup(a, b)
            if nd:
                nondup_count += 1
                continue
            review.append((a, b, '器具/工具类需人工确认'))
            continue

        # 非重复规则检查
        nd, reason = is_nondup(a, b)
        if nd:
            nondup_count += 1
            continue

        # ★ 名称不同的配对 ★
        if a['name'] != b['name']:
            norm_na = re.sub(r'[\s\-_/\\(),;:.，：；、。{}\[\]<>]+', '', a['name'].lower())
            norm_nb = re.sub(r'[\s\-_/\\(),;:.，：；、。{}\[\]<>]+', '', b['name'].lower())

            if norm_na == norm_nb:
                confirmed.append((a, b, '命名不规范（名称仅符号差异）'))
                continue

            # 中文字符交集（字符级，避免「微型漏电断路器」vs「微型断路器」被误判）
            chars_a = set(re.findall(r'[\u4e00-\u9fff]', a['name']))
            chars_b = set(re.findall(r'[\u4e00-\u9fff]', b['name']))
            common_chars = chars_a & chars_b
            min_chars = min(len(chars_a), len(chars_b))
            # 共享字符少于较小名称字符数的50% = 不相关
            # 例外: 描述型号token完全匹配时, 不受此规则限制
            desc_models_common = _extract_model_tokens(a['name'] + ' ' + a['desc']) & \
                                _extract_model_tokens(b['name'] + ' ' + b['desc'])
            if min_chars > 0 and len(common_chars) / min_chars < 0.5 and not desc_models_common:
                nondup_count += 1
                continue

            # 材质差异
            mat_pairs_chk = [
                ('不锈钢', '镀彩'), ('不锈钢', '镀锌'), ('不锈钢', '镀蓝锌'),
                ('镀铝锌', '镀锌'), ('紫铜', '黄铜'), ('铜', '铝'),
                ('不锈钢', '发黑'), ('白锌', '镀彩'),
            ]
            if any((m1 in a['name'] and m2 in b['name']) or (m2 in a['name'] and m1 in b['name'])
                   for m1, m2 in mat_pairs_chk):
                nondup_count += 1
                continue

            # 类型/功能差异
            type_pairs_chk = [
                ('熔芯', '底座'), ('三通', '活接'), ('三通', '直通'), ('定向', '万向'),
                ('手动', '电动'), ('燕尾', '防爆'), ('滚轮', '脚轮'),
                ('轻轨', '重轨'), ('盲板', '凸面'), ('管夹', '垫片'),
                ('车轮', '手轮'), ('驱动大头', '驱动链条'), ('驱动器', '顶升'),
                ('大门', '加热包'), ('绝缘靴', '绝缘手套'),
                # 注意: 电度表/电能表 是同义词，不在此排除
                ('电压', '电流'), ('两孔', '三孔'),
                ('简易版', '全功能'),
            ]
            if any((t1 in a['name'] and t2 in b['name']) or (t2 in a['name'] and t1 in b['name'])
                   for t1, t2 in type_pairs_chk):
                nondup_count += 1
                continue

            # 形状差异
            shape_pairs_chk = [('方形', '圆形'), ('方形', '管装'), ('U型', '对接')]
            if any((s1 in a['name'] and s2 in b['name']) or (s2 in a['name'] and s1 in b['name'])
                   for s1, s2 in shape_pairs_chk):
                nondup_count += 1
                continue

            # 同义词检查：名称虽不同但为同义词 + 描述一致 = 确认重复
            if _are_synonyms(a['name'], b['name']):
                if ultra(a['desc']) == ultra(b['desc']):
                    confirmed.append((a, b, f'确认重复（同义词: {a["name"]}={b["name"]})'))
                    continue
                else:
                    review.append((a, b, f'同义词待确认: {a["name"]}={b["name"]}'))
                    continue

            review.append((a, b, '名称不同但描述一致'))
            continue

        # 名称相同，检查描述
        if ultra(a['desc']) == ultra(b['desc']):
            confirmed.append((a, b, '命名不规范（仅符号/格式差异）'))
            continue

        # 名称相同 + 描述高度相似: ultra一致（仅标点/空格/大小写差异）或完全相同
        if ultra(a['desc']) == ultra(b['desc']):
            confirmed.append((a, b, '命名不规范（仅符号/格式差异）'))
            continue

        if a['desc'] == b['desc']:
            confirmed.append((a, b, '描述完全相同'))
            continue

        # 名称相同 + 描述高度相似: 必须数字一致 + sim >= 0.97（几乎完全一致）
        if sim(a['desc'], b['desc']) >= 0.97 and desc_numbers_match(a, b):
            confirmed.append((a, b, '描述高度相似'))
            continue

        review.append((a, b, '需人工确认'))

    return confirmed, review, nondup_count, skipped_processed


# ─── Excel 输出 ─────────────────────────────────────────────────────────────

BLUE = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
ORANGE = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
RED = Font(color='FF0000', size=10)
BLACK = Font(size=10)
thin = Side(style='thin', color='D0D0D0')
BORDER = Border(top=thin, left=thin, right=thin, bottom=thin)
WRAP = Alignment(vertical='center', wrap_text=True)
CENTER = Alignment(vertical='center')


def make_sheet(wb_out, title, data):
    """生成一个 Excel sheet"""
    ws_out = wb_out.create_sheet(title)
    for ci, (fname, _) in enumerate(FIELDS, 1):
        c = ws_out.cell(row=1, column=ci, value=fname)
        c.font = Font(bold=True, size=10)
        c.alignment = CENTER
        c.border = BORDER

    ri = 2
    for idx, (a, b, reason) in enumerate(data, 1):
        for label, mat, fill in [(f'A-{idx}', a, BLUE), (f'B-{idx}', b, ORANGE)]:
            for ci, (_, key) in enumerate(FIELDS, 1):
                if key == 'label':
                    val = label
                elif key == 'remark':
                    val = reason
                else:
                    val = mat.get(key, '')
                    if isinstance(val, float) and val == int(val):
                        val = int(val)
                cell = ws_out.cell(row=ri, column=ci, value=val)
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = WRAP if key == 'desc' else CENTER
                if key in DIFF_FIELDS:
                    va = str(a.get(key, '') or '').strip()
                    vb = str(b.get(key, '') or '').strip()
                    cell.font = RED if va != vb else BLACK
                else:
                    cell.font = BLACK
            ri += 1

    # 自动调整列宽：根据表头和所有数据行的内容长度取最大值
    for ci in range(1, len(FIELDS) + 1):
        max_len = len(str(ws_out.cell(row=1, column=ci).value or ''))
        for row in range(2, ri):
            val = str(ws_out.cell(row=row, column=ci).value or '')
            # 取单元格内容最长行（按换行拆分后最长行），避免超长描述撑爆
            for line in val.split('\n'):
                if len(line) > max_len:
                    max_len = len(line)
        # 字符宽度 → Excel 列宽（中文约2个字符宽，加padding）
        col_width = min(max_len * 1.3 + 3, 60)
        col_width = max(col_width, 8)  # 最小8
        ws_out.column_dimensions[get_column_letter(ci)].width = col_width
    ws_out.freeze_panes = 'C2'
    if ri > 2:
        ws_out.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}{ri - 1}"
    return ws_out


def generate_excel(confirmed, review, output_path):
    """生成结果 Excel"""
    print(f"\n📝 生成 Excel: {output_path}")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    make_sheet(wb_out, '确认重复', confirmed)
    make_sheet(wb_out, '待人工确认', review)
    wb_out.save(output_path)
    print(f"✅ 已保存")


# ─── 对比报告 ───────────────────────────────────────────────────────────────

def read_pairs_from_sheet(file_path, sheet_name):
    """从结果 Excel 中读取配对"""
    if not os.path.exists(file_path):
        print(f"❌ 文件不存在: {file_path}")
        sys.exit(1)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]

    pairs = []
    current_pair = []
    for row in range(2, ws.max_row + 1):
        mark = str(ws.cell(row=row, column=1).value or '').strip()
        code = str(ws.cell(row=row, column=2).value or '').strip()
        if mark.startswith('A-') or mark.startswith('B-'):
            if current_pair and len(current_pair) >= 2:
                pairs.append(current_pair)
                current_pair = []
            current_pair.append({
                'mark': mark, 'code': code,
                'name': str(ws.cell(row=row, column=3).value or '').strip(),
                'desc': str(ws.cell(row=row, column=4).value or '').strip(),
                'cat': str(ws.cell(row=row, column=5).value or '').strip(),
            })
        else:
            continue
    if current_pair and len(current_pair) >= 2:
        pairs.append(current_pair)
    wb.close()
    return pairs


def get_pair_key(pair):
    """物料对的唯一标识（按物料编号排序）"""
    return tuple(sorted([pair[0]['code'], pair[1]['code']]))


def compare_reports(old_path, new_path, output_path=None):
    """对比两次检查结果，生成变更报告"""
    print(f"📊 对比物料重复检查结果")
    print(f"   旧文件: {old_path}")
    print(f"   新文件: {new_path}")

    sheets_to_check = ['确认重复', '待人工确认']
    report_lines = []

    # 提取日期
    def extract_date(path):
        m = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(path))
        return m.group(1) if m else 'unknown'

    old_date = extract_date(old_path)
    new_date = extract_date(new_path)

    report_lines.append(f"📊 物料重复检查对比报告")
    report_lines.append(f"📅 {old_date} → {new_date}")
    report_lines.append("=" * 60)

    total_old = 0
    total_new = 0
    total_resolved = 0
    total_new_added = 0
    total_unresolved = 0

    for sheet_name in sheets_to_check:
        old_pairs = read_pairs_from_sheet(old_path, sheet_name)
        new_pairs = read_pairs_from_sheet(new_path, sheet_name)

        old_dict = {get_pair_key(p): p for p in old_pairs}
        new_dict = {get_pair_key(p): p for p in new_pairs}

        resolved = [(k, v) for k, v in old_dict.items() if k not in new_dict]
        new_added = [(k, v) for k, v in new_dict.items() if k not in old_dict]
        unresolved = [(k, old_dict[k], new_dict[k]) for k in old_dict if k in new_dict]

        total_old += len(old_pairs)
        total_new += len(new_pairs)
        total_resolved += len(resolved)
        total_new_added += len(new_added)
        total_unresolved += len(unresolved)

        report_lines.append(f"\n### {sheet_name} ###")
        report_lines.append(f"  上次: {len(old_pairs)} 对 | 本次: {len(new_pairs)} 对")
        report_lines.append(f"  ✅ 已解决: {len(resolved)} 对")
        if resolved:
            for k, p in resolved[:10]:
                report_lines.append(f"    - {p[0]['code']} ({p[0]['name']}) vs {p[1]['code']} ({p[1]['name']})")
            if len(resolved) > 10:
                report_lines.append(f"    ... 还有 {len(resolved) - 10} 对")
        report_lines.append(f"  🆕 新增: {len(new_added)} 对")
        if new_added:
            for k, p in new_added[:10]:
                report_lines.append(f"    - {p[0]['code']} ({p[0]['name']}) vs {p[1]['code']} ({p[1]['name']})")
            if len(new_added) > 10:
                report_lines.append(f"    ... 还有 {len(new_added) - 10} 对")
        report_lines.append(f"  ⚠️ 未解决: {len(unresolved)} 对")

    report_lines.append(f"\n{'=' * 60}")
    report_lines.append(f"📈 总计:")
    report_lines.append(f"  上次: {total_old} 对 → 本次: {total_new} 对")
    report_lines.append(f"  ✅ 已解决 {total_resolved} | 🆕 新增 {total_new_added} | ⚠️ 未解决 {total_unresolved}")

    report = '\n'.join(report_lines)
    print(report)

    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(report)
        print(f"\n📄 报告已保存: {output_path}")

    return report


# ─── 自动模式 ───────────────────────────────────────────────────────────────

def find_latest_export(data_dir='IN3数据'):
    """找到最新的物料主数据导出文件"""
    pattern = re.compile(r'物料主数据导出结果.*\.xlsx$')
    candidates = []
    for f in os.listdir(data_dir):
        if pattern.match(f) and 'v2' not in f.lower():
            path = os.path.join(data_dir, f)
            candidates.append((os.path.getmtime(path), path, f))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def find_latest_result(data_dir='IN3数据'):
    """找到最新的可疑重复物料结果文件"""
    pattern = re.compile(r'可疑重复物料-\d{4}-\d{2}-\d{2}-v3\.xlsx$')
    candidates = []
    for f in os.listdir(data_dir):
        if pattern.match(f):
            path = os.path.join(data_dir, f)
            candidates.append((os.path.getmtime(path), path, f))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


# ─── CLI 入口 ───────────────────────────────────────────────────────────────

def cmd_analyze(args):
    """分析命令"""
    materials = load_materials(args.input)
    pairs = find_pairs(materials)
    print(f"📊 候选配对: {len(pairs)} 对")

    confirmed, review, nondup_count, skipped_processed = classify_pairs(pairs)

    print(f"\n📈 分类结果:")
    print(f"  ⏭️ 已处理（跳过）: {skipped_processed}")
    print(f"  ❌ 非重复（已排除）: {nondup_count}")
    print(f"  ✅ 确认重复: {len(confirmed)}")
    print(f"  ❓ 待人工确认: {len(review)}")

    # 自动生成输出路径
    output = args.output
    if not output:
        m = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(args.input))
        date_str = m.group(1) if m else time.strftime('%Y-%m-%d')
        output = os.path.join(os.path.dirname(args.input) or '.', f'可疑重复物料-{date_str}-v3.xlsx')
        print(f"   自动输出路径: {output}")

    generate_excel(confirmed, review, output)
    return output


def cmd_analyze_incremental(args):
    """增量分析模式：只查当天新建/修改物料 vs 全量历史"""
    import datetime
    materials = load_materials(args.input)
    print(f"\n📅 增量模式：筛选当天新建/修改物料")

    # 从文件名提取日期
    m = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(args.input))
    if m:
        target_date = m.group(1)
    else:
        target_date = datetime.date.today().strftime('%Y-%m-%d')
    print(f"   目标日期: {target_date}")

    # 筛选当天新建/修改的物料
    # IN3 导出日期格式可能是 YYYY/MM/DD 或 YYYY-MM-DD
    target_date_alt = target_date.replace('-', '/')
    today_materials = []
    old_materials = []
    for mat in materials:
        create_date = str(mat.get('create_date', ''))[:10]
        modify_date = str(mat.get('modify_date', ''))[:10]
        if target_date in create_date or target_date in modify_date or \
           target_date_alt in create_date or target_date_alt in modify_date:
            today_materials.append(mat)
        else:
            old_materials.append(mat)

    print(f"   当天新建/修改: {len(today_materials)} 条")
    print(f"   历史物料: {len(old_materials)} 条")

    if not today_materials:
        print("\n✅ 当天无新建/修改物料，跳过查重")
        return None

    # 对当天物料与全量历史做配对检查
    # 方法：把当天物料和历史物料放一起跑配对，但只保留至少一方是当天物料的配对
    all_materials = materials  # 全量，确保同组配对能正常工作
    pairs = find_pairs(all_materials)

    # 过滤：只保留至少一方是当天物料的配对
    today_codes = {m['code'] for m in today_materials}
    filtered_pairs = []
    for a, b in pairs:
        if a['code'] in today_codes or b['code'] in today_codes:
            filtered_pairs.append((a, b))

    print(f"📊 候选配对（含当天物料）: {len(filtered_pairs)} 对")

    confirmed, review, nondup_count, skipped_processed = classify_pairs(filtered_pairs)

    print(f"\n📈 分类结果:")
    print(f"  ⏭️ 已处理（跳过）: {skipped_processed}")
    print(f"  ❌ 非重复（已排除）: {nondup_count}")
    print(f"  ✅ 确认重复: {len(confirmed)}")
    print(f"  ❓ 待人工确认: {len(review)}")

    output = args.output
    if not output:
        output = os.path.join(os.path.dirname(args.input) or '.', f'增量查重结果-{target_date}.xlsx')
        print(f"   自动输出路径: {output}")

    if confirmed or review:
        generate_excel(confirmed, review, output)
    else:
        print("✅ 无重复，不生成文件")
    return output if (confirmed or review) else None


def cmd_compare(args):
    """对比命令"""
    output = args.output
    if not output:
        m = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(args.new))
        date_str = m.group(1) if m else time.strftime('%Y-%m-%d')
        output = os.path.join(os.path.dirname(args.new) or '.', f'对比报告-{date_str}.txt')
    compare_reports(args.old, args.new, output)
    return output


def cmd_auto(args):
    """自动模式：找最新文件，分析+对比"""
    data_dir = args.dir or 'IN3数据'
    print(f"🤖 自动模式，数据目录: {data_dir}")

    latest_export = find_latest_export(data_dir)
    if not latest_export:
        print(f"❌ 在 {data_dir} 中找不到物料主数据导出文件")
        sys.exit(1)

    print(f"📦 最新导出文件: {os.path.basename(latest_export)}")

    if getattr(args, 'incremental', False):
        print("🔍 增量模式启动")
        output = cmd_analyze_incremental(argparse.Namespace(input=latest_export, output=None))
        return output

    # 先分析
    output = cmd_analyze(argparse.Namespace(input=latest_export, output=None))

    # 如果有上次的结果，自动对比
    prev_result = find_latest_result(data_dir)
    if prev_result and prev_result != output:
        print(f"\n--- 自动对比 ---")
        old_date = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(prev_result))
        new_date = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(output))
        report_path = os.path.join(data_dir, f'对比报告-{new_date.group(1) if new_date else "latest"}.txt')
        compare_reports(prev_result, output, report_path)
    else:
        print("\nℹ️ 没有找到上次的结果文件，跳过对比")

    return output


def main():
    parser = argparse.ArgumentParser(
        description='IN3 物料重复检查工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
子命令:
  analyze   分析物料主数据，输出可疑重复 Excel
  compare   对比两次检查结果，生成变更报告
  auto      自动找最新文件，分析并对比
        """,
    )
    sub = parser.add_subparsers(dest='command', required=True)

    # analyze
    p_analyze = sub.add_parser('analyze', help='分析物料主数据')
    p_analyze.add_argument('input', help='IN3 导出的物料主数据 Excel 文件路径')
    p_analyze.add_argument('-o', '--output', help='输出文件路径（默认自动生成）')

    # compare
    p_compare = sub.add_parser('compare', help='对比两次检查结果')
    p_compare.add_argument('old', help='上次的结果 Excel 文件')
    p_compare.add_argument('new', help='本次的结果 Excel 文件')
    p_compare.add_argument('-o', '--output', help='对比报告输出路径（默认自动生成）')

    # auto
    p_auto = sub.add_parser('auto', help='自动分析+对比')
    p_auto.add_argument('--dir', default='IN3数据', help='数据目录（默认 IN3数据）')
    p_auto.add_argument('--incremental', action='store_true', help='增量模式：只查当天新建/修改物料')

    args = parser.parse_args()

    if args.command == 'analyze':
        cmd_analyze(args)
    elif args.command == 'compare':
        cmd_compare(args)
    elif args.command == 'auto':
        cmd_auto(args)


if __name__ == '__main__':
    main()
