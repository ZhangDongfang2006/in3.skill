#!/usr/bin/env python3
"""
采购价格异常分析 v2
- 近 N 天采购 vs 历史均价
- 去重：同采购订单+同物料编号只保留一行
- 历史对比订单编号放入备注列
"""

import sys
from collections import defaultdict
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def load_data(filepath):
    """加载全量采购数据（从异常报告的明细sheet）"""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb['异常物料明细']
    headers = None
    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        vals = list(row)
        rec = dict(zip(headers, vals))
        records.append(rec)
    wb.close()
    return records

def analyze(records, days=60, threshold=50):
    """分析价格异常"""
    cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    
    recent = []
    historical = defaultdict(list)
    
    for rec in records:
        date = str(rec.get('下单时间', ''))
        if date >= cutoff:
            recent.append(rec)
        else:
            historical[rec['物料编号']].append(rec)
    
    print(f"近期采购 (>= {cutoff}): {len(recent)} 条")
    print(f"历史采购 (< {cutoff}): {sum(len(v) for v in historical.values())} 条")
    
    # Historical stats per 物料编号
    hist_stats = {}
    for mat_id, recs in historical.items():
        prices = [r['含税单价'] for r in recs if r['含税单价'] and r['含税单价'] > 0]
        if not prices:
            continue
        sorted_prices = sorted(prices)
        hist_stats[mat_id] = {
            'avg': sum(prices) / len(prices),
            'median': sorted_prices[len(sorted_prices) // 2],
            'min': min(prices),
            'max': max(prices),
            'count': len(prices),
            'po_refs': '; '.join(
                f"{r['采购订单号']}(¥{r['含税单价']:.2f},{str(r['下单时间'])})"
                for r in recs
            ),
        }
    
    # Check recent purchases
    anomalies = []
    seen = set()
    
    for rec in recent:
        mat_id = rec['物料编号']
        po = rec['采购订单号']
        price = rec['含税单价']
        if not price or price <= 0:
            continue
        
        # Dedup: same PO + same material
        key = (po, mat_id)
        if key in seen:
            continue
        seen.add(key)
        
        stats = hist_stats.get(mat_id)
        if not stats or stats['count'] < 1:
            continue
        
        avg = stats['avg']
        deviation = (price - avg) / avg * 100 if avg > 0 else 0
        
        if abs(deviation) >= threshold:
            anomalies.append({
                '物料编号': mat_id,
                '物料名称': rec['物料名称'],
                '物料描述': rec['物料描述'] or '',
                '采购订单号': po,
                '下单时间': str(rec['下单时间']),
                '供应商': rec['供应商名称'],
                '采购数量': rec['采购数量'],
                '含税单价': price,
                '价税合计': rec['价税合计'],
                '状态': rec['状态'],
                '历史均价': round(avg, 2),
                '历史中位价': round(stats['median'], 2),
                '历史最低': stats['min'],
                '历史最高': stats['max'],
                '历史笔数': stats['count'],
                '偏离幅度': f"{deviation:+.1f}%",
                '偏离值': round(deviation, 1),
                '严重程度': '🔴高' if abs(deviation) >= 100 else ('🟡中' if abs(deviation) >= 50 else '🟢低'),
                '历史订单编号': stats['po_refs'],
            })
    
    anomalies.sort(key=lambda x: abs(x['偏离值']), reverse=True)
    return anomalies

def export_excel(anomalies, output_path, days):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'近{days}天价格异常'
    
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    columns = [
        ('序号', 6), ('严重程度', 10), ('采购订单号', 22), ('物料编号', 16),
        ('物料名称', 14), ('规格描述', 36), ('本次单价', 12), ('本次数量', 10),
        ('本次金额', 12), ('供应商', 24), ('下单时间', 12),
        ('历史均价', 12), ('历史中位价', 12),
        ('历史最低', 12), ('历史最高', 12), ('历史笔数', 10), ('偏离幅度', 10),
        ('历史订单编号(备注)', 80),
    ]
    
    for col_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    for row_idx, a in enumerate(anomalies, 2):
        row_data = [
            row_idx - 1,
            a['严重程度'],
            a['采购订单号'],
            a['物料编号'],
            a['物料名称'],
            a['物料描述'],
            a['含税单价'],
            a['采购数量'],
            a['价税合计'],
            a['供应商'],
            a['下单时间'],
            a['历史均价'],
            a['历史中位价'],
            a['历史最低'],
            a['历史最高'],
            a['历史笔数'],
            a['偏离幅度'],
            a['历史订单编号'],
        ]
        
        severity = a['严重程度']
        fill = red_fill if '高' in severity else (yellow_fill if '中' in severity else green_fill)
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'
    wb.save(output_path)
    print(f"\n导出完成: {output_path}")
    print(f"共 {len(anomalies)} 条异常")

def main():
    days = 60
    threshold = 50
    
    if len(sys.argv) > 1:
        days = int(sys.argv[1])
    if len(sys.argv) > 2:
        threshold = int(sys.argv[2])
    
    data_file = '采购价格异常报告-20260527.xlsx'
    output_file = f'近{days}天价格异常-{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    print(f"=== 近 {days} 天采购价格异常分析 ===")
    print(f"偏差阈值: ±{threshold}%")
    
    records = load_data(data_file)
    anomalies = analyze(records, days=days, threshold=threshold)
    
    high = sum(1 for a in anomalies if '高' in a['严重程度'])
    mid = sum(1 for a in anomalies if '中' in a['严重程度'])
    low = sum(1 for a in anomalies if '低' in a['严重程度'])
    print(f"\n异常分布: 🔴高 {high}, 🟡中 {mid}, 🟢低 {low}")
    
    print("\n--- 全部异常清单 ---")
    for a in anomalies:
        print(f"  {a['严重程度']} {a['采购订单号']} | {a['物料名称']} {a['物料描述'][:30]} | ¥{a['含税单价']} vs ¥{a['历史均价']} ({a['偏离幅度']})")
    
    export_excel(anomalies, output_file, days)

if __name__ == '__main__':
    main()
