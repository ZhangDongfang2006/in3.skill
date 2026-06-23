#!/usr/bin/env python3
"""
每日采购价格分析（增强版 v2）
- 读取宁波+湖北采购订单明细(当天导出的全量文件)
- 筛选当天下单的采购记录
- 三类比价：
  1. 铜排/电线 → 按当日铜价计算基准价
  2. 标准元器件/其他 → 按历史均价对比
  3. 螺丝/标准件/走线槽 → 跳过
- 全部列出（不筛选百分比），按偏离度从高到低排列
"""

import sys
import os
import re
import json
import urllib.request
import urllib.error
from collections import defaultdict
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_DIR = os.path.dirname(os.path.abspath(__file__))

# ============ 物料分类规则 ============

# 跳过的物料类型（螺丝/标准件/走线槽等）
SKIP_KEYWORDS = [
    '螺丝', '螺钉', '螺柱', '螺栓', '螺帽', '螺母',
    '垫片', '弹簧垫', '平垫', '弹垫',
    '铆钉', '销钉', '卡簧', '挡圈',
    '走线槽', '线槽', '扎带', '尼龙',
    '铰链', '门锁', '锁具', '把手', '密封条',
    '标签', '铭牌', '标牌', '贴纸',
    '包装', '木箱', '纸箱', '泡沫',
    '焊锡', '焊条', '焊丝',
    '胶水', '胶带', '双面胶', '热缩管', '冷缩管',
    '编织带', '钢钉',
]

# 铜价基准类物料
COPPER_KEYWORDS = ['铜排', '铜母线', '铜编织', '铜带']
# 电线电缆类
WIRE_KEYWORDS = ['电线', '电缆', 'BVR', 'BV ', 'BVV', 'YJV', 'RV ', 'RVS']

# 钢板类
STEEL_KEYWORDS = ['覆铝锌板', '钢板', '不锈钢板', '冷轧板', '热轧板', '镀锌板']

# 标准元器件类（可网上搜价）
COMPONENT_KEYWORDS = [
    '断路器', '接触器', '继电器', '变频器', '互感器', '浪涌保护器',
    '熔芯', '熔断器', '开关电源', '热继电', '脱扣器', '操作机构',
    '控制器', '电动机', '马达保护', '滤波器', '滤波控制器',
    '无功补偿', '多功能表', '电能表', '电度表', '仪表',
    '转换开关', '隔离开关', '负荷开关', '刀开关',
    '按钮', '指示灯', '蜂鸣器', '电流表', '电压表',
    '接线端子', '端子排', '接线排',
    '插拔件', '插件', '连接器', '接插件',
    '变压器', '稳压器', 'UPS', '电源',
    '传感器', '行程开关', '限位开关', '接近开关', '光电开关',
    '微型断路器', '小型断路器', '塑壳断路器', '万能式',
    '漏电断路器', '框架断路器',
    '后备保护器', '电容器',
    '铜鼻子', '线鼻子', 'OT端子', 'UT端子',
    '二次夹头', '一次夹头',
    '母线框', '母线夹', '绝缘件',
    '热缩', '冷缩', '绝缘罩', '绝缘盖',
]

# 铜密度 (g/cm³)
COPPER_DENSITY = 8.96


def classify_material(name, desc):
    """分类物料：copper(铜价基准) / wire(电线铜价基准) / steel(钢材) / component(元器件) / skip(跳过) / other(其他历史比价)"""
    full = f"{name} {desc}"

    # 先检查是否需要跳过
    for kw in SKIP_KEYWORDS:
        if kw in name:
            return 'skip', kw

    # 铜排
    for kw in COPPER_KEYWORDS:
        if kw in name:
            return 'copper', kw

    # 电线电缆
    for kw in WIRE_KEYWORDS:
        if kw in name or kw in desc:
            return 'wire', kw

    # 钢板
    for kw in STEEL_KEYWORDS:
        if kw in name:
            return 'steel', kw

    # 标准元器件
    for kw in COMPONENT_KEYWORDS:
        if kw in name:
            return 'component', kw

    return 'other', ''


# ============ 铜价获取 ============

def get_copper_price():
    """
    从有色宝长江(changjiangyouse.com)获取当日长江现货铜价
    该网站服务端渲染，urllib 可直接抓取
    返回: (铜价元/吨, 日期字符串 YYYY-MM-DD) 或 (None, None)
    """
    try:
        url = "https://www.changjiangyouse.com"
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        })
        with urllib.request.urlopen(req, timeout=10) as resp:
            html = resp.read().decode('utf-8', errors='ignore')

        # 页面格式：铜 105240-105260 105250 -380 ↓ ... 06-18
        # 提取铜行数据：名称(铜) 价格区间 均价 涨跌 ... 日期
        # 先找到「铜」所在的行
        # 去掉HTML标签后搜索
        clean = re.sub(r'<[^>]+>', ' ', html)
        clean = re.sub(r'\s+', ' ', clean).strip()

        # 匹配：铜 价格低-价格高 均价 涨跌 前三日均价 前五日均价 日期
        m = re.search(r'铜\s+(\d{5,7})\s*[-–]\s*(\d{5,7})\s+(\d{5,7})\s+[-+]?\d+\s*[↑↓]?\s+\d+\.?\d*\s+\d+\.?\d*\s+(\d{2})-(\d{2})', clean)
        if m:
            low, high, avg, month, day = m.groups()
            # 推断年份（当前年）
            year = datetime.now().year
            price_date = f"{year}-{month}-{day}"
            return int(avg), price_date

        # 降级：找任何 1XXXXX 格式的5-6位数字（铜价范围 50000-120000）
        m2 = re.search(r'铜\s+(\d{5,7})\s*[-–]\s*(\d{5,7})\s+(\d{5,7})', clean)
        if m2:
            low, high, avg = m2.groups()
            # 找最近的日期
            date_m = re.search(r'(\d{2})-(\d{2})', clean[clean.find('铜'):clean.find('铜')+200])
            year = datetime.now().year
            if date_m:
                price_date = f"{year}-{date_m.group(1)}-{date_m.group(2)}"
            else:
                price_date = date.today().strftime('%Y-%m-%d')
            return int(avg), price_date

        print(f"  ⚠️ 页面中未找到铜价数据")
        return None, None
    except Exception as e:
        print(f"  ⚠️ 获取铜价失败: {e}")
        return None, None


# ============ 铜排/电线基准价计算 ============

def parse_busbar_dimensions(desc):
    """从描述中解析铜排尺寸，返回截面积 mm²。如 '80×6' → 480"""
    # 匹配 数字×数字 或 数字*数字 或 数字x数字
    m = re.search(r'(\d+)\s*[×xX\*]\s*(\d+)', str(desc))
    if m:
        w, h = int(m.group(1)), int(m.group(2))
        if 2 <= w <= 200 and 2 <= h <= 30:
            return w * h
    return None


def parse_wire_cross_section(name, desc):
    """从电线名称/描述中解析截面积 mm²。如 'BVR 4' → 4"""
    full = f"{name} {desc}"
    # 尝试匹配 "BVR 2.5"、"BV 4"、"YJV 3×16" 等
    # 先尝试 数字² 或 数字2 的写法
    m = re.search(r'(\d+(?:\.\d+)?)\s*[²2]', full)
    if m:
        return float(m.group(1))

    # 匹配 BVR/BV/RV + 数字
    m = re.search(r'(?:BVR|BV|RV|RVS|BVV|YJV)\s*(\d+(?:\.\d+)?)', full, re.IGNORECASE)
    if m:
        return float(m.group(1))

    # 匹配 "X平方" 或 "X平"
    m = re.search(r'(\d+(?:\.\d+)?)\s*(?:平方|平)', full)
    if m:
        return float(m.group(1))

    # 匹配 "数字×数字×数字" 格式（如3×16，取16）
    m = re.search(r'\d+\s*[×xX\*]\s*(\d+(?:\.\d+)?)', full)
    if m:
        val = float(m.group(1))
        if val <= 400:  # 合理的电线截面范围
            return val

    return None


# T2铜排加工费（元/kg），来源：SMM上海有色网
# 宽度≤160mm: 约 2.4 元/kg
# 宽度160-300mm: 约 4.0 元/kg
COPPER_BUSBAR_PROCESSING_FEE = 2.4  # 元/kg
# 电线加工费（含绝缘材料），约 3-5 元/kg，取3.5
WIRE_PROCESSING_FEE = 3.5  # 元/kg


def calc_copper_benchmark(category, desc, name, copper_price_per_kg):
    """
    计算铜排/电线的铜价基准（元/米）
    公式：重量(kg/m) × (铜价 + 加工费)
    copper_price_per_kg: 当日铜价 元/kg
    返回: (基准价元/米, 计算说明) 或 (None, None)
    """
    if category == 'copper':
        area = parse_busbar_dimensions(desc)
        if area and copper_price_per_kg:
            weight_kg = area * COPPER_DENSITY / 1000  # kg/m
            # 铜排基准 = 重量 × (铜价 + 加工费)
            fair_price_per_kg = copper_price_per_kg + COPPER_BUSBAR_PROCESSING_FEE
            benchmark = weight_kg * fair_price_per_kg
            note = f"截面{area}mm², 重{weight_kg:.3f}kg/m, 铜价{copper_price_per_kg:.1f}+加工费{COPPER_BUSBAR_PROCESSING_FEE}={fair_price_per_kg:.1f}元/kg"
            return round(benchmark, 2), note
    elif category == 'wire':
        area = parse_wire_cross_section(name, desc)
        if area and copper_price_per_kg:
            weight_kg = area * COPPER_DENSITY / 1000  # kg/m
            # 电线基准 = 重量 × (铜价 + 加工费)
            fair_price_per_kg = copper_price_per_kg + WIRE_PROCESSING_FEE
            benchmark = weight_kg * fair_price_per_kg
            note = f"截面{area}mm², 重{weight_kg:.4f}kg/m, 铜价{copper_price_per_kg:.1f}+加工费{WIRE_PROCESSING_FEE}={fair_price_per_kg:.1f}元/kg"
            return round(benchmark, 2), note

    return None, None


# ============ IN3 数据读取 ============

def find_columns(headers):
    """根据列名动态建立映射"""
    col_map = {}
    for i, h in enumerate(headers):
        h_str = str(h).strip() if h else ''
        if h_str == '采购订单号': col_map['po'] = i
        elif h_str == '合同编号': col_map['contract_no'] = i
        elif h_str == '合同名称': col_map['contract_name'] = i
        elif h_str == '销售订单编号': col_map['sales_no'] = i
        elif h_str == '销售订单名称': col_map['sales_name'] = i
        elif h_str == '供应商名称': col_map['supplier'] = i
        elif h_str == '采购申请人': col_map['buyer'] = i
        elif h_str == '物料编号': col_map['material_id'] = i
        elif h_str == '物料名称': col_map['material_name'] = i
        elif h_str == '物料描述': col_map['material_desc'] = i
        elif h_str == '物料类别': col_map['material_cat'] = i
        elif h_str == '物料子类别': col_map['material_subcat'] = i
        elif h_str == '采购数量': col_map['qty'] = i
        elif h_str == '采购单位': col_map['unit'] = i
        elif h_str == '含税单价': col_map['price_with_tax'] = i
        elif h_str == '未税单价': col_map['price_no_tax'] = i
        elif h_str == '价税合计': col_map['amount'] = i
        elif h_str in ('下单时间', '下单日期'): col_map['order_date'] = i
        elif h_str == '创建时间': col_map['create_time'] = i
        elif h_str == '制造商': col_map['maker'] = i
        elif h_str == '制造商物料编码': col_map['maker_model'] = i
    return col_map


def load_all_records(nb_file, hb_file):
    """读取宁波+湖北采购订单明细,返回全部记录"""
    all_records = []

    for filepath, factory in [(nb_file, '宁波'), (hb_file, '湖北')]:
        if not filepath or not os.path.exists(filepath):
            print(f"  ⚠️ 文件不存在: {filepath}")
            continue

        wb = openpyxl.load_workbook(filepath)  # 不用 read_only！IN3 xlsx dimension bug
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        col_map = find_columns(headers)

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)

            def get(key):
                idx = col_map.get(key)
                if idx is None or idx >= len(vals):
                    return None
                return vals[idx]

            po = str(get('po') or '')
            if not po.startswith('PO'):
                continue

            material_id = str(get('material_id') or '')
            if not material_id:
                continue

            price = get('price_with_tax')
            try:
                price = float(price) if price else 0
            except (ValueError, TypeError):
                price = 0

            price_no_tax = get('price_no_tax')
            try:
                price_no_tax = float(price_no_tax) if price_no_tax else 0
            except (ValueError, TypeError):
                price_no_tax = 0

            # 日期解析：优先创建时间（更精确），否则下单时间
            create_time = get('create_time')
            order_date = get('order_date')

            if create_time and hasattr(create_time, 'strftime'):
                date_str = create_time.strftime('%Y-%m-%d')
            elif order_date and hasattr(order_date, 'strftime'):
                date_str = order_date.strftime('%Y-%m-%d')
            elif create_time:
                date_str = str(create_time)[:10]
            elif order_date:
                date_str = str(order_date)[:10]
            else:
                date_str = ''

            qty = get('qty')
            try:
                qty = float(qty) if qty else 0
            except (ValueError, TypeError):
                qty = 0

            amount = get('amount')
            try:
                amount = float(amount) if amount else 0
            except (ValueError, TypeError):
                amount = 0

            record = {
                'factory': factory,
                'po': po,
                'material_id': material_id,
                'material_name': str(get('material_name') or ''),
                'material_desc': str(get('material_desc') or ''),
                'material_cat': str(get('material_cat') or ''),
                'material_subcat': str(get('material_subcat') or ''),
                'supplier': str(get('supplier') or ''),
                'buyer': str(get('buyer') or ''),
                'price': price,  # 含税单价
                'price_no_tax': price_no_tax,
                'qty': qty,
                'amount': amount,
                'unit': str(get('unit') or ''),
                'order_date': date_str,
                'contract_no': str(get('contract_no') or ''),
                'contract_name': str(get('contract_name') or ''),
                'sales_no': str(get('sales_no') or ''),
                'sales_name': str(get('sales_name') or ''),
                'maker': str(get('maker') or ''),
                'maker_model': str(get('maker_model') or ''),
            }
            all_records.append(record)
            count += 1

        wb.close()
        print(f"  {factory}: {count} 条记录")

    return all_records


# ============ 分析逻辑 ============

def analyze_daily(all_records, target_date, copper_price_per_kg=None):
    """
    筛选 target_date 的采购，分三类分析
    返回: (copper_results, hist_results, skipped_count)
    """
    target_str = target_date.strftime('%Y-%m-%d')

    recent = []
    historical = defaultdict(list)

    for rec in all_records:
        if rec['order_date'] == target_str:
            recent.append(rec)
        else:
            if rec['price'] > 0:
                historical[rec['material_id']].append(rec)

    print(f"\n当天采购 ({target_str}): {len(recent)} 条")
    print(f"历史采购: {sum(len(v) for v in historical.values())} 条, {len(historical)} 种物料")

    # 历史均价统计
    hist_stats = {}
    for mat_id, recs in historical.items():
        prices = [r['price'] for r in recs if r['price'] > 0]
        if not prices:
            continue
        sorted_prices = sorted(prices)
        hist_stats[mat_id] = {
            'avg': sum(prices) / len(prices),
            'median': sorted_prices[len(sorted_prices) // 2],
            'min': min(prices),
            'max': max(prices),
            'count': len(prices),
            'po_refs': '; '.join(
                f"{r['po']}(¥{r['price']:.2f},{r['order_date']})"
                for r in recs[:10]
            ),
        }

    copper_results = []  # 铜价基准分析结果
    hist_results = []    # 历史均价分析结果
    skipped_count = 0
    new_materials = []   # 新物料（无历史价格）
    seen = set()

    for rec in recent:
        mat_id = rec['material_id']
        po = rec['po']
        price = rec['price']

        if price <= 0:
            continue

        key = (po, mat_id)
        if key in seen:
            continue
        seen.add(key)

        name = rec['material_name']
        desc = rec['material_desc']
        category, matched_kw = classify_material(name, desc)

        if category == 'skip':
            skipped_count += 1
            continue

        # 铜排/电线 → 铜价基准
        if category in ('copper', 'wire') and copper_price_per_kg:
            benchmark, note = calc_copper_benchmark(category, desc, name, copper_price_per_kg)
            if benchmark:
                deviation = (price - benchmark) / benchmark * 100 if benchmark > 0 else 0
                copper_results.append({
                    '类型': '铜排' if category == 'copper' else '电线',
                    '采购订单号': po,
                    '物料编号': mat_id,
                    '物料名称': name,
                    '规格描述': desc,
                    '本次单价': round(price, 4),
                    '本次数量': rec['qty'],
                    '本次金额': round(rec['amount'], 2),
                    '采购单位': rec['unit'],
                    '供应商': rec['supplier'],
                    '工厂': rec['factory'],
                    '采购申请人': rec['buyer'],
                    '铜价基准': benchmark,
                    '计算说明': note,
                    '偏离幅度': round(deviation, 1),
                    '比价方式': '当日铜价基准',
                    'order_date': rec['order_date'],
                })
                continue

        # 钢板/元器件/其他 → 历史均价对比
        stats = hist_stats.get(mat_id)
        if stats and stats['count'] >= 1:
            avg = stats['avg']
            deviation = (price - avg) / avg * 100 if avg > 0 else 0

            # 项目信息
            contract_name = rec.get('contract_name', '')
            contract_no = rec.get('contract_no', '')
            sales_name = rec.get('sales_name', '')
            parts = []
            if contract_name:
                parts.append(contract_name)
            if contract_no:
                parts.append(f"合同:{contract_no}")
            if sales_name and sales_name != contract_name:
                parts.append(f"销售订单:{sales_name}")
            project_info = ' | '.join(parts) if parts else ''

            hist_results.append({
                '类型': {'steel': '钢板', 'component': '元器件', 'other': '其他'}.get(category, '其他'),
                '采购订单号': po,
                '物料编号': mat_id,
                '物料名称': name,
                '规格描述': desc,
                '本次单价': round(price, 4),
                '本次数量': rec['qty'],
                '本次金额': round(rec['amount'], 2),
                '采购单位': rec['unit'],
                '供应商': rec['supplier'],
                '工厂': rec['factory'],
                '采购申请人': rec['buyer'],
                '历史均价': round(avg, 4),
                '历史中位价': round(stats['median'], 4),
                '历史最低': round(stats['min'], 4),
                '历史最高': round(stats['max'], 4),
                '历史笔数': stats['count'],
                '偏离幅度': round(deviation, 1),
                '历史订单': stats['po_refs'],
                '比价方式': '历史均价',
                '关联项目': project_info,
                'order_date': rec['order_date'],
            })
        else:
            # 新物料，无历史价格
            new_materials.append({
                '类型': {'steel': '钢板', 'component': '元器件', 'copper': '铜排', 'wire': '电线', 'other': '其他'}.get(category, '其他'),
                '采购订单号': po,
                '物料编号': mat_id,
                '物料名称': name,
                '规格描述': desc,
                '本次单价': round(price, 4),
                '本次数量': rec['qty'],
                '本次金额': round(rec['amount'], 2),
                '采购单位': rec['unit'],
                '供应商': rec['supplier'],
                '工厂': rec['factory'],
                '采购申请人': rec['buyer'],
                '销售订单号': rec.get('sales_no', ''),
                '销售订单名称': rec.get('sales_name', ''),
                '合同编号': rec.get('contract_no', ''),
                '合同名称': rec.get('contract_name', ''),
                '比价方式': '新物料（无历史）',
                'order_date': rec['order_date'],
            })

    # 排序：偏离幅度从高到低
    copper_results.sort(key=lambda x: x['偏离幅度'], reverse=True)
    hist_results.sort(key=lambda x: x['偏离幅度'], reverse=True)

    return copper_results, hist_results, new_materials, skipped_count


# ============ Excel 导出 ============

def export_excel(copper_results, hist_results, new_materials, output_path, target_date, copper_price, copper_date):
    """导出Excel — 全量数据，3个Sheet"""
    wb = openpyxl.Workbook()

    # 样式
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_fill_green = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    hdr_fill_orange = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    hdr_fill_gray = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    def style_header(ws, columns, fill):
        for col_idx, (col_name, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = hdr_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(columns))}1"

    def write_data(ws, columns, data):
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, (_, key) in enumerate(columns, 1):
                val = row_data.get(key, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                # 偏离幅度着色
                if key == '偏离幅度' and isinstance(val, (int, float)):
                    if val >= 20:
                        cell.fill = red_fill
                    elif val <= -20:
                        cell.fill = green_fill
                    elif abs(val) >= 10:
                        cell.fill = yellow_fill

    # ====== Sheet 1: 铜价基准比价 ======
    ws1 = wb.active
    ws1.title = f"铜价基准比价"

    # 铜价信息行
    ws1.cell(row=1, column=1, value=f"长江现货铜价: {copper_price}元/吨 ({copper_date})" if copper_price else "铜价: 未获取")
    ws1.cell(row=1, column=1).font = Font(bold=True, size=12, color="C55A11")
    ws1.cell(row=2, column=1, value=f"计算公式: 截面积(mm²) × 铜密度(8.96) / 1000 × 铜价(元/kg) × 加工费系数(铜排1.10, 电线1.15)")
    ws1.cell(row=2, column=1).font = Font(size=9, italic=True)

    start_row = 4
    copper_cols = [
        ('序号', 'seq'), ('类型', '类型'), ('采购订单号', '采购订单号'),
        ('物料名称', '物料名称'), ('规格描述', '规格描述'),
        ('本次单价(含税)', '本次单价'), ('采购单位', '采购单位'),
        ('本次数量', '本次数量'), ('本次金额', '本次金额'),
        ('铜价基准(元/单位)', '铜价基准'), ('偏离幅度%', '偏离幅度'),
        ('供应商', '供应商'), ('工厂', '工厂'), ('采购申请人', '采购申请人'),
        ('计算说明', '计算说明'),
    ]
    for col_idx, (col_name, _) in enumerate(copper_cols, 1):
        cell = ws1.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill_orange
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = [5, 8, 22, 25, 35, 14, 10, 10, 12, 14, 12, 25, 8, 10, 50][col_idx-1]

    for i, r in enumerate(copper_results, 1):
        r['seq'] = i
        row_idx = start_row + i
        for col_idx, (_, key) in enumerate(copper_cols, 1):
            val = r.get(key, '')
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if key == '偏离幅度%' and isinstance(val, (int, float)):
                if val >= 20:
                    cell.fill = red_fill
                elif val <= -20:
                    cell.fill = green_fill
                elif abs(val) >= 10:
                    cell.fill = yellow_fill

    ws1.freeze_panes = f'A{start_row+1}'
    if copper_results:
        ws1.auto_filter.ref = f"A{start_row}:{openpyxl.utils.get_column_letter(len(copper_cols))}{start_row+len(copper_results)}"

    # ====== Sheet 2: 历史均价比价 ======
    ws2 = wb.create_sheet(f"历史均价比价")
    hist_cols = [
        ('序号', 'seq'), ('类型', '类型'), ('采购订单号', '采购订单号'),
        ('物料名称', '物料名称'), ('规格描述', '规格描述'),
        ('本次单价', '本次单价'), ('采购单位', '采购单位'),
        ('本次数量', '本次数量'), ('本次金额', '本次金额'),
        ('历史均价', '历史均价'), ('历史中位价', '历史中位价'),
        ('历史最低', '历史最低'), ('历史最高', '历史最高'),
        ('历史笔数', '历史笔数'), ('偏离幅度%', '偏离幅度'),
        ('供应商', '供应商'), ('工厂', '工厂'), ('采购申请人', '采购申请人'),
        ('历史订单', '历史订单'), ('关联项目', '关联项目'),
    ]
    for col_idx, (col_name, _) in enumerate(hist_cols, 1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill_blue
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = [5, 8, 22, 25, 35, 12, 10, 10, 12, 12, 12, 12, 12, 8, 12, 25, 8, 10, 55, 45][col_idx-1]

    for i, r in enumerate(hist_results, 1):
        r['seq'] = i
        for col_idx, (_, key) in enumerate(hist_cols, 1):
            val = r.get(key, '')
            cell = ws2.cell(row=i+1, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if key == '偏离幅度%' and isinstance(val, (int, float)):
                if val >= 40:
                    cell.fill = red_fill
                elif val <= -40:
                    cell.fill = green_fill
                elif abs(val) >= 20:
                    cell.fill = yellow_fill

    ws2.freeze_panes = 'A2'
    if hist_results:
        ws2.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(hist_cols))}{len(hist_results)+1}"

    # ====== Sheet 3: 新物料（无历史价格） ======
    ws3 = wb.create_sheet(f"新物料(无历史)")
    new_cols = [
        ('序号', 'seq'), ('类型', '类型'), ('物料编号', '物料编号'),
        ('采购订单号', '采购订单号'),
        ('物料名称', '物料名称'), ('规格描述', '规格描述'),
        ('本次单价', '本次单价'), ('采购单位', '采购单位'),
        ('本次数量', '本次数量'), ('本次金额', '本次金额'),
        ('供应商', '供应商'), ('采购申请人', '采购申请人'),
        ('销售订单号', '销售订单号'), ('销售订单名称', '销售订单名称'),
        ('合同编号', '合同编号'), ('合同名称', '合同名称'),
    ]
    new_widths = [5, 8, 15, 22, 18, 35, 12, 8, 10, 12, 25, 10, 18, 30, 15, 25]
    for col_idx, (col_name, _) in enumerate(new_cols, 1):
        cell = ws3.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill_gray
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = new_widths[col_idx-1]

    for i, r in enumerate(new_materials, 1):
        r['seq'] = i
        for col_idx, (_, key) in enumerate(new_cols, 1):
            val = r.get(key, '')
            cell = ws3.cell(row=i+1, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    ws3.freeze_panes = 'A2'

    # ====== 保存 ======
    wb.save(output_path)
    total = len(copper_results) + len(hist_results) + len(new_materials)
    print(f"✅ 已导出: {output_path}")
    print(f"   铜价基准: {len(copper_results)} 条 | 历史均价: {len(hist_results)} 条 | 新物料: {len(new_materials)} 条 | 跳过: 跳过")


# ============ 文件查找 ============

def find_latest_files():
    """找最新的宁波/湖北采购订单明细文件"""
    nb_file = None
    hb_file = None
    nb_date = ''
    hb_date = ''

    for f in os.listdir(DATA_DIR):
        m = re.match(r'采购订单明细-宁波-(\d{8})\.xlsx$', f)
        if m:
            d = m.group(1)
            if d > nb_date:
                nb_date = d
                nb_file = os.path.join(DATA_DIR, f)
        m = re.match(r'采购订单明细-湖北-(\d{8})\.xlsx$', f)
        if m:
            d = m.group(1)
            if d > hb_date:
                hb_date = d
                hb_file = os.path.join(DATA_DIR, f)

    return nb_file, hb_file


# ============ 主程序 ============

def main():
    target_date = date.today()

    # 命令行参数: [YYYY-MM-DD]
    if len(sys.argv) > 1:
        try:
            target_date = datetime.strptime(sys.argv[1], '%Y-%m-%d').date()
        except ValueError:
            pass

    print(f"=== 每日采购价格分析（增强版 v2）===")
    print(f"检查日期: {target_date}")

    # 获取当日铜价
    print("\n获取长江现货铜价...")
    copper_price_ton, copper_date = get_copper_price()
    copper_price_kg = copper_price_ton / 1000 if copper_price_ton else None
    if copper_price_kg:
        print(f"  铜价: {copper_price_ton} 元/吨 ({copper_price_kg:.2f} 元/kg) [{copper_date}]")
    else:
        print(f"  ⚠️ 未能获取铜价，铜排/电线将改用历史均价对比")

    # 找最新文件
    nb_file, hb_file = find_latest_files()
    print(f"\n数据文件:")
    print(f"  宁波: {os.path.basename(nb_file) if nb_file else '❌ 未找到'}")
    print(f"  湖北: {os.path.basename(hb_file) if hb_file else '❌ 未找到'}")

    if not nb_file and not hb_file:
        print("\n❌ 未找到任何采购订单明细文件")
        sys.exit(1)

    # 加载数据
    print("\n加载数据...")
    all_records = load_all_records(nb_file or '', hb_file or '')

    if not all_records:
        print("❌ 无数据")
        sys.exit(1)

    # 分析
    copper_results, hist_results, new_materials, skipped = analyze_daily(
        all_records, target_date, copper_price_kg
    )

    # 如果铜价获取失败，铜排/电线会落入 new_materials 或 hist_results
    # 没有铜价时把铜排/电线也加入历史比价
    if not copper_price_kg:
        # 铜排/电线没有历史价格的情况下会进入 new_materials，这是合理的
        pass

    # 汇总
    print(f"\n=== 分析结果 ===")
    print(f"铜价基准比价: {len(copper_results)} 条")
    print(f"历史均价比价: {len(hist_results)} 条")
    print(f"新物料(无历史): {len(new_materials)} 条")
    print(f"跳过(螺丝/标准件等): {skipped} 条")

    total_analyzed = len(copper_results) + len(hist_results) + len(new_materials)
    if total_analyzed == 0:
        print(f"\nℹ️ {target_date} 无需分析的采购记录")
        return

    # 打印摘要
    if copper_results:
        print(f"\n--- 铜价基准 TOP5（偏离最高）---")
        for r in copper_results[:5]:
            print(f"  {r['物料名称']} | {r['规格描述'][:20]} | 采购¥{r['本次单价']} vs 基准¥{r['铜价基准']} ({r['偏离幅度']:+.1f}%) | {r['工厂']}")

    if hist_results:
        print(f"\n--- 历史均价 TOP5（偏离最高）---")
        for r in hist_results[:5]:
            print(f"  {r['物料名称']} | {r['规格描述'][:20]} | 采购¥{r['本次单价']} vs 均价¥{r['历史均价']} ({r['偏离幅度']:+.1f}%) | {r['工厂']}")

    if new_materials:
        print(f"\n--- 新物料 {len(new_materials)} 条 ---")
        for r in new_materials[:10]:
            print(f"  {r['物料名称']} | {r['规格描述'][:20]} | ¥{r['本次单价']} | {r['工厂']}")

    # 导出 Excel
    output = os.path.join(DATA_DIR, f'采购价格分析-{target_date.strftime("%Y%m%d")}.xlsx')
    export_excel(copper_results, hist_results, new_materials, output, target_date, copper_price_ton, copper_date)

    # 输出 JSON 摘要供 cron 使用
    summary = {
        'date': str(target_date),
        'copper_price': copper_price_ton,
        'copper_date': copper_date,
        'copper_count': len(copper_results),
        'hist_count': len(hist_results),
        'new_count': len(new_materials),
        'skipped': skipped,
        'output': output,
        # 标记需要关注的
        'copper_high': [r for r in copper_results if r['偏离幅度'] >= 15],
        'hist_high': [r for r in hist_results if r['偏离幅度'] >= 40],
    }
    summary_path = os.path.join(DATA_DIR, f'price_analysis_summary_{target_date.strftime("%Y%m%d")}.json')
    with open(summary_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2, default=str)

    print(f"\n[SUMMARY] {target_date}: 铜{len(copper_results)}条 历史{len(hist_results)}条 新{len(new_materials)}条 跳过{skipped}")
    print(f"[OUTPUT] {output}")


if __name__ == '__main__':
    main()
