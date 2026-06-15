#!/usr/bin/env python3
"""过滤AI检测的待人工确认配对，去掉明确不相关的，只留可疑的"""
import openpyxl
import re

INPUT_FILE = 'AI检测重复物料-20260526-v9.xlsx'

wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb['待人工确认']

pairs = []
current = None
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    if row[0] == '序号':
        continue
    seq = row[0]
    mark = row[1]
    code = row[2]
    name = str(row[3] or '')
    desc = str(row[4] or '')
    cat = str(row[6] or '')
    subcat = str(row[7] or '')
    remark = str(row[15] or '')

    if mark and str(mark).startswith('A-'):
        current = {
            'seq': seq, 'a_code': code, 'a_name': name,
            'a_desc': desc, 'a_cat': cat, 'a_subcat': subcat, 'remark': remark
        }
    elif mark and str(mark).startswith('B-') and current:
        current['b_code'] = code
        current['b_name'] = name
        current['b_desc'] = desc
        current['b_cat'] = cat
        current['b_subcat'] = subcat
        pairs.append(current)
        current = None

def is_obviously_different(p):
    """返回 (bool, reason) — True = 明确不是重复"""
    a = p['a_name']
    b = p['b_name']
    a_desc = p['a_desc']
    b_desc = p['b_desc']

    # 1. 完全不同的东西（常识判断）
    obviously_diff = [
        ('手轮', '红砂纸'), ('手轮', '橡皮圈'), ('红砂纸', '橡皮圈'),
        ('风顶', '车轮'), ('号码管', '航空针'), ('航空针', '透明接地线'),
        ('铜宝塔接头', 'U型管夹'), ('铜宝塔接头', '金属石墨垫片'),
        ('插销拉手锁', '定位吸盘'), ('方形拉手', '管装把手A型'),
        ('4代柜B型拉手', '低压电工绝缘胶带'),
        ('导向滑槽', '圆钢'), ('圆钢', '缠绕管'),
        ('变压器低压侧接头可铅封绝缘护套', '电表箱'),
        ('底框', '非标箱'),
        ('驱动大头', '驱动链条'),
        ('油漆刷', '毛刷'), ('铲刀', '毛刷'),
        ('5#引风圈', '扁铁'),
        ('低压验电笔', '低压接地线'),
        ('波纹管', '橡皮圈'),
        ('无缝钢管', '软连接'),
    ]
    for x, y in obviously_diff:
        if (x in a and y in b) or (y in a and x in b):
            return True, f'完全不同的东西: {x} ≠ {y}'

    # 2. 不同类型的漆
    if '漆' in a and '漆' in b and a != b:
        qi_types = ['环氧云铁中间漆', '环氧富锌底漆', '高附着力底漆', '磷化底漆', '丙烯酸聚氨酯']
        a_type = next((t for t in qi_types if t in a), None)
        b_type = next((t for t in qi_types if t in b), None)
        if a_type and b_type and a_type != b_type:
            return True, f'不同类型的漆: {a_type} ≠ {b_type}'

    # 3. 稀释剂 vs 漆 vs 固化剂
    if '稀释剂' in a and '稀释剂' not in b:
        return True, f'稀释剂 ≠ 非稀释剂'
    if '固化剂' in a and '固化剂' not in b:
        return True, f'固化剂 ≠ 非固化剂'
    if '固化剂' in b and '固化剂' not in a:
        return True, f'固化剂 ≠ 非固化剂'
    if '稀释剂' in b and '稀释剂' not in a:
        return True, f'稀释剂 ≠ 非稀释剂'

    # 4. 螺母类型不同
    if '螺母' in a and '螺母' in b:
        lomu_types = ['方形螺母', '六角焊接螺母', '法兰螺母', '法兰防松螺母', '蓝白锌法兰螺母', '镀彩锌法兰防松螺母']
        a_type = next((t for t in lomu_types if t in a), None)
        b_type = next((t for t in lomu_types if t in b), None)
        if a_type and b_type and a_type != b_type:
            return True, f'不同类型螺母: {a_type} ≠ {b_type}'

    # 5. 组合螺丝 vs 带垫螺丝
    if a == '组合' and b == '带垫':
        return True, '组合螺丝 ≠ 带垫螺丝（不同紧固件类型）'

    # 6. 插件型号不同
    if '插件' in a and '插件' in b:
        if 'XCT9' in a and 'XCT6' in b:
            return True, 'XCT9系列 ≠ XCT6系列（不同型号系列）'
        if 'XCZ9' in a and 'XCZ6' in b:
            return True, 'XCZ9系列 ≠ XCZ6系列（不同型号系列）'
        if 'RC4' in a and 'RC8' in b:
            return True, 'RC4 ≠ RC8（不同型号）'

    # 7. 玻璃 vs 小母线端子
    if a == '玻璃' and b == '小母线端子':
        return True, '玻璃 ≠ 小母线端子（完全不同的东西）'

    # 8. 拉手型号不同
    if '拉手' in a and '拉手' in b:
        if '1/2新型拉手' in a and '4代柜A型拉手' in b:
            return True, '1/2新型拉手 ≠ 4代柜A型拉手（不同型号拉手）'

    # 9. 软管 vs 波纹管（不同材料/类型）
    if '软管' in a and '波纹管' in b:
        return True, '耐高温软管 ≠ 波纹管（不同管类）'
    if '波纹管' in a and '软管' in b:
        return True, '波纹管 ≠ 耐高温软管（不同管类）'

    # 10. 卤素灯珠 vs 烤箱灯
    if '卤素灯珠' in a and '烤箱灯' in b:
        return True, '卤素灯珠 ≠ 烤箱灯（不同灯具）'
    if '烤箱灯' in a and '卤素灯珠' in b:
        return True, '烤箱灯 ≠ 卤素灯珠（不同灯具）'

    # 11. 信号线束 vs LED灯到主板（可能其实重复，需要看）
    # 先标记为可疑

    return False, ''


# 分类
sure_not_dup = []
suspicious = []

for p in pairs:
    is_diff, reason = is_obviously_different(p)
    if is_diff:
        p['reason'] = reason
        sure_not_dup.append(p)
    else:
        suspicious.append(p)

print(f'总计: {len(pairs)} 对')
print(f'明确非重复（排除）: {len(sure_not_dup)} 对')
print(f'可疑（保留给人工确认）: {len(suspicious)} 对')
print()
print('=== 🔍 可疑配对（需要人工确认）===')
for p in suspicious:
    print(f"  {p['seq']}. [{p['a_code']}] {p['a_name']} ({p['a_desc'][:50]})")
    print(f"       [{p['b_code']}] {p['b_name']} ({p['b_desc'][:50]})")
print()
print('=== ✅ 明确非重复（已排除）===')
for p in sure_not_dup:
    print(f"  {p['seq']}. {p['a_name']} vs {p['b_name']} → {p['reason']}")
