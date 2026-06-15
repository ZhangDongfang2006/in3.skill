#!/usr/bin/env python3
import json, re, sys
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

print("Loading data...", flush=True)

with open('confirmed_dup_pairs_v2.json', 'r') as f:
    all_pairs = json.load(f)
with open('not_dup_pairs_v2.json', 'r') as f:
    not_dup_pairs = json.load(f)

confirmed_dup = []
suspected_dup = []
needs_review = []
false_positive = []

for p in all_pairs:
    a, b = p['a'], p['b']
    reason = p['reason']
    if '名称描述制造商均相同' in reason:
        confirmed_dup.append(p)
    elif '名称描述相同制造商不同' in reason:
        suspected_dup.append({**p, 'reason': '名称描述相同，制造商不同'})
    elif '描述高度相似' in reason:
        ratio = SequenceMatcher(None, a['desc'], b['desc']).ratio()
        if ratio >= 0.95:
            suspected_dup.append({**p, 'reason': '描述相似度' + str(round(ratio, 2))})
        elif ratio >= 0.92:
            needs_review.append({**p, 'reason': '描述相似度' + str(round(ratio, 2))})
        else:
            false_positive.append({**p, 'reason': '描述相似度' + str(round(ratio, 2))})

print("confirmed_dup: %d" % len(confirmed_dup), flush=True)
print("suspected_dup: %d" % len(suspected_dup), flush=True)
print("needs_review: %d" % len(needs_review), flush=True)
print("false_positive: %d" % len(false_positive), flush=True)
print("not_dup: %d" % len(not_dup_pairs), flush=True)

OUTPUT = '可疑重复物料_已验证_v2.xlsx'
wb = Workbook()

BLUE = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
ORANGE = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
RED = Font(color='FF0000')
BLACK = Font(color='000000')
HFONT = Font(color='000000', bold=True)
HFILL = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

NO_RED = {'物料编号'}
FIELDS = [('物料编号','id'),('物料名称','name'),('物料描述','desc'),
          ('物料类别','category'),('物料子类别','subcategory'),
          ('制造商','manufacturer'),('物料来源','source'),('主计量单位','unit')]

def write_sheet(wb, title, pairs):
    ws = wb.create_sheet(title[:31])
    ws.append([title + '（' + str(len(pairs)) + ' 对）', '', ''])
    ws.merge_cells('A1:C1')
    
    hdrs = []
    for prefix in ['A', 'B']:
        for label, _ in FIELDS:
            hdrs.append(prefix + '-' + label)
    hdrs.append('差异/备注')
    
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HFONT; c.fill = HFILL; c.border = BORDER
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    
    for i, p in enumerate(pairs):
        row = 3 + i
        fill = BLUE if i % 2 == 0 else ORANGE
        av = [p['a'].get(k, '') for _, k in FIELDS]
        bv = [p['b'].get(k, '') for _, k in FIELDS]
        
        for col, val in enumerate(av + bv + [p['reason']], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill; c.border = BORDER
            c.alignment = Alignment(wrap_text=True)
            
            fi = (col - 1) % len(FIELDS)
            label = FIELDS[fi][0]
            is_a = col <= len(FIELDS)
            other = bv[fi] if is_a else av[fi]
            self_v = av[fi] if is_a else bv[fi]
            
            if label not in NO_RED and self_v != other and self_v and other:
                c.font = RED
            else:
                c.font = BLACK
    
    for col in range(1, len(hdrs) + 1):
        label = FIELDS[(col-1) % len(FIELDS)][0]
        if label == '物料描述': w = 50
        elif label == '物料编号': w = 16
        elif label in ('物料名称','制造商','差异/备注'): w = 25
        else: w = 15
        ws.column_dimensions[get_column_letter(col)].width = w

# Overview sheet
ws = wb.active
ws.title = '总览'
ws['A1'] = 'IN3 物料重复检查报告'
ws['A1'].font = Font(bold=True, size=14)
ts = time.strftime('%Y-%m-%d %H:%M')
ws['A3'] = '数据来源'
ws['B3'] = '物料主数据导出 (' + ts + ')'
ws['A4'] = '检查范围'
ws['B4'] = '非成品物料 18,049 条'
ws['A6'] = '分类'
ws['B6'] = '数量'
ws['A7'] = '确认重复（名称+描述+制造商均相同）'
ws['B7'] = len(confirmed_dup)
ws['A8'] = '高度可疑（描述相似度>=0.95）'
ws['B8'] = len(suspected_dup)
ws['A9'] = '待人工确认（描述相似度0.92-0.95）'
ws['B9'] = len(needs_review)
ws['A10'] = '非重复（规则排除）'
ws['B10'] = len(not_dup_pairs)
ws['A11'] = '误报（描述相似度<0.92）'
ws['B11'] = len(false_positive)

write_sheet(wb, '确认重复', confirmed_dup)
write_sheet(wb, '高度可疑', suspected_dup)
write_sheet(wb, '待人工确认', needs_review)
write_sheet(wb, '非重复示例', not_dup_pairs[:200])

wb.save(OUTPUT)
print("Done: " + OUTPUT, flush=True)
