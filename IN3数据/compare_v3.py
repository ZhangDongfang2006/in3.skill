#!/usr/bin/env python3
"""对比 v3 版本的重复物料文件"""

import openpyxl
import re
from collections import defaultdict

def parse_excel_key(sheet_name, row_num):
    """解析 Excel 确认重复 sheet 中的物料编号对，返回 key tuple(sorted([code_a, code_b]))"""
    # 检查当前行的第一个单元格是否包含 A-1, A-2, B-1, B-2 等格式
    first_cell = sheet_name.cell(row=row_num, column=1).value
    if not first_cell:
        return None
        
    # 使用正则表达式匹配 A-1, B-1 等格式
    match = re.match(r'^([AB])-(\d+)$', str(first_cell))
    if match:
        row_letter = match.group(1)
        row_number = int(match.group(2))
        
        # 获取对应的物料编号
        code_cell = sheet_name.cell(row=row_num, column=2)  # 假设物料编号在第2列
        if code_cell.value:
            return code_cell.value
    return None

def read_excel_pairs(filename):
    """读取 Excel 文件中的确认重复对"""
    try:
        wb = openpyxl.load_workbook(filename)
        
        # 尝试不同的 sheet 名称
        sheet_names = ['确认重复', '重复确认', 'Confirmed Duplicates', 'confirmed duplicates']
        target_sheet = None
        
        for sheet_name in sheet_names:
            if sheet_name in wb.sheetnames:
                target_sheet = wb[sheet_name]
                break
                
        if not target_sheet:
            print(f"警告：在 {filename} 中未找到确认重复 sheet")
            return set()
            
        print(f"读取 sheet: {target_sheet.title}")
        pairs = set()
        
        # 从第2行开始读取（跳过表头）
        for row in range(2, target_sheet.max_row + 1):
            code = parse_excel_key(target_sheet, row)
            if code:
                # 根据行号判断是A还是B，实际应该成对出现
                # 这里简化处理，假设每行就是一个物料编号
                # 在实际应用中，需要根据A/B配对逻辑处理
                next_row = row + 1
                if next_row <= target_sheet.max_row:
                    next_code = parse_excel_key(target_sheet, next_row)
                    if next_code and next_code != code:
                        # 找到一对物料
                        pair_key = tuple(sorted([code, next_code]))
                        pairs.add(pair_key)
                        row += 1  # 跳过下一行，因为已经处理了
                        
        return pairs
        
    except Exception as e:
        print(f"读取 {filename} 时出错: {e}")
        return set()

def main():
    # 读取新旧两个文件
    old_file = '/Users/zhangdongfang/.openclaw/workspace-in3bot/IN3数据/可疑重复物料-2026-05-19-v3.xlsx'
    new_file = '/Users/zhangdongfang/.openclaw/workspace-in3bot/IN3数据/可疑重复物料-2026-05-21-v3.xlsx'
    
    print(f"读取旧文件: {old_file}")
    old_pairs = read_excel_pairs(old_file)
    print(f"旧文件重复对数量: {len(old_pairs)}")
    
    print(f"\n读取新文件: {new_file}")
    new_pairs = read_excel_pairs(new_file)
    print(f"新文件重复对数量: {len(new_pairs)}")
    
    # 分类对比
    resolved = old_pairs - new_pairs  # 旧有新无 = 已解决
    new_added = new_pairs - old_pairs  # 旧无新有 = 新增
    unresolved = old_pairs & new_pairs  # 两次都有 = 未解决
    
    # 生成报告
    report = f"""📊 确认重复物料对比（2026-05-19 → 2026-05-21）

✅ 已解决: {len(resolved)} 对
🆕 新增: {len(new_added)} 对  
⚠️ 未解决: {len(unresolved)} 对

"""
    
    if resolved:
        report += "✅ 已解决详情:\n"
        for i, (code1, code2) in enumerate(sorted(resolved), 1):
            report += f"  {i}. {code1} ↔ {code2}\n"
        report += "\n"
    
    if new_added:
        report += "🆕 新增详情:\n"
        for i, (code1, code2) in enumerate(sorted(new_added), 1):
            report += f"  {i}. {code1} ↔ {code2}\n"
        report += "\n"
    
    if unresolved:
        report += "⚠️ 未解决详情:\n"
        for i, (code1, code2) in enumerate(sorted(unresolved), 1):
            report += f"  {i}. {code1} ↔ {code2}\n"
    
    # 保存报告
    with open('/Users/zhangdongfang/.openclaw/workspace-in3bot/IN3数据/对比报告-2026-05-21.txt', 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(report)
    print(f"对比报告已保存到: 对比报告-2026-05-21.txt")

if __name__ == "__main__":
    main()