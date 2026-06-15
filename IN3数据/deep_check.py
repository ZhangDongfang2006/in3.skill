#!/usr/bin/env python3
"""IN3 物料重复检查 - AI 深度查重脚本"""
import json, re, sys

INPUT = "/Users/zhangdongfang/.openclaw/workspace-in3bot/IN3数据/ai_final_unsure_0601.json"
OUTPUT = "/Users/zhangdongfang/.openclaw/workspace-in3bot/IN3数据/AI深度查重结果-20260601.xlsx"

def norm(s):
    if not s: return ""
    s = s.upper()
    s = re.sub(r'\s+', '', s)
    s = s.replace('（','(').replace('）',')').replace('，',',').replace('：',':').replace('；',';')
    return s

def deep_norm(s):
    """去空格+去分隔符+去括号"""
    s = norm(s)
    s = re.sub(r'[/\\\-_\(\)（）]+', '', s)
    return s

def compare(desc_a, desc_b, mfr_a, mfr_b):
    """返回 (result, reason): result in ('dup','nondup','unsure')"""
    
    na = norm(desc_a)
    nb = norm(desc_b)
    a = desc_a.upper().strip()
    b = desc_b.upper().strip()
    
    # 1. 完全相同
    if na == nb:
        return ("dup", "仅格式/空格差异")
    if deep_norm(a) == deep_norm(b):
        return ("dup", "仅格式/分隔符差异")
    if na.rstrip('/') == nb.rstrip('/'):
        return ("dup", "仅末尾分隔符差异")
    
    # === 制造商不同 ===
    mfrs = {}
    for label, m in [('A', mfr_a), ('B', mfr_b)]:
        if m:
            mfrs[label] = m.replace('有限公司','').replace('(中国)','').replace('（中国）','').replace('（','(').replace('）',')').strip()
    if len(mfrs) == 2:
        m1, m2 = list(mfrs.values())
        if m1 != m2 and not (m1 in m2 or m2 in m1):
            return ("nondup", f"制造商不同: {mfr_a} vs {mfr_b}")
    
    ca = na  # normalized
    cb = nb
    
    # === 极数差异 ===
    def find_poles(s):
        # 3P, 4P, 3P3D, 4P3D, 1P+N, 2P
        p = re.findall(r'(?<!\w)(\d)P(?![A-Z])', s)
        return set(p)
    pa, pb = find_poles(ca), find_poles(cb)
    if pa and pb and pa != pb:
        return ("nondup", f"极数不同: {','.join(pa)}P vs {','.join(pb)}P")
    
    # === 分断能力/性能等级 ===
    def perf_diff(s1, s2):
        # NSX100F vs NSX100N etc
        for prefix in [r'NSX\d*', r'TMXE?-?\d*', r'NDM\d*-?\d*', r'CVS\d*', r'NXM-?W?\d*']:
            m1 = re.search(prefix + r'(M?[FHNSECL]|HPN|NA)', s1)
            m2 = re.search(prefix + r'(M?[FHNSECL]|HPN|NA)', s2)
            if m1 and m2 and m1.group(1) != m2.group(1):
                # mF vs F: also different
                v1, v2 = m1.group(1), m2.group(1)
                # Normalize: MF==mF, etc
                if v1.replace('M','m') != v2.replace('M','m'):
                    return (v1, v2)
            # ABB XT series
        for prefix in [r'XT\d*']:
            m1 = re.search(prefix + r'([NSHL])\d+', s1)
            m2 = re.search(prefix + r'([NSHL])\d+', s2)
            if m1 and m2 and m1.group(1) != m2.group(1):
                return (m1.group(1), m2.group(1))
        return None
    
    pd = perf_diff(ca, cb)
    if pd:
        return ("nondup", f"性能等级不同: {pd[0]} vs {pd[1]}")
    
    # === 脱扣曲线 C vs D ===
    def get_curve(s):
        # C20, D63, C/4P
        return re.findall(r'(?<![A-Z])([CD])(\d+)', s)
    ca_c = get_curve(ca)
    cb_c = get_curve(cb)
    if ca_c and cb_c:
        for cv_a, cur_a in ca_c:
            for cv_b, cur_b in cb_c:
                if cur_a == cur_b and cv_a != cv_b:
                    return ("nondup", f"脱扣曲线不同: {cv_a}{cur_a} vs {cv_b}{cur_b}")
    
    # === 脱扣单元 ===
    def get_trip(s):
        units = set()
        if 'TAM' in s: units.add('TAM')
        if 'TMD' in s: units.add('TMD')
        if 'TMF' in s: units.add('TMF')
        # MA as trip unit (not part of model name like "MA3P")
        if re.search(r'\bMA\b', s) or re.search(r'MA\s', s): units.add('MA')
        if 'EKIPDIPLSIG' in s or 'EKIPDIPLIG' in s: units.add('EKIPDIPLIG')
        elif 'EKIPDIPLSI' in s: units.add('EKIPDIPLSI')
        elif 'EKIPDIP' in s: units.add('EKIPDIP')
        return units
    ta, tb = get_trip(ca), get_trip(cb)
    if ta and tb and ta != tb:
        only_a = ta - tb
        only_b = tb - ta
        if only_a or only_b:
            return ("nondup", f"脱扣单元不同: {','.join(ta)} vs {','.join(tb)}")
    
    # === 附件差异 ===
    def get_acc(s):
        accs = set()
        if re.search(r'AUX', s): accs.add('AUX')
        if '辅助' in s.lower(): accs.add('辅助')
        if '报警' in s.lower(): accs.add('报警')
        if '分励' in s.lower(): accs.add('分励')
        if re.search(r'\bSOR\b', s): accs.add('SOR')
        if re.search(r'\bMX\b', s) and 'TMX' not in s.upper(): accs.add('MX')
        if re.search(r'\bOF\b', s): accs.add('OF')
        if re.search(r'1SY', s): accs.add('1SY')
        if re.search(r'1Q', s): accs.add('1Q')
        if '锁定' in s.lower(): accs.add('锁定')
        if '挂锁' in s.lower(): accs.add('挂锁')
        if '钥匙锁' in s.lower(): accs.add('钥匙锁')
        if '接线片' in s.lower(): accs.add('接线片')
        if '相间隔板' in s.lower(): accs.add('相间隔板')
        return accs
    aa, ab = get_acc(ca), get_acc(cb)
    if aa != ab:
        diff = (aa | ab) - (aa & ab)
        if diff:
            return ("nondup", f"附件差异: {','.join(sorted(diff))}")
    
    # === 安装方式 ===
    def get_mount(s):
        t = set()
        if '固定式' in s: t.add('固定式')
        if '插入式' in s: t.add('插入式')
        if '抽屉式' in s: t.add('抽屉式')
        return t
    ma_m, mb_m = get_mount(a), get_mount(b)
    if ma_m and mb_m and ma_m != mb_m:
        return ("nondup", f"安装方式不同: {','.join(ma_m)} vs {','.join(mb_m)}")
    
    # === 接线方式 ===
    def get_wire(s):
        t = set()
        if '板前' in s: t.add('板前')
        if '板后' in s: t.add('板后')
        return t
    wa, wb = get_wire(a), get_wire(b)
    if wa and wb and wa != wb:
        return ("nondup", f"接线方式不同: {','.join(wa)} vs {','.join(wb)}")
    
    # === 漏电类型 ===
    def get_leak(s):
        t = set()
        if re.search(r'AC型', s): t.add('AC型')
        if re.search(r'(?<!\w)A型', s) and '漏电' in s: t.add('A型')
        if re.search(r'(?<!\w)B型', s) and '漏电' in s: t.add('B型')
        return t
    la_l, lb_l = get_leak(a), get_leak(b)
    if la_l != lb_l and (la_l or lb_l):
        return ("nondup", f"漏电类型不同: {','.join(la_l) if la_l else '无'} vs {','.join(lb_l) if lb_l else '无'}")
    
    # === (R) 后缀 ===
    if bool(re.search(r'\(R\)', a)) != bool(re.search(r'\(R\)', b)):
        return ("nondup", "(R)后缀差异")
    
    # === NA vs 断路器 ===
    has_na_a = bool(re.search(r'\dNA', ca))
    has_na_b = bool(re.search(r'\dNA', cb))
    if has_na_a != has_na_b:
        return ("nondup", "NA(隔离开关) vs 非NA(断路器)")
    
    # === 安装附件 PHR vs FF vs PMP ===
    def get_install(s):
        for x in ['PHR', 'PMP']:
            if x in s: return x
        if 'FF' in s and re.search(r'\bFF\b', s): return 'FF'
        if re.search(r'\bF\b', s) and 'NSX' in s: 
            # F could be perf level or install accessory
            return None  # ambiguous
        return None
    ia, ib = get_install(ca), get_install(cb)
    if ia and ib and ia != ib:
        return ("nondup", f"安装附件不同: {ia} vs {ib}")
    
    # === 数字差异分析 ===
    # 用token比较
    dna = deep_norm(a)
    dnb = deep_norm(b)
    
    # SequenceMatcher similarity
    from difflib import SequenceMatcher
    ratio = SequenceMatcher(None, dna, dnb).ratio()
    
    if ratio >= 0.95:
        # Very similar - check specific diffs
        sm = SequenceMatcher(None, dna, dnb)
        has_num_diff = False
        num_diffs = []
        for op, i1, i2, j1, j2 in sm.get_opcodes():
            if op != 'equal':
                sa = dna[i1:i2]
                sb = dnb[j1:j2]
                if re.search(r'\d', sa) or re.search(r'\d', sb):
                    has_num_diff = True
                    num_diffs.append(f"'{sa}' vs '{sb}'")
        
        if has_num_diff:
            return ("nondup", f"数字差异: {'; '.join(num_diffs)}")
        else:
            return ("dup", "仅表述/格式差异，无参数差异")
    
    # Check if one is substring of other (without special chars)
    if dna in dnb or dnb in dna:
        longer = dnb if dna in dnb else dna
        shorter = dna if dna in dnb else dnb
        extra = longer.replace(shorter, '', 1)
        if re.search(r'\d', extra):
            # Extra part has numbers - likely different
            return ("nondup", f"参数差异: 额外部分含数字 '{extra}'")
        else:
            return ("dup", f"描述为子集关系，额外部分无数字")
    
    # Token-based comparison for remaining
    tokens_a = set(re.findall(r'[A-Z]+\d+|[A-Z]+|\d+', ca))
    tokens_b = set(re.findall(r'[A-Z]+\d+|[A-Z]+|\d+', cb))
    
    only_a = tokens_a - tokens_b
    only_b = tokens_b - tokens_a
    
    if only_a or only_b:
        # Check if exclusive tokens contain numbers
        num_a = [t for t in only_a if re.match(r'^\d+$', t)]
        num_b = [t for t in only_b if re.match(r'^\d+$', t)]
        
        if num_a and num_b:
            return ("nondup", f"数字参数差异: A独有{','.join(num_a)}, B独有{','.join(num_b)}")
        
        if num_a:
            return ("nondup", f"数字参数差异: A独有{','.join(num_a)}")
        if num_b:
            return ("nondup", f"数字参数差异: B独有{','.join(num_b)}")
        
        # Letter-only differences
        alpha_a = [t for t in only_a if re.match(r'^[A-Z]+$', t)]
        alpha_b = [t for t in only_b if re.match(r'^[A-Z]+$', t)]
        
        perf_keywords = {'F','H','N','S','L','M','U','NA','HPN','E','S','C'}
        for t in alpha_a:
            if t in perf_keywords:
                return ("nondup", f"性能标识差异: A有{t}而B无")
        for t in alpha_b:
            if t in perf_keywords:
                return ("nondup", f"性能标识差异: B有{t}而A无")
    
    # Remaining: unsure
    return ("unsure", "需人工确认")


def process():
    with open(INPUT, 'r', encoding='utf-8') as f:
        data = json.load(f)
    print(f"共 {len(data)} 对")
    
    dup, nondup, unsure = [], [], []
    
    for i, p in enumerate(data):
        a, b = p['a'], p['b']
        result, reason = compare(a['desc'], b['desc'], a.get('mfr',''), b.get('mfr',''))
        
        entry = {
            'a_code': a['code'], 'a_name': a.get('name',''), 'a_desc': a['desc'],
            'a_mfr': a.get('mfr',''), 'a_cat': a.get('cat',''), 'a_subcat': a.get('subcat',''),
            'b_code': b['code'], 'b_name': b.get('name',''), 'b_desc': b['desc'],
            'b_mfr': b.get('mfr',''), 'b_cat': b.get('cat',''), 'b_subcat': b.get('subcat',''),
            'reason': reason
        }
        
        if result == 'dup': dup.append(entry)
        elif result == 'nondup': nondup.append(entry)
        else: unsure.append(entry)
        
        if (i+1) % 2000 == 0:
            print(f"  {i+1}/{len(data)} done...")
    
    print(f"\n确认重复: {len(dup)}")
    print(f"非重复:   {len(nondup)}")
    print(f"待确认:   {len(unsure)}")
    
    # Stats
    print("\n非重复原因分布:")
    from collections import Counter
    reason_counts = Counter()
    for e in nondup:
        key = e['reason'].split(':')[0] if ':' in e['reason'] else e['reason'][:15]
        reason_counts[key] += 1
    for k, v in reason_counts.most_common(20):
        print(f"  {k}: {v}")
    
    return dup, nondup, unsure


def make_excel(dup_list, unsure_list):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    wb = Workbook()
    
    blue = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
    orange = PatternFill(start_color='FDEBD0', end_color='FDEBD0', fill_type='solid')
    red_ft = Font(color='FF0000', bold=True)
    hdr_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    hdr_ft = Font(color='FFFFFF', bold=True)
    bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    center = Alignment(horizontal='center', vertical='center')
    wrap = Alignment(wrap_text=True, vertical='center')
    
    headers = ['序号','A物料编码','A名称','A描述','A制造商','B物料编码','B名称','B描述','B制造商','判断','差异原因']
    widths = [6, 15, 12, 55, 22, 15, 12, 55, 22, 10, 30]
    
    def write_sheet(ws, entries, judgment):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill; c.font = hdr_ft; c.border = bdr; c.alignment = center
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+ci)].width = w
        
        for idx, e in enumerate(entries):
            r = idx + 2
            fill = blue if idx % 2 == 0 else orange
            vals = [idx+1, e['a_code'], e['a_name'], e['a_desc'], e['a_mfr'],
                    e['b_code'], e['b_name'], e['b_desc'], e['b_mfr'],
                    judgment, e['reason']]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=ci, value=v)
                c.fill = fill; c.border = bdr
                c.alignment = wrap if ci in [4,8,11] else center
            
            # Diff fields in red
            diff_cols = set()
            if e['a_desc'] != e['b_desc']: diff_cols |= {4, 8}
            if e['a_mfr'] != e['b_mfr']: diff_cols |= {5, 9}
            if e['a_name'] != e['b_name']: diff_cols |= {3, 7}
            for ci in diff_cols:
                ws.cell(row=r, column=ci).font = red_ft
    
    if dup_list:
        ws1 = wb.active
        ws1.title = "确认重复"
        write_sheet(ws1, dup_list, "确认重复")
    
    if unsure_list:
        ws2 = wb.create_sheet("待人工确认")
        write_sheet(ws2, unsure_list, "待确认")
    elif not dup_list:
        # Create empty sheet
        wb.active.title = "无数据"
    
    wb.save(OUTPUT)
    print(f"\nExcel saved: {OUTPUT}")


if __name__ == '__main__':
    dup, nondup, unsure = process()
    make_excel(dup, unsure)
    # Also dump nondup to json for reference
    nondup_out = OUTPUT.replace('.xlsx', '_nondup.json')
    with open(nondup_out, 'w', encoding='utf-8') as f:
        json.dump(nondup, f, ensure_ascii=False, indent=2)
    print(f"Non-dup detail saved: {nondup_out}")
