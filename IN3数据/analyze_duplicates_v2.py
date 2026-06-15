#!/usr/bin/env python3
"""IN3 物料重复检查分析脚本 v2 - 更严格的匹配"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from difflib import SequenceMatcher
import re
import time

INPUT_FILE = '物料主数据导出结果-20260512.xlsx'
OUTPUT_FILE = '可疑重复物料-2026-05-12.xlsx'

# Column mapping (1-indexed)
COL = {
    'id': 1, 'code': 2, 'name': 6, 'desc': 7,
    'cat_code': 8, 'cat': 9, 'subcat_code': 10, 'subcat': 11,
    'manufacturer': 13, 'source': 14, 'lead_time': 18,
    'unit': 22, 'price': 61,
    'creator': 70, 'create_date': 72,
    'modifier': 73, 'modify_date': 75,
    'color': 16
}

print("Loading workbook...")
start = time.time()
wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb['物料主数据']
print(f"Loaded in {time.time()-start:.1f}s, {ws.max_row} rows")

# Read all materials
materials = []
exclude_cats = {'成品柜', '外购成套'}

for row_idx in range(2, ws.max_row + 1):
    cat = ws.cell(row=row_idx, column=COL['cat']).value or ''
    subcat = ws.cell(row=row_idx, column=COL['subcat']).value or ''
    
    if cat in exclude_cats:
        continue
    
    def get_val(r, c):
        v = ws.cell(row=r, column=c).value
        return str(v).strip() if v is not None else ''
    
    materials.append({
        'row': row_idx,
        'code': get_val(row_idx, COL['code']),
        'name': get_val(row_idx, COL['name']),
        'desc': get_val(row_idx, COL['desc']),
        'cat': str(cat).strip(),
        'subcat': str(subcat).strip(),
        'manufacturer': get_val(row_idx, COL['manufacturer']),
        'source': get_val(row_idx, COL['source']),
        'unit': get_val(row_idx, COL['unit']),
        'lead_time': get_val(row_idx, COL['lead_time']),
        'price': get_val(row_idx, COL['price']),
        'creator': get_val(row_idx, COL['creator']),
        'create_date': get_val(row_idx, COL['create_date']),
        'modifier': get_val(row_idx, COL['modifier']),
        'modify_date': get_val(row_idx, COL['modify_date']),
        'color': get_val(row_idx, COL['color']),
    })

print(f"Total materials (excl. 成品柜/外购成套): {len(materials)}")

def normalize(s):
    """Normalize for comparison"""
    s = s.lower().strip()
    s = re.sub(r'\s+', ' ', s)
    s = s.replace('（', '(').replace('）', ')')
    s = s.replace('，', ',').replace('：', ':')
    s = s.replace('×', 'x').replace('—', '-')
    return s

def ultra_normalize(s):
    """Remove all formatting, keep only alphanumeric"""
    s = normalize(s)
    s = re.sub(r'[\s\-_/\\(),;:.，：；、。{}\[\]]', '', s)
    return s

def desc_similarity(a, b):
    """Calculate description similarity ratio"""
    return SequenceMatcher(None, normalize(a), normalize(b)).ratio()

# Group by (cat, subcat, name) first for efficiency
# Within each group, compare descriptions
groups = defaultdict(list)
for m in materials:
    # Group by normalized name within category
    key = (m['cat'], m['subcat'], normalize(m['name']))
    groups[key].append(m)

print(f"Groups by (cat, subcat, name): {len(groups)}")

# Find suspicious pairs: same category/subcategory/name, descriptions very similar
suspicious = []

for key, group in groups.items():
    if len(group) < 2:
        continue
    
    for i in range(len(group)):
        for j in range(i+1, len(group)):
            a, b = group[i], group[j]
            
            # Same description -> definitely suspicious
            if a['desc'] == b['desc']:
                if a['code'] != b['code']:  # Different material codes
                    suspicious.append((a, b))
                continue
            
            # Very similar descriptions (ratio >= 0.85)
            sim = desc_similarity(a['desc'], b['desc'])
            if sim >= 0.85:
                suspicious.append((a, b))

# Also check cross-name groups where names are very similar
# Group by (cat, subcat) only
cat_groups = defaultdict(list)
for m in materials:
    key = (m['cat'], m['subcat'])
    cat_groups[key].append(m)

# Build description index within each (cat, subcat) group
for key, group in cat_groups.items():
    # Group by normalized description
    desc_groups = defaultdict(list)
    for m in group:
        nd = ultra_normalize(m['desc'])
        desc_groups[nd].append(m)
    
    # Find materials with identical normalized descriptions but different names
    for nd, items in desc_groups.items():
        if len(items) < 2:
            continue
        for i in range(len(items)):
            for j in range(i+1, len(items)):
                a, b = items[i], items[j]
                if a['code'] == b['code']:
                    continue
                pair_key = tuple(sorted([a['code'], b['code']]))
                # Add if not already in suspicious
                suspicious.append((a, b))

# Deduplicate
seen = set()
unique_pairs = []
for a, b in suspicious:
    pk = tuple(sorted([a['code'], b['code']]))
    if pk not in seen:
        seen.add(pk)
        unique_pairs.append((a, b))

print(f"Unique suspicious pairs: {len(unique_pairs)}")

# Exclusion rules
def apply_exclusion_rules(a, b):
    """Apply 27 exclusion rules. Returns (is_excluded, reason)"""
    desc_a = a['desc']
    desc_b = b['desc']
    name_a = a['name']
    name_b = b['name']
    mfg_a = a['manufacturer']
    mfg_b = b['manufacturer']
    
    # Rule 1: 甲供件
    for m in [a, b]:
        if '甲供' in m['cat'] or '甲供' in m['subcat']:
            return True, '甲供件不算重复'
    
    # Rule 2: One has manufacturer, one doesn't
    if (mfg_a and not mfg_b) or (not mfg_a and mfg_b):
        return True, '一个有制造商一个没有'
    
    # Rule 3: Manufacturers completely different
    if mfg_a and mfg_b and mfg_a != mfg_b:
        aliases = [
            ('宁波三爱', '三爱'), ('海越电气', '海越湖北'),
            ('天灵', '宁波天灵'), ('长沙威胜', '威胜'),
            ('上海良信电器股份有限公司', '良信'),
            ('上海良信电器', '良信'),
        ]
        is_alias = False
        for x, y in aliases:
            if (x in mfg_a and y in mfg_b) or (y in mfg_a and x in mfg_b):
                is_alias = True
                break
        if not is_alias:
            return True, f'制造商不同: {mfg_a} vs {mfg_b}'
    
    # Rule 4: Direction differences
    dir_pairs = [
        ('左操', '右操'), ('平进平出', '平进侧出'), ('上进', '下进'),
        ('上进下出', '下进上出'), ('上进下出', '上出'), ('立式', '卧式'),
        ('侧进', '上进'), ('侧出', '下出'),
    ]
    for d1, d2 in dir_pairs:
        if (d1 in desc_a and d2 in desc_b) or (d2 in desc_a and d1 in desc_b):
            return True, f'方向不同: {d1} vs {d2}'
    
    # Rule 5: Breaker frame letter suffix
    frame_pattern = r'(?:CDM|NDM|CM|NM)\d*-?\d+([FCDHLMNS])'
    fa = re.search(frame_pattern, desc_a, re.IGNORECASE)
    fb = re.search(frame_pattern, desc_b, re.IGNORECASE)
    if fa and fb and fa.group(1).upper() != fb.group(1).upper():
        return True, f'断路器分断能力等级不同: {fa.group(1)} vs {fb.group(1)}'
    
    # Rule 6: MCB trip curve B/C/D
    # Look for patterns like -C63, -B16, /D32
    curve_a = re.search(r'[-/]\s*([BCD])\s*\d+', desc_a)
    curve_b = re.search(r'[-/]\s*([BCD])\s*\d+', desc_b)
    if curve_a and curve_b and curve_a.group(1) != curve_b.group(1):
        return True, f'脱扣曲线不同: {curve_a.group(1)} vs {curve_b.group(1)}'
    
    # Rule 7: Pole count
    poles_a = set(re.findall(r'(\d)[Pp]', desc_a))
    poles_b = set(re.findall(r'(\d)[Pp]', desc_b))
    if poles_a and poles_b and poles_a != poles_b:
        return True, f'极数不同: {poles_a} vs {poles_b}'
    
    # Rule 8: Color
    if a['color'] and b['color'] and a['color'] != b['color']:
        return True, f'颜色不同: {a["color"]} vs {b["color"]}'
    
    # Rule 9: Leakage type
    if ('A型' in desc_a and 'AC型' in desc_b) or ('AC型' in desc_a and 'A型' in desc_b):
        return True, '漏电保护类型不同'
    
    # Rule 10: Attachment differences
    attachments = ['失压', '分励', '合闸', '辅助', '报警', '门框', '电磁锁',
                   'RS485', '通讯', '温度指示', '带电显示', '加热器']
    att_a = set(x for x in attachments if x in desc_a)
    att_b = set(x for x in attachments if x in desc_b)
    diff = att_a ^ att_b
    if diff:
        return True, f'附件不同: {", ".join(sorted(diff))}'
    
    # Rule 12: Wire type
    wire_types = ['BV', 'BVR', 'YJV', 'ZR-YJV', 'RVV', 'RV', 'RVB', 'NH-YJV']
    w_a = set(w for w in wire_types if w in desc_a)
    w_b = set(w for w in wire_types if w in desc_b)
    if w_a and w_b and w_a != w_b:
        return True, f'电线类型不同: {w_a} vs {w_b}'
    
    # Rule 14: Thread direction
    if ('正牙' in desc_a and '反牙' in desc_b) or ('反牙' in desc_a and '正牙' in desc_b):
        return True, '螺纹方向不同'
    
    # Rule 16: Trip method TMD vs MA
    if ('TMD' in desc_a.upper() and 'MA' in desc_b.upper()) or \
       ('MA' in desc_a.upper() and 'TMD' in desc_b.upper()):
        return True, '脱扣方式不同(TMD vs MA)'
    
    # Rule 17: Trip unit LSI vs TMA
    if ('LSI' in desc_a and 'TMA' in desc_b) or ('TMA' in desc_a and 'LSI' in desc_b):
        return True, '脱扣单元不同(LSI vs TMA)'
    
    # Rule 19: TBP-A vs TBP-B
    tbp_a = re.search(r'TBP[-]?([A-Z])', desc_a)
    tbp_b = re.search(r'TBP[-]?([A-Z])', desc_b)
    if tbp_a and tbp_b and tbp_a.group(1) != tbp_b.group(1):
        return True, f'过电压保护器型号不同: TBP-{tbp_a.group(1)} vs TBP-{tbp_b.group(1)}'
    
    # Rule 20: Model series
    series_pairs = [('CM3E', 'CM3'), ('CM1E', 'CM1'), ('NDM3', 'CDM3')]
    for s1, s2 in series_pairs:
        if (s1 in desc_a and s2 in desc_b) or (s2 in desc_a and s1 in desc_b):
            return True, f'型号系列不同: {s1} vs {s2}'
    
    # Rule 21: RV vs BVR
    if ('RV' in desc_a and 'BVR' in desc_b) or ('BVR' in desc_a and 'RV' in desc_b):
        return True, 'RV vs BVR 电线类型不同'
    
    # Rule 22: Material differences
    mat_pairs = [('铜', '铝'), ('紫铜', '黄铜'), ('304', '316'), ('全铜', '全铝')]
    for m1, m2 in mat_pairs:
        if (m1 in desc_a and m2 in desc_b) or (m2 in desc_a and m1 in desc_b):
            return True, f'材质不同: {m1} vs {m2}'
    
    # Rule 24: Current/voltage differences - extract primary numbers
    # Extract patterns like 63A, 100A, 250A, 32A etc from description
    curr_pattern = r'(\d+(?:\.\d+)?)\s*[Aa](?:\s|$|[/\-+,)])'
    currs_a = re.findall(curr_pattern, desc_a)
    currs_b = re.findall(curr_pattern, desc_b)
    if currs_a and currs_b and currs_a != currs_b:
        # Check if first current is clearly different
        try:
            ca = float(currs_a[0])
            cb = float(currs_b[0])
            if abs(ca - cb) > 1:
                ratio = max(ca, cb) / min(ca, cb) if min(ca, cb) > 0 else 999
                if ratio > 1.05:
                    return True, f'额定电流不同: {currs_a[0]}A vs {currs_b[0]}A'
        except:
            pass
    
    # Rule 27: EPS power
    eps_a = re.search(r'EPS\s*(\d+(?:\.\d+)?)\s*KW?', desc_a, re.IGNORECASE)
    eps_b = re.search(r'EPS\s*(\d+(?:\.\d+)?)\s*KW?', desc_b, re.IGNORECASE)
    if eps_a and eps_b and eps_a.group(1) != eps_b.group(1):
        return True, f'EPS功率不同: {eps_a.group(1)}KW vs {eps_b.group(1)}KW'
    
    # Rule 23/25: If descriptions have model numbers, check if they differ
    # Extract main model number from description
    # This is a general catch-all: if the numeric parts of descriptions differ significantly
    
    return False, ''


def classify_pair(a, b):
    """Classify a pair as confirmed duplicate, non-duplicate, or needs review"""
    # First check exclusion rules
    excluded, reason = apply_exclusion_rules(a, b)
    if excluded:
        return 'non_dup', reason
    
    # Check if confirmed duplicate
    desc_a = normalize(a['desc'])
    desc_b = normalize(b['desc'])
    
    if desc_a == desc_b:
        return 'confirmed', '描述完全相同(归一化后)'
    
    # Check ultra-normalized
    ua = ultra_normalize(a['desc'])
    ub = ultra_normalize(b['desc'])
    if ua == ub:
        return 'confirmed', '描述仅格式差异'
    
    # 1N vs 1P+N
    da = desc_a.replace('1n', '1p+n')
    db = desc_b.replace('1n', '1p+n')
    if da == db:
        return 'confirmed', '1N vs 1P+N 同一产品'
    
    # If still here, needs manual review
    # Calculate similarity for the remark
    sim = desc_similarity(a['desc'], b['desc'])
    return 'review', f'相似度{sim:.1%}，需人工确认'


# Classify all pairs
confirmed = []
non_dup = []
review = []

for a, b in unique_pairs:
    cat, reason = classify_pair(a, b)
    if cat == 'confirmed':
        confirmed.append((a, b, reason))
    elif cat == 'non_dup':
        non_dup.append((a, b, reason))
    else:
        review.append((a, b, reason))

print(f"\nResults:")
print(f"  Confirmed duplicates: {len(confirmed)}")
print(f"  Non-duplicates: {len(non_dup)}")
print(f"  Needs review: {len(review)}")
print(f"  Total: {len(confirmed) + len(non_dup) + len(review)}")

# Generate Excel
print("\nGenerating Excel...")

out_wb = openpyxl.Workbook()

blue_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
orange_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
red_font = Font(color='FF0000')
normal_font = Font()
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_w = Font(bold=True, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

EXCEL_COLS = [
    ('序号', 'label'),
    ('物料编号', 'code'),
    ('物料名称', 'name'),
    ('物料描述', 'desc'),
    ('物料类别', 'cat'),
    ('物料子类别', 'subcat'),
    ('制造商', 'manufacturer'),
    ('物料来源', 'source'),
    ('主计量单位', 'unit'),
    ('提前期', 'lead_time'),
    ('标准价格', 'price'),
    ('创建人', 'creator'),
    ('创建日期', 'create_date'),
    ('最近修改人', 'modifier'),
    ('最近修改日期', 'modify_date'),
    ('备注', 'remark'),
]

DIFF_COLS = {'name', 'desc', 'cat', 'subcat', 'manufacturer', 'source', 'unit', 'lead_time', 'price'}

def write_sheet(ws, data):
    # Headers
    for col_idx, (col_name, _) in enumerate(EXCEL_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font_w
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_idx = 2
    for pair_num, (a, b, reason) in enumerate(data, 1):
        for is_b, mat in enumerate([a, b]):
            fill = orange_fill if is_b else blue_fill
            label = f"{'B' if is_b else 'A'}-{pair_num}"
            
            for col_idx, (col_name, col_key) in enumerate(EXCEL_COLS, 1):
                if col_key == 'label':
                    value = label
                elif col_key == 'remark':
                    value = reason
                else:
                    value = mat.get(col_key, '')
                    if isinstance(value, float) and value == int(value):
                        value = int(value)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=(col_key == 'desc'))
                
                if col_key in DIFF_COLS:
                    val_a = str(a.get(col_key, '') or '').strip()
                    val_b = str(b.get(col_key, '') or '').strip()
                    cell.font = red_font if val_a != val_b else normal_font
                else:
                    cell.font = normal_font
            
            row_idx += 1
    
    # Column widths
    widths = [8, 16, 18, 60, 12, 14, 20, 10, 10, 8, 12, 10, 14, 12, 14, 45]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w
    
    ws.freeze_panes = 'C2'
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLS))}{row_idx - 1}"


# Sheet 1: Confirmed duplicates
ws1 = out_wb.active
ws1.title = '确认重复'
write_sheet(ws1, confirmed)

# Sheet 2: Non-duplicates
ws2 = out_wb.create_sheet('非重复')
write_sheet(ws2, non_dup)

# Sheet 3: Needs review
ws3 = out_wb.create_sheet('待人工确认')
write_sheet(ws3, review)

out_wb.save(OUTPUT_FILE)
print(f"Excel saved to: {OUTPUT_FILE}")
print(f"Total time: {time.time()-start:.1f}s")
