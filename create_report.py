#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成今日采购价格分析报告
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os

def create_report():
    wb = Workbook()
    ws = wb.active
    ws.title = "采购价格分析报告"
    
    # 设置标题
    ws.merge_cells('A1:D1')
    ws['A1'] = f"每日采购价格分析报告 - {datetime.now().strftime('%Y-%m-%d')}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 内容
    ws['A3'] = "检查日期:"
    ws['B3'] = datetime.now().strftime('%Y-%m-%d')
    
    ws['A4'] = "分析结果:"
    ws['B4'] = "2026-07-07 无需分析的采购记录"
    
    ws['A5'] = "原因说明:"
    ws['B5'] = "1. 今日(2026-07-07)无新的采购订单"
    ws['B6'] = "2. 所有采购订单的历史数据已分析完成"
    ws['B7'] = "3. 铜价: 103400 元/吨 (103.40 元/kg)"
    
    ws['A8'] = "数据文件:"
    ws['B8'] = "宁波: 采购订单明细-宁波-20260704.xlsx (448条记录)"
    ws['C8'] = "湖北: 采购订单明细-湖北-20260704.xlsx (587条记录)"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    
    # 添加边框样式
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if cell.value:
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                    top=Side(style='thin'), bottom=Side(style='thin'))
                cell.border = thin_border
    
    # 保存文件
    report_path = "/Users/zhangdongfang/.openclaw/media/采购价格分析报告_20260707.xlsx"
    wb.save(report_path)
    print(f"报告已生成: {report_path}")
    return report_path

if __name__ == "__main__":
    create_report()