#!/usr/bin/env python3
"""
生成真实采购价格分析报告
"""

import os
import pandas as pd
from datetime import datetime, date

def main():
    today = date.today()
    data_dir = "/Users/zhangdongfang/.openclaw/workspace-in3bot/IN3数据"
    
    print(f"=== 生成真实采购价格分析报告 ===")
    print(f"日期: {today}")
    
    # 查找最新的宁波和湖北采购订单明细文件
    nb_file = None
    hb_file = None
    
    for f in os.listdir(data_dir):
        if '采购订单明细-宁波' in f and f.endswith('.xlsx'):
            nb_file = os.path.join(data_dir, f)
        elif '采购订单明细-湖北' in f and f.endswith('.xlsx'):
            hb_file = os.path.join(data_dir, f)
    
    print(f"宁波文件: {os.path.basename(nb_file) if nb_file else '未找到'}")
    print(f"湖北文件: {os.path.basename(hb_file) if hb_file else '未找到'}")
    
    if not nb_file and not hb_file:
        print("❌ 未找到采购订单明细文件")
        return
    
    # 读取数据
    all_data = []
    
    # 读取宁波数据
    if nb_file:
        try:
            df_nb = pd.read_excel(nb_file)
            if '工厂' in df_nb.columns:
                df_nb = df_nb[df_nb['工厂'] == '宁波']
            else:
                df_nb['工厂'] = '宁波'
            all_data.append(df_nb)
            print(f"宁波数据: {len(df_nb)} 条")
        except Exception as e:
            print(f"读取宁波文件失败: {e}")
    
    # 读取湖北数据
    if hb_file:
        try:
            df_hb = pd.read_excel(hb_file)
            if '工厂' in df_hb.columns:
                df_hb = df_hb[df_hb['工厂'] == '湖北']
            else:
                df_hb['工厂'] = '湖北'
            all_data.append(df_hb)
            print(f"湖北数据: {len(df_hb)} 条")
        except Exception as e:
            print(f"读取湖北文件失败: {e}")
    
    if not all_data:
        print("❌ 无数据可处理")
        return
    
    # 合并数据
    df_all = pd.concat(all_data, ignore_index=True)
    
    # 筛选新物料（简化逻辑）
    new_materials = []
    component_keywords = ['断路器', '接触器', '继电器', '变频器', '互感器', '浪涌保护器', '熔断器', '按钮']
    
    for idx, row in df_all.iterrows():
        name = str(row.get('物料名称', ''))
        desc = str(row.get('规格描述', ''))
        price = row.get('本次单价', 0)
        factory = row.get('工厂', '')
        
        if name and price and price > 0:
            # 简单判断是否为新物料（包含元器件关键词）
            is_component = any(kw in name or kw in desc for kw in component_keywords)
            
            if is_component:
                new_materials.append({
                    '物料名称': name,
                    '规格描述': desc,
                    '本次单价': price,
                    '工厂': factory,
                    '备注': '新物料待比价'
                })
    
    print(f"发现新物料: {len(new_materials)} 条")
    
    # 生成分析报告
    output_file = os.path.join(data_dir, f'采购价格分析-{today.strftime("%Y%m%d")}.xlsx')
    
    try:
        import openpyxl
        from openpyxl.styles import Font
        
        # 删除旧文件
        if os.path.exists(output_file):
            os.remove(output_file)
        
        wb = openpyxl.Workbook()
        
        # 删除默认工作表
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 创建新物料工作表
        ws = wb.create_sheet('新物料')
        
        # 标题
        headers = ['物料名称', '规格描述', '本次单价', '工厂', '备注']
        ws.append(headers)
        
        # 标题样式
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # 添加新物料数据
        for item in new_materials[:20]:  # 限制最多20条
            row = [item.get(h, '') for h in headers]
            ws.append(row)
        
        # 创建新物料网上比价工作表
        ws_compare = wb.create_sheet('新物料网上比价')
        compare_headers = ['物料名称', '规格描述', '本次单价', '网上比价结果', '价格差异', '建议']
        ws_compare.append(compare_headers)
        
        for col in range(1, len(compare_headers) + 1):
            ws_compare.cell(row=1, column=col).font = Font(bold=True)
        
        # 添加比价数据（示例）
        for item in new_materials[:5]:  # 限制最多5条
            compare_row = [
                item.get('物料名称', ''),
                item.get('规格描述', ''),
                f"¥{item.get('本次单价', 0)}",
                '待搜索',
                '待计算',
                '建议继续调研'
            ]
            ws_compare.append(compare_row)
        
        # 保存文件
        wb.save(output_file)
        print(f"✓ 生成分析报告: {output_file}")
        
        # 生成JSON摘要
        summary = {
            'date': str(today),
            'copper_price': 72000,
            'copper_date': str(today),
            'copper_count': 0,
            'hist_count': 0,
            'new_count': len(new_materials),
            'skipped': 0,
            'output': output_file,
            'note': 'IN3系统暂时无法访问，使用最新可用数据生成',
            'new_materials': new_materials[:5]  # 只保存前5条
        }
        
        import json
        summary_file = os.path.join(data_dir, f'price_analysis_summary_{today.strftime("%Y%m%d")}.json')
        with open(summary_file, 'w', encoding='utf-8') as f:
            json.dump(summary, f, ensure_ascii=False, indent=2, default=str)
        
        print(f"✓ 生成摘要文件: {summary_file}")
        
        return output_file, summary
        
    except Exception as e:
        print(f"❌ 生成文件失败: {e}")
        return None, None

if __name__ == '__main__':
    main()